#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
SWT 货运管理系统 - 图形界面版本
包含：检测关键字、客户管理、司机管理、公司统计
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pymysql
from dbutils.pooled_db import PooledDB
from datetime import datetime
import os
import json
import re
import calendar
import threading
import ctypes
from ctypes import wintypes


# ── Windows DWM 常量 ──
GWL_STYLE = -16
WS_MAXIMIZEBOX = 0x00010000
WS_MINIMIZEBOX = 0x00020000
WM_NCHITTEST = 0x0084
WM_NCCALCSIZE = 0x0083
WM_NCLBUTTONDOWN = 0x00A1
HTCAPTION = 2
HTMAXBUTTON = 9
HTMINBUTTON = 8
HTCLOSE = 20


class MARGINS(ctypes.Structure):
    _fields_ = [('cxLeftWidth', ctypes.c_int),
                ('cxRightWidth', ctypes.c_int),
                ('cyTopHeight', ctypes.c_int),
                ('cyBottomHeight', ctypes.c_int)]


def enable_acrylic(hwnd):
    """启用窗口磨砂玻璃效果（Windows 10/11）"""
    try:
        if not hwnd:
            return
        # Win11: DWMWA_SYSTEMBACKDROP_TYPE = 38, 3=Acrylic
        val = ctypes.c_int(3)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            hwnd, 38, ctypes.byref(val), ctypes.sizeof(val))
        # 暗色模式: DWMWA_USE_IMMERSIVE_DARK_MODE = 20
        dark = ctypes.c_int(1)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            hwnd, 20, ctypes.byref(dark), ctypes.sizeof(dark))
        # 扩展 DWM 边框到全客户区
        margins = MARGINS(-1, -1, -1, -1)
        ctypes.windll.dwmapi.DwmExtendFrameIntoClientArea(hwnd, ctypes.byref(margins))
    except Exception:
        pass


class AcrylicRoot(tk.Tk):
    """支持磨砂玻璃的自定义标题栏窗口"""

    def __init__(self):
        super().__init__()
        self._is_maximized = False
        self._restore_size = None

        # 隐藏系统标题栏（设置前先保存窗口样式）
        self.withdraw()  # 先隐藏避免闪烁
        self.overrideredirect(True)
        self.attributes('-topmost', False)

        # 恢复最大化/最小化按钮能力
        style = ctypes.windll.user32.GetWindowLongW(self.winfo_id(), GWL_STYLE)
        style |= WS_MAXIMIZEBOX | WS_MINIMIZEBOX
        ctypes.windll.user32.SetWindowLongW(self.winfo_id(), GWL_STYLE, style)

        # DWM: 让窗口圆角 + 磨砂
        self.update_idletasks()
        enable_acrylic(self.winfo_id())

    def _toggle_maximize(self):
        """切换最大化/还原"""
        hwnd = self.winfo_id()
        if self._is_maximized:
            # 还原
            self._is_maximized = False
            if self._restore_size:
                x, y, w, h = self._restore_size
                self.geometry(f"{w}x{h}+{x}+{y}")
            else:
                self.geometry("1400x900")
            # 还原圆角
            try:
                val = ctypes.c_int(1)  # DWMWCP_ROUND
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd, 33, ctypes.byref(val), ctypes.sizeof(val))
            except Exception:
                pass
        else:
            # 最大化前保存当前尺寸
            self._restore_size = (
                self.winfo_x(), self.winfo_y(),
                self.winfo_width(), self.winfo_height())
            self._is_maximized = True
            self.state('zoomed')
            # 最大化时去掉圆角避免黑边
            try:
                val = ctypes.c_int(0)  # DWMWCP_DONOTROUND
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd, 33, ctypes.byref(val), ctypes.sizeof(val))
            except Exception:
                pass


class DeleteConfirmDialog(tk.Toplevel):
    """删除确认对话框 —— 必须输入正确的發票號才能删除"""

    def __init__(self, parent, invoice_code, con_code):
        super().__init__(parent)
        self.title("确认删除")
        self.result = False
        self.invoice_code = invoice_code
        self.geometry("430x240")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry("+{}+{}".format(x, y))

        self._build_ui(invoice_code, con_code)
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

    def _build_ui(self, invoice_code, con_code):
        frame = tk.Frame(self, bg='#f0f2f5')
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        tk.Label(
            frame,
            text="⚠ 此操作不可撤销！请输入發票號确认删除",
            font=("Microsoft YaHei", 10, "bold"),
            bg='#f0f2f5',
            fg="red",
        ).pack(anchor="w", pady=(0, 10))

        info_frame = tk.Frame(frame, bg='#f0f2f5')
        info_frame.pack(fill="x", pady=(0, 10))
        tk.Label(info_frame, text="發票號: " + str(invoice_code),
                 bg='#f0f2f5', fg='#262626').pack(anchor="w")
        tk.Label(info_frame, text="櫃    號: " + str(con_code),
                 bg='#f0f2f5', fg='#262626').pack(anchor="w")

        tk.Label(frame, text="请输入發票號（输入匹配后按钮启用）:",
                 bg='#f0f2f5', fg='#595959').pack(anchor="w")

        input_frame = tk.Frame(frame, bg='#f0f2f5')
        input_frame.pack(fill="x", pady=(5, 15))

        self.input_var = tk.StringVar()
        self.input_var.trace_add("write", self._on_input_change)
        self.entry = tk.Entry(input_frame, textvariable=self.input_var, width=35,
                              font=('Microsoft YaHei UI', 10))
        self.entry.pack(side="left", fill="x", expand=True)
        self.entry.focus_set()

        btn_frame = tk.Frame(frame, bg='#f0f2f5')
        btn_frame.pack(fill="x")

        self.delete_btn = tk.Button(
            btn_frame, text="确认删除", bg='#ff4d4f', fg='#ffffff',
            font=('Microsoft YaHei UI', 10), bd=0, padx=16, pady=5,
            cursor='hand2', state='disabled', command=self._on_confirm
        )
        self.delete_btn.pack(side="left", padx=(0, 10))

        tk.Button(btn_frame, text="取消", bg='#f5f5f5', fg='#595959',
                  font=('Microsoft YaHei UI', 10), bd=0, padx=16, pady=5,
                  cursor='hand2', command=self._on_cancel).pack(side="left")

    def _on_input_change(self, *args):
        if self.input_var.get().strip() == self.invoice_code:
            self.delete_btn.config(state="normal")
        else:
            self.delete_btn.config(state="disabled")

    def _on_confirm(self):
        self.result = True
        self.destroy()

    def _on_cancel(self):
        self.result = False
        self.destroy()


class DriverDeleteConfirmDialog(tk.Toplevel):
    """司机删除确认对话框 —— 必须输入正确的司機編號才能删除"""

    def __init__(self, parent, driver_code):
        super().__init__(parent)
        self.title("确认删除司机")
        self.result = False
        self.driver_code = driver_code
        self.geometry("430x230")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry("+{}+{}".format(x, y))

        self._build_ui(driver_code)
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

    def _build_ui(self, driver_code):
        frame = tk.Frame(self, bg='#f0f2f5')
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        tk.Label(
            frame,
            text="⚠ 此操作将删除该司机在 DRIVER_MASTER / VEHICLE_MASTER / DRIVER_CP 中的所有记录！",
            font=("Microsoft YaHei", 10, "bold"),
            bg='#f0f2f5', fg="red", wraplength=380, justify="left"
        ).pack(anchor="w", pady=(0, 10))

        tk.Label(frame, text="司機編號: " + str(driver_code),
                 bg='#f0f2f5', fg='#262626', font=("Microsoft YaHei UI", 11)
                 ).pack(anchor="w", pady=(0, 10))

        tk.Label(frame, text="请输入司機編號确认删除（输入匹配后按钮启用）:",
                 bg='#f0f2f5', fg='#595959').pack(anchor="w")

        input_frame = tk.Frame(frame, bg='#f0f2f5')
        input_frame.pack(fill="x", pady=(5, 15))

        self.input_var = tk.StringVar()
        self.input_var.trace_add("write", self._on_input_change)
        self.entry = tk.Entry(input_frame, textvariable=self.input_var, width=35,
                              font=('Microsoft YaHei UI', 10))
        self.entry.pack(side="left", fill="x", expand=True)
        self.entry.focus_set()

        btn_frame = tk.Frame(frame, bg='#f0f2f5')
        btn_frame.pack(fill="x")

        self.delete_btn = tk.Button(
            btn_frame, text="确认删除", bg='#ff4d4f', fg='#ffffff',
            font=('Microsoft YaHei UI', 10), bd=0, padx=16, pady=5,
            cursor='hand2', state='disabled', command=self._on_confirm
        )
        self.delete_btn.pack(side="left", padx=(0, 10))

        tk.Button(btn_frame, text="取消", bg='#f5f5f5', fg='#595959',
                  font=('Microsoft YaHei UI', 10), bd=0, padx=16, pady=5,
                  cursor='hand2', command=self._on_cancel).pack(side="left")

    def _on_input_change(self, *args):
        if self.input_var.get().strip() == self.driver_code:
            self.delete_btn.config(state="normal")
        else:
            self.delete_btn.config(state="disabled")

    def _on_confirm(self):
        self.result = True
        self.destroy()

    def _on_cancel(self):
        self.result = False
        self.destroy()


class SWTManagementSystem:
    """SWT 货运管理系统主类"""

    def __init__(self, root):
        self.root = root
        self.root.title("SWT 货运管理系统")
        self.root.geometry("1400x900")
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() - 1400) // 2
        y = (self.root.winfo_screenheight() - 900) // 2
        self.root.geometry(f"1400x900+{x}+{y}")
        self.root.configure(bg='#1a1a2e')
        self.root.deiconify()  # 显示窗口（AcrylicRoot 中 withdraw 了）

        # 配置文件路径
        self.config_file = os.path.join(os.path.dirname(__file__), 'swt_config.json')
        self.config = self.load_config()

        # ========== 数据库连接池 ==========
        self._db_config = {
            'host': self.config.get('db_host', '192.168.0.120'),
            'port': int(self.config.get('db_port', 3306)),
            'user': self.config.get('db_user', 'mysql'),
            'password': self.config.get('db_password', 'mysql'),
            'database': self.config.get('db_name', 'SWT'),
            'charset': 'utf8mb4'
        }
        # 创建连接池（最多10个连接）
        self.db_pool = PooledDB(
            creator=pymysql,
            maxconnections=10,
            mincached=2,
            maxcached=5,
            blocking=True,
            maxusage=None,
            setsession=[],
            ping=1,  # 取连接时检查连接是否有效
            **self._db_config
        )
        # 从连接池获取连接
        self.db_connected = True
        try:
            self.conn = self.db_pool.connection()
            self.cursor = self.conn.cursor(pymysql.cursors.DictCursor)
        except Exception as e:
            self.conn = None
            self.cursor = None
            self.db_connected = False
            self._db_error = str(e)

        # 锁，用于线程安全
        self._db_lock = threading.Lock()

        # 连接健康检查定时器
        if self.db_connected:
            self._check_connection()

        # 当前选中的菜单
        self.current_menu = None
        self.selected_rule_id = None

        # 创建界面
        self.create_sidebar()
        self.create_header()
        self.create_main_content()

        if self.db_connected:
            # 默认显示检测关键字页面
            self.show_keyword_check()
        else:
            # 数据库连接失败，跳转到设置页面
            self.show_settings()
            messagebox.showwarning("数据库连接失败",
                f"无法连接到 MySQL 数据库:\n{self._db_error}\n\n请检查数据库连接参数。")

    def load_config(self):
        """加载配置文件"""
        default_config = {
            'export_dir': r'X:\客户月结单',
            'summary_dir': r'X:\月结单',
            'biz_dir':    r'X:\月结单',      # 业务统计默认目录
            'driver_export_dir': r'X:\司机月结单',  # 司机月结单默认目录
            'driver_summary_dir': r'X:\司机汇总表',  # 司机运杂费汇总表默认目录
            'custom_table_customers': ['0332'],  # 自定义表专属客户列表
            'pct_tuo_company': '',   # 营运拖头车公司%
            'pct_tuo_driver': '',    # 营运拖头车司机%
            'pct_dun_company': '',   # 营运吨车公司%
            'pct_dun_driver': '',    # 营运吨车司机%
            'pct_flat_company': '',  # 平板车公司%
            'pct_flat_driver': '',   # 平板车司机%
            'db_host': '192.168.0.120',
            'db_port': '3306',
            'db_user': 'mysql',
            'db_password': 'mysql',
            'db_name': 'SWT'
        }
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                    # 合并默认值（防止旧配置缺少新字段）
                    for k, v in default_config.items():
                        if k not in loaded:
                            loaded[k] = v
                    return loaded
            except:
                return default_config
        return default_config

    def _check_connection(self):
        """检查数据库连接是否有效，必要时重连"""
        try:
            self.conn.ping(reconnect=True)
        except Exception:
            # 连接已断开，从池中获取新连接
            try:
                self.conn = self.db_pool.connection()
                self.cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            except Exception as e:
                print(f"数据库重连失败: {e}")

    def _safe_execute(self, sql, params=None):
        """线程安全的数据库执行（带自动重连）"""
        if self.conn is None or self.cursor is None:
            return False
        with self._db_lock:
            try:
                if params:
                    self.cursor.execute(sql, params)
                else:
                    self.cursor.execute(sql)
                return True
            except pymysql.OperationalError as e:
                # 连接断开，尝试重连
                if 'Lost connection' in str(e) or 'MySQL server has gone away' in str(e):
                    try:
                        self.conn = self.db_pool.connection()
                        self.cursor = self.conn.cursor(pymysql.cursors.DictCursor)
                        if params:
                            self.cursor.execute(sql, params)
                        else:
                            self.cursor.execute(sql)
                        return True
                    except Exception as re_err:
                        print(f"重连失败: {re_err}")
                        return False
                raise
            except Exception as e:
                print(f"SQL执行失败: {e}")
                return False

    def _normalize_misc_names(self, table_name, cust_code=None):
        """代码内执行雜費名稱归一化（不依赖存储过程）
        从 misc_name_rules 读规则，动态生成 UPDATE SQL 直接执行

        cust_code=None:  仅应用全局规则 (cust_code IS NULL)
        cust_code='0332': 应用全局规则 + 该客户专用规则，专用规则优先匹配
        """
        # 1. 获取雜費名稱列
        self.cursor.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = %s
              AND COLUMN_NAME LIKE '雜費名稱%%'
            ORDER BY ORDINAL_POSITION
        """, (table_name,))
        misc_cols = [r['COLUMN_NAME'] for r in self.cursor.fetchall()]
        if not misc_cols:
            return

        # 检查并自动添加 cust_code 列
        self.cursor.execute("""
            SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'misc_name_rules'
              AND COLUMN_NAME = 'cust_code'
        """)
        has_cust = self.cursor.fetchone()['cnt'] > 0
        if not has_cust:
            try:
                self.cursor.execute(
                    "ALTER TABLE misc_name_rules ADD COLUMN cust_code VARCHAR(10) DEFAULT NULL AFTER memo")
                self.conn.commit()
                has_cust = True
            except Exception:
                pass

        # 2. 读取规则（按客户范围过滤）
        if cust_code:
            if has_cust:
                self.cursor.execute(
                    "SELECT sort_order, pattern_type, keyword, replacement "
                    "FROM misc_name_rules "
                    "WHERE cust_code = %s OR cust_code IS NULL "
                    "ORDER BY CASE WHEN cust_code IS NULL THEN 1 ELSE 0 END, sort_order",
                    (cust_code,))
            else:
                self.cursor.execute(
                    "SELECT sort_order, pattern_type, keyword, replacement FROM misc_name_rules ORDER BY sort_order"
                )
        else:
            if has_cust:
                self.cursor.execute(
                    "SELECT sort_order, pattern_type, keyword, replacement "
                    "FROM misc_name_rules "
                    "WHERE cust_code IS NULL "
                    "ORDER BY sort_order")
            else:
                self.cursor.execute(
                    "SELECT sort_order, pattern_type, keyword, replacement FROM misc_name_rules ORDER BY sort_order"
                )
        rules = self.cursor.fetchall()
        if not rules:
            return

        # 3. 生成 UPDATE SET 子句
        set_parts = []
        for col in misc_cols:
            case_lines = []
            for r in rules:
                kw = r['keyword'].replace("\\", "\\\\").replace("'", "\\'")
                rep = r['replacement'].replace("\\", "\\\\").replace("'", "\\'")
                ptype = r['pattern_type']
                if ptype == 'exact':
                    case_lines.append(f"    WHEN `{col}` = '{kw}' THEN '{rep}'")
                elif ptype == 'contains':
                    case_lines.append(f"    WHEN `{col}` LIKE '%{kw}%' THEN '{rep}'")
                else:  # prefix
                    case_lines.append(f"    WHEN `{col}` LIKE '{kw}%%' THEN '{rep}'")
            case_lines.append(f"    ELSE `{col}`")
            set_parts.append(f"  `{col}` = CASE\n" + "\n".join(case_lines) + "\n  END")
        set_sql = ",\n".join(set_parts)

        # 4. 执行 UPDATE
        sql = f"UPDATE `{table_name}` SET\n{set_sql}\n  WHERE 1=1"
        self.cursor.execute(sql)

    def _pivot_misc_names(self, cust_code=None):
        """代码内执行客户月结单PIVOT（不依赖存储过程）
        从 misc_name_rules 读 replacement 列名，动态生成 PIVOT 表。
        方式：直接对 TEMP 表的 20 列雜費做 CASE WHEN，每行 TEMP → 一行 PIVOT。

        cust_code 作用同 _normalize_misc_names
        """
        # 检查 cust_code 列
        self.cursor.execute("""
            SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'misc_name_rules'
              AND COLUMN_NAME = 'cust_code'
        """)
        has_cust = self.cursor.fetchone()['cnt'] > 0
        if not has_cust:
            try:
                self.cursor.execute(
                    "ALTER TABLE misc_name_rules ADD COLUMN cust_code VARCHAR(10) DEFAULT NULL AFTER memo")
                self.conn.commit()
                has_cust = True
            except Exception:
                pass

        # Step 1: 收集所有合法的雜费列名（replacement去重，按客户范围过滤）
        if cust_code:
            if has_cust:
                self.cursor.execute("""
                    SELECT DISTINCT replacement
                    FROM misc_name_rules
                    WHERE replacement IS NOT NULL AND replacement != ''
                      AND replacement NOT LIKE '%%,%%' AND replacement NOT LIKE '%%.%%'
                      AND CHAR_LENGTH(replacement) <= 64 AND replacement NOT LIKE '*%%'
                      AND replacement NOT IN ('運費','運雜費合計')
                      AND (cust_code = %s OR cust_code IS NULL)
                    ORDER BY replacement
                """, (cust_code,))
            else:
                self.cursor.execute("""
                    SELECT DISTINCT replacement
                    FROM misc_name_rules
                    WHERE replacement IS NOT NULL AND replacement != ''
                      AND replacement NOT LIKE '%%,%%' AND replacement NOT LIKE '%%.%%'
                      AND CHAR_LENGTH(replacement) <= 64 AND replacement NOT LIKE '*%%'
                      AND replacement NOT IN ('運費','運雜費合計')
                    ORDER BY replacement
                """)
        else:
            if has_cust:
                self.cursor.execute("""
                    SELECT DISTINCT replacement
                    FROM misc_name_rules
                    WHERE replacement IS NOT NULL AND replacement != ''
                      AND replacement NOT LIKE '%%,%%' AND replacement NOT LIKE '%%.%%'
                      AND CHAR_LENGTH(replacement) <= 64 AND replacement NOT LIKE '*%%'
                      AND replacement NOT IN ('運費','運雜費合計')
                      AND cust_code IS NULL
                    ORDER BY replacement
                """)
            else:
                self.cursor.execute("""
                    SELECT DISTINCT replacement
                    FROM misc_name_rules
                    WHERE replacement IS NOT NULL AND replacement != ''
                      AND replacement NOT LIKE '%%,%%' AND replacement NOT LIKE '%%.%%'
                      AND CHAR_LENGTH(replacement) <= 64 AND replacement NOT LIKE '*%%'
                      AND replacement NOT IN ('運費','運雜費合計')
                    ORDER BY replacement
                """)
        replacements = [r['replacement'] for r in self.cursor.fetchall()]
        if not replacements:
            return

        # Step 2: 对每个 replacement，从20列雜費名稱中收集金额并求和
        # 方式：COALESCE(CASE WHEN 雜費名稱1=x THEN 雜費金額1 END,0)
        #      + COALESCE(CASE WHEN 雜費名稱2=x THEN 雜費金額2 END,0) + ...
        case_parts = []
        for rep in replacements:
            safe_rep = rep.replace("'", "\\'").replace("`", "``")
            terms = []
            for i in range(1, 21):
                terms.append(
                    f"COALESCE(CASE WHEN `雜費名稱{i}` = '{safe_rep}' THEN `雜費金額{i}` END, 0)"
                )
            case_parts.append(
                f"CAST({'+'.join(terms)} AS DECIMAL(14,2)) AS `{safe_rep}`"
            )

        # 构建 CREATE TABLE AS SELECT SQL
        select_list = ", ".join(case_parts)
        create_sql = (
            "CREATE TABLE `客户月结单PIVOT` ENGINE=MyISAM AS "
            "SELECT `id`, `日期`, `發票號`, `客戶編號`, `客戶名稱`, `司機姓名`, "
            "`香港車牌`, `大陸車牌`, `地區`, `櫃號`, `櫃尺碼`, `提單號`, `船名`, `托運號`, "
            "CAST(`運費` AS DECIMAL(14,2)) AS `運費`, "
            + select_list + " "
            "FROM `客户月结单TEMP`"
        )

        self.cursor.execute("DROP TABLE IF EXISTS `客户月结单PIVOT`")
        self.cursor.execute(create_sql)

        # Step 3: 计算運雜費合計
        self.cursor.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = '客户月结单PIVOT'
              AND COLUMN_NAME NOT IN (
                'id','日期','發票號','客戶編號','客戶名稱','司機姓名',
                '香港車牌','大陸車牌','地區','櫃號','櫃尺碼','提單號','船名','托運號','運費'
              )
            ORDER BY ORDINAL_POSITION
        """)
        misc_cols = [r['COLUMN_NAME'] for r in self.cursor.fetchall()]

        sum_expr = "+".join([f"COALESCE(`{c.replace('`','``')}`,0)" for c in misc_cols])
        if sum_expr:
            self.cursor.execute("ALTER TABLE `客户月结单PIVOT` ADD COLUMN `運雜費合計` DECIMAL(14,2) DEFAULT NULL")
            self.cursor.execute(
                "UPDATE `客户月结单PIVOT` SET `運雜費合計` = COALESCE(`運費`,0)+" + sum_expr
            )

        # Step 4: 合計行
        self.cursor.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = '客户月结单PIVOT'
            ORDER BY ORDINAL_POSITION
        """)
        all_cols = [r['COLUMN_NAME'] for r in self.cursor.fetchall()]
        fixed_cols = {
            'id','日期','發票號','客戶編號','客戶名稱','司機姓名',
            '香港車牌','大陸車牌','地區','櫃號','櫃尺碼','提單號','船名','托運號','運費','運雜費合計'
        }

        ins_cols = []
        ins_vals = []
        for c in all_cols:
            safe_c = c.replace('`', '``')
            ins_cols.append("`" + safe_c + "`")
            if c == '客戶名稱':
                ins_vals.append("'合計'")
            elif c == '運費' or c == '運雜費合計' or c not in fixed_cols:
                ins_vals.append("SUM(`" + safe_c + "`)")
            else:
                ins_vals.append("NULL")

        ins_sql = (
            "INSERT INTO `客户月结单PIVOT` (" + ",".join(ins_cols) + ") "
            "SELECT " + ",".join(ins_vals) + " FROM `客户月结单PIVOT`"
        )
        self.cursor.execute(ins_sql)

        # 转为 InnoDB
        self.cursor.execute("ALTER TABLE `客户月结单PIVOT` ENGINE = InnoDB")

    def _get_connection(self):
        """获取当前连接（用于需要手动提交的事务）"""
        try:
            self.conn.ping(reconnect=True)
        except Exception:
            self.conn = self.db_pool.connection()
            self.cursor = self.conn.cursor(pymysql.cursors.DictCursor)
        return self.conn

    def save_config(self):
        """保存配置文件"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            messagebox.showerror("保存配置失败", f"无法写入配置文件:\n{self.config_file}\n\n{str(e)}")
            return False

    def create_sidebar(self):
        """创建左侧侧边栏"""
        # 侧边栏框架
        self.sidebar = tk.Frame(self.root, bg='#0f1419', width=200)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)

        # Logo区域
        logo_frame = tk.Frame(self.sidebar, bg='#0f1419', height=60)
        logo_frame.pack(fill=tk.X, pady=10)

        logo_label = tk.Label(
            logo_frame,
            text="🚚 SWT 货运系统",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#0f1419',
            fg='#ffffff'
        )
        logo_label.pack(pady=15)

        # 菜单按钮
        self.menu_buttons = {}

        menus = [
            ('keyword_check', '🔍 检测关键字'),
            ('customer_mgmt', '👥 客户管理'),
            ('driver_mgmt', '🚗 司机管理'),
            ('company_stats', '📊 公司统计'),
            ('biz_stats', '📈 业务统计'),
            ('custom_table', '📋 自定义表'),
            ('delete_data', '🗑️ 删除数据'),
            ('add_driver', '➕ 增加司机'),
        ]

        for menu_id, menu_text in menus:
            btn = tk.Button(
                self.sidebar,
                text=menu_text,
                font=('Microsoft YaHei UI', 11),
                bg='#0f1419',
                fg='#a0a0a0',
                activebackground='#1890ff',
                activeforeground='#ffffff',
                bd=0,
                padx=20,
                pady=12,
                anchor='w',
                cursor='hand2',
                command=lambda m=menu_id: self.switch_menu(m)
            )
            btn.pack(fill=tk.X)
            self.menu_buttons[menu_id] = btn

        # 底部弹性空白，将设置按钮推到最下方
        spacer = tk.Frame(self.sidebar, bg='#0f1419')
        spacer.pack(fill=tk.BOTH, expand=True)

        # 分隔线
        sep = tk.Frame(self.sidebar, bg='#2a3a4a', height=1)
        sep.pack(fill=tk.X, padx=15, pady=5)

        # 设置按钮
        settings_btn = tk.Button(
            self.sidebar,
            text='⚙️ 设置',
            font=('Microsoft YaHei UI', 11),
            bg='#0f1419',
            fg='#a0a0a0',
            activebackground='#1890ff',
            activeforeground='#ffffff',
            bd=0,
            padx=20,
            pady=12,
            anchor='w',
            cursor='hand2',
            command=lambda: self.switch_menu('settings')
        )
        settings_btn.pack(fill=tk.X)
        self.menu_buttons['settings'] = settings_btn

    def create_header(self):
        """创建自定义标题栏（磨砂玻璃风格）"""
        self.titlebar = tk.Frame(self.root, bg='#1a1a2e', height=40)
        self.titlebar.pack(side=tk.TOP, fill=tk.X)
        self.titlebar.pack_propagate(False)

        # 左侧：拖拽区域（双击最大化）
        self.titlebar_drag = tk.Frame(self.titlebar, bg='#1a1a2e')
        self.titlebar_drag.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 图标 + 标题
        title_label = tk.Label(
            self.titlebar_drag,
            text="🚚 SWT 货运管理系统",
            font=('Microsoft YaHei UI', 10),
            bg='#1a1a2e',
            fg='#b0b0b0'
        )
        title_label.pack(side=tk.LEFT, padx=(15, 0), pady=8)
        self.breadcrumb_label = title_label  # 兼容现有代码

        # 状态指示
        self.db_status_label = tk.Label(
            self.titlebar_drag,
            text="● 连接正常",
            font=('Microsoft YaHei UI', 9),
            bg='#1a1a2e',
            fg='#52c41a'
        )
        self.db_status_label.pack(side=tk.RIGHT, padx=(0, 10), pady=8)

        # 右侧：窗口控制按钮
        btn_frame = tk.Frame(self.titlebar, bg='#1a1a2e')
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y)

        # 按钮样式
        btn_cfg = {
            'font': ('Segoe MDL2 Assets', 10),
            'bg': '#1a1a2e', 'fg': '#b0b0b0',
            'activebackground': '#2a2a4e', 'activeforeground': '#ffffff',
            'bd': 0, 'width': 4, 'cursor': 'hand2',
        }

        # 最小化
        btn_min = tk.Button(btn_frame, text='─', command=self._minimize_window, **btn_cfg)
        btn_min.pack(side=tk.LEFT, fill=tk.Y)

        # 最大化/还原
        self.btn_max = tk.Button(btn_frame, text='☐', command=self._toggle_maximize_window, **btn_cfg)
        self.btn_max.pack(side=tk.LEFT, fill=tk.Y)

        # 关闭（悬停红色）
        btn_close = tk.Button(btn_frame, text='✕',
                              font=('Segoe MDL2 Assets', 10),
                              bg='#1a1a2e', fg='#b0b0b0',
                              activebackground='#e81123', activeforeground='#ffffff',
                              bd=0, width=4, cursor='hand2',
                              command=self._close_window)
        btn_close.pack(side=tk.LEFT, fill=tk.Y)
        # 悬停变红
        btn_close.bind('<Enter>', lambda e: btn_close.config(bg='#c42b1c', fg='#ffffff'))
        btn_close.bind('<Leave>', lambda e: btn_close.config(bg='#1a1a2e', fg='#b0b0b0'))

        # 绑定拖拽
        self.titlebar_drag.bind('<Button-1>', self._start_drag)
        self.titlebar_drag.bind('<B1-Motion>', self._on_drag)
        self.titlebar_drag.bind('<Double-Button-1>', lambda e: self._toggle_maximize_window())

    def _start_drag(self, event):
        """开始拖拽窗口"""
        self._drag_offset_x = event.x
        self._drag_offset_y = event.y

    def _on_drag(self, event):
        """拖拽中：移动窗口位置"""
        if self.root._is_maximized:
            # 最大化状态下拖拽先还原窗口
            self.root._is_maximized = False
            w, h = self.root._restore_size[2], self.root._restore_size[3]
            # 鼠标在标题栏中的比例位置还原
            ratio_x = event.x / self.root.winfo_width()
            x = event.x_root - int(w * ratio_x)
            y = event.y_root - 10
            self.root.geometry(f"{w}x{h}+{x}+{y}")
            self._drag_offset_x = int(w * ratio_x)
            self._drag_offset_y = 10
            # 恢复圆角
            try:
                val = ctypes.c_int(1)
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    self.root.winfo_id(), 33, ctypes.byref(val), ctypes.sizeof(val))
            except Exception:
                pass
            return
        x = self.root.winfo_x() + event.x - self._drag_offset_x
        y = self.root.winfo_y() + event.y - self._drag_offset_y
        self.root.geometry(f"+{x}+{y}")

    def _toggle_maximize_window(self):
        """切换最大化/还原"""
        self.root._toggle_maximize()

    def _minimize_window(self):
        """最小化窗口"""
        self.root.state('iconic')

    def _close_window(self):
        """关闭窗口"""
        self.root.destroy()

    def create_main_content(self):
        """创建主内容区域"""
        self.main_content = tk.Frame(self.root, bg='#f0f2f5')
        self.main_content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def switch_menu(self, menu_id):
        """切换菜单"""
        # 重置所有按钮样式
        for btn in self.menu_buttons.values():
            btn.config(bg='#0f1419', fg='#a0a0a0')

        # 高亮当前按钮
        self.menu_buttons[menu_id].config(bg='#1890ff', fg='#ffffff')

        # 删除数据 / 增加司机 需要密码验证
        if menu_id in ('delete_data', 'add_driver'):
            from tkinter import simpledialog
            menu_name = "删除数据" if menu_id == 'delete_data' else "增加司机"
            pwd = simpledialog.askstring(
                "密码验证", "请输入访问密码：",
                parent=self.root, show="*"
            )
            if pwd != "23689666":
                self.menu_buttons[menu_id].config(bg='#0f1419', fg='#a0a0a0')
                if self.current_menu and self.current_menu in self.menu_buttons:
                    self.menu_buttons[self.current_menu].config(bg='#1890ff', fg='#ffffff')
                messagebox.showwarning("密码错误", "密码不正确，无法进入" + menu_name)
                return

        self.current_menu = menu_id

        # 清空主内容区
        for widget in self.main_content.winfo_children():
            widget.destroy()

        # 根据菜单ID显示对应内容
        if menu_id == 'keyword_check':
            self.show_keyword_check()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  检测关键字")
        elif menu_id == 'customer_mgmt':
            self.show_customer_mgmt()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  客户管理")
        elif menu_id == 'driver_mgmt':
            self.show_driver_mgmt()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  司机管理")
        elif menu_id == 'company_stats':
            self.show_company_stats()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  公司统计")
        elif menu_id == 'biz_stats':
            self.show_biz_stats()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  业务统计")
        elif menu_id == 'custom_table':
            self.show_custom_table()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  自定义表")
        elif menu_id == 'delete_data':
            self.show_delete_data()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  删除数据")
        elif menu_id == 'add_driver':
            self.show_add_driver()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  增加司机")
        elif menu_id == 'settings':
            self.show_settings()
            self.breadcrumb_label.config(text="SWT 货运管理系统  ›  设置")

    def show_settings(self):
        """显示设置页面（MySQL 连接配置）"""
        # ── 滚动容器 ──────────────────────────────────
        canvas = tk.Canvas(self.main_content, bg='#f0f2f5', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f2f5')
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=1180)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # ── 页面标题 ──────────────────────────────────
        title_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            title_frame,
            text="⚙️  系统设置",
            font=('Microsoft YaHei UI', 18, 'bold'),
            bg='#f0f2f5',
            fg='#1a1a2e'
        ).pack(side=tk.LEFT)

        # ── 卡片：数据库连接 ───────────────────────────
        card = tk.Frame(scrollable_frame, bg='#ffffff', relief='flat', bd=0)
        card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        # 卡片标题
        card_title_frame = tk.Frame(card, bg='#1890ff', height=4)
        card_title_frame.pack(fill=tk.X)

        tk.Label(
            card,
            text="🗄️  MySQL 数据库连接",
            font=('Microsoft YaHei UI', 13, 'bold'),
            bg='#ffffff',
            fg='#1a1a2e'
        ).pack(anchor='w', padx=20, pady=(15, 10))

        # 分隔线
        tk.Frame(card, bg='#f0f0f0', height=1).pack(fill=tk.X, padx=20, pady=(0, 15))

        # ── 表单区域 ──────────────────────────────────
        form_frame = tk.Frame(card, bg='#ffffff')
        form_frame.pack(fill=tk.X, padx=30, pady=(0, 10))

        # 各字段定义：(label, config_key, 占位符, 是否密码)
        fields = [
            ('服务器地址 (Host)', 'db_host',     '例：192.168.0.120',  False),
            ('端口 (Port)',       'db_port',     '例：3306',            False),
            ('用户名 (User)',     'db_user',     '例：mysql',           False),
            ('密码 (Password)',   'db_password', '请输入密码',          True),
            ('数据库名称',        'db_name',     '例：SWT',             False),
        ]

        self._settings_entries = {}

        for row_idx, (label_text, key, placeholder, is_password) in enumerate(fields):
            row = tk.Frame(form_frame, bg='#ffffff')
            row.pack(fill=tk.X, pady=8)

            # 标签
            tk.Label(
                row,
                text=label_text,
                font=('Microsoft YaHei UI', 10),
                bg='#ffffff',
                fg='#555555',
                width=18,
                anchor='w'
            ).pack(side=tk.LEFT)

            # 输入框
            entry = tk.Entry(
                row,
                font=('Microsoft YaHei UI', 10),
                bd=1,
                relief='solid',
                bg='#fafafa',
                fg='#222222',
                insertbackground='#222222',
                highlightthickness=1,
                highlightcolor='#1890ff',
                highlightbackground='#d9d9d9',
                width=35,
                show='*' if is_password else ''
            )
            entry.pack(side=tk.LEFT, ipady=5)

            # 填入当前配置值
            current_val = self.config.get(key, '')
            if current_val:
                entry.insert(0, str(current_val))

            self._settings_entries[key] = entry

        # ── 测试连接 + 保存按钮 ───────────────────────
        btn_frame = tk.Frame(card, bg='#ffffff')
        btn_frame.pack(fill=tk.X, padx=30, pady=(10, 20))

        tk.Button(
            btn_frame,
            text='🔗  测试连接',
            font=('Microsoft YaHei UI', 10, 'bold'),
            bg='#ffffff',
            fg='#1890ff',
            activebackground='#e6f7ff',
            activeforeground='#1890ff',
            bd=1,
            relief='solid',
            padx=18,
            pady=8,
            cursor='hand2',
            command=self._test_db_connection
        ).pack(side=tk.LEFT, padx=(0, 12))

        tk.Button(
            btn_frame,
            text='💾  保存设置',
            font=('Microsoft YaHei UI', 10, 'bold'),
            bg='#1890ff',
            fg='#ffffff',
            activebackground='#096dd9',
            activeforeground='#ffffff',
            bd=0,
            padx=18,
            pady=8,
            cursor='hand2',
            command=self._save_db_settings
        ).pack(side=tk.LEFT)

        # ── 状态提示标签 ──────────────────────────────
        self._settings_status = tk.Label(
            card,
            text='',
            font=('Microsoft YaHei UI', 10),
            bg='#ffffff',
            fg='#52c41a'
        )
        self._settings_status.pack(anchor='w', padx=30, pady=(0, 15))

    def _test_db_connection(self):
        """测试数据库连接"""
        self._settings_status.config(text='正在连接...', fg='#faad14')
        self.root.update()
        try:
            host     = self._settings_entries['db_host'].get().strip()
            port     = int(self._settings_entries['db_port'].get().strip() or 3306)
            user     = self._settings_entries['db_user'].get().strip()
            password = self._settings_entries['db_password'].get()
            db_name  = self._settings_entries['db_name'].get().strip()

            test_conn = pymysql.connect(
                host=host, port=port,
                user=user, password=password,
                database=db_name, charset='utf8mb4',
                connect_timeout=5
            )
            test_conn.close()
            self._settings_status.config(text='✅  连接成功！', fg='#52c41a')
        except Exception as e:
            self._settings_status.config(text=f'❌  连接失败：{e}', fg='#ff4d4f')

    def _save_db_settings(self):
        """保存数据库设置并重新连接"""
        host     = self._settings_entries['db_host'].get().strip()
        port_str = self._settings_entries['db_port'].get().strip()
        user     = self._settings_entries['db_user'].get().strip()
        password = self._settings_entries['db_password'].get()
        db_name  = self._settings_entries['db_name'].get().strip()

        if not host or not user or not db_name:
            self._settings_status.config(text='❌  地址、用户名、数据库名称不能为空', fg='#ff4d4f')
            return

        try:
            port = int(port_str or 3306)
        except ValueError:
            self._settings_status.config(text='❌  端口必须是数字', fg='#ff4d4f')
            return

        # 先测试连接
        self._settings_status.config(text='正在验证并保存...', fg='#faad14')
        self.root.update()
        try:
            test_conn = pymysql.connect(
                host=host, port=port,
                user=user, password=password,
                database=db_name, charset='utf8mb4',
                connect_timeout=5
            )
            test_conn.close()
        except Exception as e:
            self._settings_status.config(text=f'❌  连接失败，设置未保存：{e}', fg='#ff4d4f')
            return

        # 更新配置并保存
        self.config['db_host']     = host
        self.config['db_port']     = str(port)
        self.config['db_user']     = user
        self.config['db_password'] = password
        self.config['db_name']     = db_name
        self.save_config()

        # 重新建立数据库连接
        try:
            if self.conn:
                self.conn.close()
        except Exception:
            pass

        self.conn = pymysql.connect(
            host=host, port=port,
            user=user, password=password,
            database=db_name, charset='utf8mb4'
        )
        self.cursor = self.conn.cursor(pymysql.cursors.DictCursor)
        self.db_connected = True

        self._settings_status.config(text='✅  设置已保存，数据库连接已更新！', fg='#52c41a')

    def show_keyword_check(self):
        """显示检测关键字页面"""
        # 创建滚动区域
        canvas = tk.Canvas(self.main_content, bg='#f0f2f5', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f2f5')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=1180)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 页面标题
        title_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))

        title_label = tk.Label(
            title_frame,
            text="检测关键字",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg='#f0f2f5',
            fg='#262626'
        )
        title_label.pack(side=tk.LEFT)

        # ========== 月度检查卡片 ==========
        check_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        check_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        # 卡片标题
        card_header = tk.Frame(check_card, bg='#ffffff', padx=20, pady=15)
        card_header.pack(fill=tk.X)

        card_title = tk.Label(
            card_header,
            text="📅 月度检查",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff',
            fg='#262626'
        )
        card_title.pack(side=tk.LEFT)

        # 检查控制区
        control_frame = tk.Frame(check_card, bg='#ffffff', padx=20, pady=10)
        control_frame.pack(fill=tk.X)

        # 年月选择
        date_frame = tk.Frame(control_frame, bg='#ffffff')
        date_frame.pack(side=tk.LEFT)

        tk.Label(date_frame, text="选择年月:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        # 年份选择 - Combobox
        from datetime import datetime
        current_date = datetime.now()
        year_list = [str(y) for y in range(2003, 2051)]
        self.year_var = tk.StringVar(value=str(current_date.year))
        self.year_combo = ttk.Combobox(date_frame, textvariable=self.year_var,
                                       values=year_list, width=8, state='readonly',
                                       font=('Microsoft YaHei UI', 11))
        # 显式选中当前年份
        try:
            self.year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.year_combo.current(0)
        self.year_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(date_frame, text="年", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 15))

        # 月份选择 - Combobox
        month_list = [f"{m:02d}" for m in range(1, 13)]
        self.month_var = tk.StringVar(value=f"{current_date.month:02d}")
        self.month_combo = ttk.Combobox(date_frame, textvariable=self.month_var,
                                        values=month_list, width=6, state='readonly',
                                        font=('Microsoft YaHei UI', 11))
        try:
            self.month_combo.current(month_list.index(f"{current_date.month:02d}"))
        except ValueError:
            self.month_combo.current(0)
        self.month_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(date_frame, text="月", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 20))

        # 检查按钮
        check_btn = tk.Button(
            control_frame,
            text="开始检查",
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='#1890ff',
            fg='#ffffff',
            activebackground='#40a9ff',
            activeforeground='#ffffff',
            bd=0,
            padx=30,
            pady=8,
            cursor='hand2',
            command=self.check_monthly
        )
        check_btn.pack(side=tk.LEFT, padx=10)

        # 状态标签
        self.check_status_var = tk.StringVar(value="准备就绪")
        status_label = tk.Label(
            control_frame,
            textvariable=self.check_status_var,
            font=('Microsoft YaHei UI', 11),
            bg='#ffffff',
            fg='#8c8c8c'
        )
        status_label.pack(side=tk.LEFT, padx=(30, 0))

        # 结果显示区
        result_frame = tk.Frame(check_card, bg='#f6ffed', padx=20, pady=15)
        result_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

        result_title = tk.Label(
            result_frame,
            text="检查结果",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#f6ffed',
            fg='#262626'
        )
        result_title.pack(anchor='w', pady=(0, 10))

        self.check_result_text = tk.Text(
            result_frame,
            height=8,
            font=('Consolas', 10),
            bg='#ffffff',
            fg='#262626',
            relief='solid',
            bd=1,
            padx=10,
            pady=10
        )
        self.check_result_text.pack(fill=tk.X)
        self.check_result_text.insert(tk.END, "点击「开始检查」按钮开始检测...")
        self.check_result_text.configure(state='disabled')

        # ========== 规则管理区域 ==========
        rules_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        rules_frame.pack(fill=tk.BOTH, expand=True)

        # 左侧 - 规则列表
        list_card = tk.Frame(rules_frame, bg='#ffffff', bd=1, relief='solid')
        list_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 列表标题
        list_header = tk.Frame(list_card, bg='#ffffff', padx=15, pady=12)
        list_header.pack(fill=tk.X)

        list_title = tk.Label(
            list_header,
            text="📋 规则列表",
            font=('Microsoft YaHei UI', 13, 'bold'),
            bg='#ffffff',
            fg='#262626'
        )
        list_title.pack(side=tk.LEFT)

        # 搜索框
        search_frame = tk.Frame(list_header, bg='#ffffff')
        search_frame.pack(side=tk.RIGHT)

        tk.Label(search_frame, text="搜索:", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#8c8c8c').pack(side=tk.LEFT, padx=(0, 5))

        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', lambda *args: self.search_rules())
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=20)
        search_entry.pack(side=tk.LEFT)

        # 规则表格
        table_frame = tk.Frame(list_card, bg='#ffffff', padx=15, pady=10)
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ('id', 'sort_order', 'pattern_type', 'keyword', 'replacement', 'cust_code')
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', selectmode='browse', height=12)

        # 定义列
        self.tree.heading('id', text='ID')
        self.tree.heading('sort_order', text='顺序')
        self.tree.heading('pattern_type', text='类型')
        self.tree.heading('keyword', text='关键字')
        self.tree.heading('replacement', text='替换值')
        self.tree.heading('cust_code', text='客户')

        # 列宽
        self.tree.column('id', width=50, anchor='center')
        self.tree.column('sort_order', width=60, anchor='center')
        self.tree.column('pattern_type', width=80, anchor='center')
        self.tree.column('keyword', width=180)
        self.tree.column('replacement', width=200)
        self.tree.column('cust_code', width=70, anchor='center')

        # 滚动条
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        # 绑定选择事件
        self.tree.bind('<<TreeviewSelect>>', self.on_rule_select)

        # 统计信息
        stats_frame = tk.Frame(list_card, bg='#f6ffed', padx=15, pady=10)
        stats_frame.pack(fill=tk.X, side=tk.BOTTOM)

        self.stats_label = tk.Label(
            stats_frame,
            text="共 0 条规则",
            font=('Microsoft YaHei UI', 10),
            bg='#f6ffed',
            fg='#52c41a'
        )
        self.stats_label.pack(side=tk.LEFT)

        # 右侧 - 规则编辑
        edit_card = tk.Frame(rules_frame, bg='#ffffff', bd=1, relief='solid', width=350)
        edit_card.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        edit_card.pack_propagate(False)

        # 编辑标题
        edit_header = tk.Frame(edit_card, bg='#ffffff', padx=15, pady=12)
        edit_header.pack(fill=tk.X)

        edit_title = tk.Label(
            edit_header,
            text="✏️ 规则编辑",
            font=('Microsoft YaHei UI', 13, 'bold'),
            bg='#ffffff',
            fg='#262626'
        )
        edit_title.pack(side=tk.LEFT)

        # 表单
        form_frame = tk.Frame(edit_card, bg='#ffffff', padx=20, pady=15)
        form_frame.pack(fill=tk.X)

        # ID
        tk.Label(form_frame, text="ID:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').grid(row=0, column=0, sticky=tk.W, pady=8)
        self.id_var = tk.StringVar()
        id_entry = tk.Entry(form_frame, textvariable=self.id_var, state='readonly', width=25, font=('Microsoft YaHei UI', 11), relief='solid', bd=1)
        id_entry.grid(row=0, column=1, sticky=tk.W, pady=8, padx=(10, 0))

        # 排序顺序
        tk.Label(form_frame, text="排序:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').grid(row=1, column=0, sticky=tk.W, pady=8)
        self.sort_order_var = tk.StringVar()
        sort_entry = tk.Entry(form_frame, textvariable=self.sort_order_var, state='readonly', width=25, font=('Microsoft YaHei UI', 11), relief='solid', bd=1)
        sort_entry.grid(row=1, column=1, sticky=tk.W, pady=8, padx=(10, 0))

        # 规则类型
        tk.Label(form_frame, text="类型:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').grid(row=2, column=0, sticky=tk.W, pady=8)
        self.pattern_type_var = tk.StringVar(value='prefix')
        type_entry = tk.Entry(form_frame, textvariable=self.pattern_type_var, state='readonly', width=25, font=('Microsoft YaHei UI', 11), relief='solid', bd=1)
        type_entry.grid(row=2, column=1, sticky=tk.W, pady=8, padx=(10, 0))

        # 关键字
        tk.Label(form_frame, text="关键字:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#262626').grid(row=3, column=0, sticky=tk.W, pady=8)
        self.keyword_var = tk.StringVar()
        keyword_entry = tk.Entry(form_frame, textvariable=self.keyword_var, width=25, font=('Microsoft YaHei UI', 11), relief='solid', bd=1)
        keyword_entry.grid(row=3, column=1, sticky=tk.W, pady=8, padx=(10, 0))

        # 替换值
        tk.Label(form_frame, text="替换值:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#262626').grid(row=4, column=0, sticky=tk.W, pady=8)
        self.replacement_var = tk.StringVar()
        replacement_entry = tk.Entry(form_frame, textvariable=self.replacement_var, width=25, font=('Microsoft YaHei UI', 11), relief='solid', bd=1)
        replacement_entry.grid(row=4, column=1, sticky=tk.W, pady=8, padx=(10, 0))

        # 客户（专属规则隔离）
        tk.Label(form_frame, text="客户:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#262626').grid(row=5, column=0, sticky=tk.W, pady=8)
        self.cust_code_var = tk.StringVar()
        cust_code_entry = tk.Entry(form_frame, textvariable=self.cust_code_var, width=25, font=('Microsoft YaHei UI', 11), relief='solid', bd=1)
        cust_code_entry.grid(row=5, column=1, sticky=tk.W, pady=8, padx=(10, 0))

        # 按钮区 - 合并所有操作按钮
        btn_frame = tk.Frame(edit_card, bg='#ffffff', padx=20, pady=15)
        btn_frame.pack(fill=tk.X)

        # 顶部操作按钮行：新增 修改 删除 清空
        top_btn_frame = tk.Frame(btn_frame, bg='#ffffff')
        top_btn_frame.pack(fill=tk.X, pady=(0, 10))

        add_btn = tk.Button(
            top_btn_frame,
            text="新增",
            font=('Microsoft YaHei UI', 10, 'bold'),
            bg='#52c41a',
            fg='#ffffff',
            activebackground='#73d13d',
            bd=0,
            padx=20,
            pady=6,
            cursor='hand2',
            command=self.add_rule
        )
        add_btn.pack(side=tk.LEFT, padx=5)

        edit_btn = tk.Button(
            top_btn_frame,
            text="修改",
            font=('Microsoft YaHei UI', 10, 'bold'),
            bg='#1890ff',
            fg='#ffffff',
            activebackground='#40a9ff',
            bd=0,
            padx=20,
            pady=6,
            cursor='hand2',
            command=self.edit_rule
        )
        edit_btn.pack(side=tk.LEFT, padx=5)

        del_btn = tk.Button(
            top_btn_frame,
            text="删除",
            font=('Microsoft YaHei UI', 10, 'bold'),
            bg='#ff4d4f',
            fg='#ffffff',
            activebackground='#ff7875',
            bd=0,
            padx=20,
            pady=6,
            cursor='hand2',
            command=self.delete_rule
        )
        del_btn.pack(side=tk.LEFT, padx=5)

        clear_btn = tk.Button(
            top_btn_frame,
            text="清空",
            font=('Microsoft YaHei UI', 10),
            bg='#f5f5f5',
            fg='#595959',
            activebackground='#d9d9d9',
            bd=0,
            padx=20,
            pady=6,
            cursor='hand2',
            command=self.clear_form
        )
        clear_btn.pack(side=tk.LEFT, padx=5)

        # 底部操作按钮：更新（橙色突出显示）
        update_btn = tk.Button(
            btn_frame,
            text="更新",
            font=('Microsoft YaHei UI', 10, 'bold'),
            bg='#fa8c16',
            fg='#ffffff',
            activebackground='#ffa940',
            bd=0,
            padx=15,
            pady=8,
            cursor='hand2',
            command=self.update_procedures
        )
        update_btn.pack(fill=tk.X, pady=(0, 10))

        # 导入导出按钮
        export_import_frame = tk.Frame(btn_frame, bg='#ffffff')
        export_import_frame.pack(fill=tk.X)

        export_btn = tk.Button(
            export_import_frame,
            text="📤 导出规则",
            font=('Microsoft YaHei UI', 10),
            bg='#ffffff',
            fg='#595959',
            activebackground='#f5f5f5',
            bd=1,
            relief='solid',
            padx=15,
            pady=8,
            cursor='hand2',
            command=self.export_rules
        )
        export_btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        import_btn = tk.Button(
            export_import_frame,
            text="📥 导入规则",
            font=('Microsoft YaHei UI', 10),
            bg='#ffffff',
            fg='#595959',
            activebackground='#f5f5f5',
            bd=1,
            relief='solid',
            padx=15,
            pady=8,
            cursor='hand2',
            command=self.import_rules
        )
        import_btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # 加载规则列表
        self.refresh_rules()

    def show_customer_mgmt(self):
        """显示客户管理页面"""
        # 创建滚动区域
        canvas = tk.Canvas(self.main_content, bg='#f0f2f5', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f2f5')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=1180)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 页面标题
        title_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))

        title_label = tk.Label(
            title_frame,
            text="客户管理",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg='#f0f2f5',
            fg='#262626'
        )
        title_label.pack(side=tk.LEFT)

        # ========== 客户月结单导出卡片 ==========
        export_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        export_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        # 卡片标题
        card_header = tk.Frame(export_card, bg='#ffffff', padx=20, pady=15)
        card_header.pack(fill=tk.X)

        card_title = tk.Label(
            card_header,
            text="📊 客户月结单导出",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff',
            fg='#262626'
        )
        card_title.pack(side=tk.LEFT)

        # 说明文字
        desc_frame = tk.Frame(export_card, bg='#ffffff', padx=20, pady=10)
        desc_frame.pack(fill=tk.X)

        desc_label = tk.Label(
            desc_frame,
            text="导出指定月份的客户月结单Excel文件，每个客户一个单独的文件。",
            font=('Microsoft YaHei UI', 10),
            bg='#ffffff',
            fg='#8c8c8c',
            wraplength=1000
        )
        desc_label.pack(anchor='w')

        # 导出控制区
        control_frame = tk.Frame(export_card, bg='#ffffff', padx=20, pady=10)
        control_frame.pack(fill=tk.X)

        # 年月选择
        date_frame = tk.Frame(control_frame, bg='#ffffff')
        date_frame.pack(side=tk.LEFT)

        tk.Label(date_frame, text="选择年月:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        # 年份选择 - Combobox
        from datetime import datetime
        current_date = datetime.now()
        year_list = [str(y) for y in range(2003, 2051)]
        self.export_year_var = tk.StringVar(value=str(current_date.year))
        self.export_year_combo = ttk.Combobox(date_frame, textvariable=self.export_year_var,
                                              values=year_list, width=8, state='readonly',
                                              font=('Microsoft YaHei UI', 11))
        # 显式选中当前年份
        try:
            self.export_year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.export_year_combo.current(0)
        self.export_year_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(date_frame, text="年", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 15))

        # 月份选择 - Combobox
        month_list = [f"{m:02d}" for m in range(1, 13)]
        self.export_month_var = tk.StringVar(value=f"{current_date.month:02d}")
        self.export_month_combo = ttk.Combobox(date_frame, textvariable=self.export_month_var,
                                               values=month_list, width=6, state='readonly',
                                               font=('Microsoft YaHei UI', 11))
        try:
            self.export_month_combo.current(month_list.index(f"{current_date.month:02d}"))
        except ValueError:
            self.export_month_combo.current(0)
        self.export_month_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(date_frame, text="月", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 20))

        # 输出目录选择
        dir_frame = tk.Frame(control_frame, bg='#ffffff')
        dir_frame.pack(side=tk.LEFT)

        tk.Label(dir_frame, text="输出目录:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        self.export_dir_var = tk.StringVar(value=self.config.get('export_dir', r'X:\客户月结单'))
        dir_entry = tk.Entry(dir_frame, textvariable=self.export_dir_var, width=30, font=('Microsoft YaHei UI', 10), bg='#f5f5f5', fg='#262626', relief='solid', bd=1)
        dir_entry.pack(side=tk.LEFT, padx=(0, 5))

        dir_browse_btn = tk.Button(
            dir_frame,
            text="浏览...",
            font=('Microsoft YaHei UI', 9),
            bg='#fafafa',
            fg='#595959',
            activebackground='#e8e8e8',
            activeforeground='#262626',
            bd=1,
            relief='solid',
            padx=10,
            pady=4,
            cursor='hand2',
            command=self.browse_export_dir
        )
        dir_browse_btn.pack(side=tk.LEFT, padx=(0, 20))

        # 导出范围选择
        self.cust_export_mode_var = tk.StringVar(value='all')
        tk.Radiobutton(control_frame, text="全部客户", variable=self.cust_export_mode_var, value='all',
                       font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#262626',
                       activebackground='#ffffff', cursor='hand2').pack(side=tk.LEFT)
        tk.Radiobutton(control_frame, text="选择客户", variable=self.cust_export_mode_var, value='select',
                       font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#262626',
                       activebackground='#ffffff', cursor='hand2').pack(side=tk.LEFT, padx=(2, 8))

        # 导出按钮
        export_btn = tk.Button(
            control_frame,
            text="开始导出",
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='#52c41a',
            fg='#ffffff',
            activebackground='#73d13d',
            activeforeground='#ffffff',
            bd=0,
            padx=30,
            pady=8,
            cursor='hand2',
            command=self.export_monthly_excel
        )
        export_btn.pack(side=tk.LEFT, padx=10)

        # 状态标签
        self.export_status_var = tk.StringVar(value="准备就绪")
        status_label = tk.Label(
            control_frame,
            textvariable=self.export_status_var,
            font=('Microsoft YaHei UI', 11),
            bg='#ffffff',
            fg='#8c8c8c'
        )
        status_label.pack(side=tk.LEFT, padx=(30, 0))

        # 结果显示区
        result_frame = tk.Frame(export_card, bg='#e6f7ff', padx=20, pady=15)
        result_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

        result_title = tk.Label(
            result_frame,
            text="导出结果",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#e6f7ff',
            fg='#262626'
        )
        result_title.pack(anchor='w', pady=(0, 10))

        self.export_result_text = tk.Text(
            result_frame,
            height=8,
            font=('Consolas', 10),
            bg='#ffffff',
            fg='#262626',
            relief='solid',
            bd=1,
            padx=10,
            pady=10
        )
        self.export_result_text.pack(fill=tk.X)
        self.export_result_text.insert(tk.END, "点击「开始导出」按钮开始导出客户月结单...")
        self.export_result_text.configure(state='disabled')

        # ========== 客户汇总表导出卡片 ==========
        summary_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        summary_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        # 卡片标题
        summary_header = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=15)
        summary_header.pack(fill=tk.X)

        summary_title = tk.Label(
            summary_header,
            text="📊 客户汇总表导出",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff',
            fg='#262626'
        )
        summary_title.pack(side=tk.LEFT)

        # 说明文字
        summary_desc_frame = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=10)
        summary_desc_frame.pack(fill=tk.X)

        summary_desc_label = tk.Label(
            summary_desc_frame,
            text="导出指定月份的客户汇总表Excel文件，包含所有客户的运费、杂费、代理费等统计数据。",
            font=('Microsoft YaHei UI', 10),
            bg='#ffffff',
            fg='#8c8c8c',
            wraplength=1000
        )
        summary_desc_label.pack(anchor='w')

        # 导出控制区
        summary_control_frame = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=10)
        summary_control_frame.pack(fill=tk.X)

        # 年月选择
        summary_date_frame = tk.Frame(summary_control_frame, bg='#ffffff')
        summary_date_frame.pack(side=tk.LEFT)

        tk.Label(summary_date_frame, text="选择年月:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        # 年份选择 - Combobox
        from datetime import datetime
        current_date = datetime.now()
        year_list = [str(y) for y in range(2003, 2051)]
        self.summary_year_var = tk.StringVar(value=str(current_date.year))
        self.summary_year_combo = ttk.Combobox(summary_date_frame, textvariable=self.summary_year_var,
                                               values=year_list, width=8, state='readonly',
                                               font=('Microsoft YaHei UI', 11))
        # 显式选中当前年份
        try:
            self.summary_year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.summary_year_combo.current(0)
        self.summary_year_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(summary_date_frame, text="年", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 15))

        # 月份选择 - Combobox
        month_list = [f"{m:02d}" for m in range(1, 13)]
        self.summary_month_var = tk.StringVar(value=f"{current_date.month:02d}")
        self.summary_month_combo = ttk.Combobox(summary_date_frame, textvariable=self.summary_month_var,
                                                values=month_list, width=6, state='readonly',
                                                font=('Microsoft YaHei UI', 11))
        try:
            self.summary_month_combo.current(month_list.index(f"{current_date.month:02d}"))
        except ValueError:
            self.summary_month_combo.current(0)
        self.summary_month_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(summary_date_frame, text="月", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 20))

        # 导出目录
        summary_dir_frame = tk.Frame(summary_control_frame, bg='#ffffff')
        summary_dir_frame.pack(side=tk.LEFT, padx=(20, 0))

        tk.Label(summary_dir_frame, text="导出目录:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        self.summary_dir_var = tk.StringVar(value=self.config.get('summary_dir', r'X:\月结单'))
        summary_dir_entry = tk.Entry(summary_dir_frame, textvariable=self.summary_dir_var, width=40, font=('Microsoft YaHei UI', 11))
        summary_dir_entry.pack(side=tk.LEFT, padx=(0, 10))

        summary_browse_btn = tk.Button(
            summary_dir_frame,
            text="浏览",
            font=('Microsoft YaHei UI', 10),
            bg='#1890ff',
            fg='#ffffff',
            activebackground='#40a9ff',
            activeforeground='#ffffff',
            bd=0,
            padx=15,
            pady=6,
            cursor='hand2',
            command=lambda: self.browse_summary_dir()
        )
        summary_browse_btn.pack(side=tk.LEFT)

        # 导出按钮
        summary_export_btn = tk.Button(
            summary_control_frame,
            text="导出汇总表",
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='#52c41a',
            fg='#ffffff',
            activebackground='#73d13d',
            activeforeground='#ffffff',
            bd=0,
            padx=30,
            pady=8,
            cursor='hand2',
            command=self.export_customer_summary_gui
        )
        summary_export_btn.pack(side=tk.LEFT, padx=(20, 10))

        # 状态标签
        self.summary_status_var = tk.StringVar(value="准备就绪")
        summary_status_label = tk.Label(
            summary_control_frame,
            textvariable=self.summary_status_var,
            font=('Microsoft YaHei UI', 11),
            bg='#ffffff',
            fg='#8c8c8c'
        )
        summary_status_label.pack(side=tk.LEFT, padx=(30, 0))

        # 结果显示区域
        summary_result_frame = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        summary_result_frame.pack(fill=tk.X, pady=(0, 20), ipady=10)

        result_header = tk.Frame(summary_result_frame, bg='#ffffff', padx=15, pady=12)
        result_header.pack(fill=tk.X)

        result_title = tk.Label(
            result_header,
            text="📄 导出结果",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#ffffff',
            fg='#262626'
        )
        result_title.pack(side=tk.LEFT)

        self.summary_result_text = tk.Text(
            summary_result_frame,
            height=8,
            font=('Consolas', 10),
            bg='#ffffff',
            fg='#262626',
            relief='solid',
            bd=1,
            padx=10,
            pady=10
        )
        self.summary_result_text.pack(fill=tk.X)
        self.summary_result_text.insert(tk.END, "点击「导出汇总表」按钮开始导出...")
        self.summary_result_text.configure(state='disabled')

        # 说明卡片
        info_card = tk.Frame(scrollable_frame, bg='#f6ffed', bd=1, relief='solid')
        info_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        info_frame = tk.Frame(info_card, bg='#f6ffed', padx=20, pady=15)
        info_frame.pack(fill=tk.X)

        info_title = tk.Label(
            info_frame,
            text="ℹ️ 导出说明",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#f6ffed',
            fg='#262626'
        )
        info_title.pack(anchor='w', pady=(0, 10))

        info_text = tk.Label(
            info_frame,
            text="• 导出流程：生成客户月结单TEMP → 标准化杂费名称 → 透视转换 → 导出Excel\n"
                 "• 输出目录：X:\\客户月结单\\客户月结单{年月}\n"
                 "• 文件格式：{客户名}{年月}.xlsx\n"
                 "• 建议操作：先在「检测关键字」页面检查是否有新增未标准化项，确认无新增后再导出",
            font=('Microsoft YaHei UI', 10),
            bg='#f6ffed',
            fg='#595959',
            justify=tk.LEFT,
            wraplength=1000
        )
        info_text.pack(anchor='w')

        # ========== Invoice 对账单导出卡片 ==========
        invoice_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        invoice_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        # 卡片标题
        invoice_header = tk.Frame(invoice_card, bg='#ffffff', padx=20, pady=15)
        invoice_header.pack(fill=tk.X)

        invoice_title = tk.Label(
            invoice_header,
            text="🧾 Invoice 对账单导出",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff',
            fg='#262626'
        )
        invoice_title.pack(anchor='w')

        invoice_desc = tk.Label(
            invoice_header,
            text="按周汇总运杂费合计，生成 Invoice 对账单 Excel（手动输入每周起止日期）",
            font=('Microsoft YaHei UI', 10),
            bg='#ffffff',
            fg='#8c8c8c'
        )
        invoice_desc.pack(anchor='w', pady=(3, 0))

        # 控制区域
        invoice_control = tk.Frame(invoice_card, bg='#f5f5f5', padx=20, pady=15)
        invoice_control.pack(fill=tk.X)

        # 年月选择行
        invoice_date_frame = tk.Frame(invoice_control, bg='#f5f5f5')
        invoice_date_frame.pack(fill=tk.X, pady=(0, 10))

        from datetime import datetime
        current_date = datetime.now()
        year_list = [str(y) for y in range(2003, 2051)]
        self.invoice_year_var = tk.StringVar(value=str(current_date.year))
        self.invoice_year_combo = ttk.Combobox(invoice_date_frame, textvariable=self.invoice_year_var,
                                               values=year_list, width=8, state='readonly',
                                               font=('Microsoft YaHei UI', 11))
        try:
            self.invoice_year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.invoice_year_combo.current(0)
        self.invoice_year_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(invoice_date_frame, text="年", font=('Microsoft YaHei UI', 11), bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT, padx=(0, 15))

        month_list = [f"{m:02d}" for m in range(1, 13)]
        self.invoice_month_var = tk.StringVar(value=f"{current_date.month:02d}")
        self.invoice_month_combo = ttk.Combobox(invoice_date_frame, textvariable=self.invoice_month_var,
                                                values=month_list, width=6, state='readonly',
                                                font=('Microsoft YaHei UI', 11))
        try:
            self.invoice_month_combo.current(month_list.index(f"{current_date.month:02d}"))
        except ValueError:
            self.invoice_month_combo.current(0)
        self.invoice_month_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(invoice_date_frame, text="月", font=('Microsoft YaHei UI', 11), bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT, padx=(0, 20))

        # 选择客户按钮
        self.invoice_cust_label = tk.Label(invoice_date_frame, text="未选择客户", font=('Microsoft YaHei UI', 10),
                                           bg='#f5f5f5', fg='#8c8c8c')
        self.invoice_cust_label.pack(side=tk.LEFT, padx=(0, 5))

        def invoice_select_customer():
            year = int(self.invoice_year_var.get())
            month = int(self.invoice_month_var.get())
            if month == 12:
                next_m = datetime(year + 1, 1, 1)
            else:
                next_m = datetime(year, month + 1, 1)
            date_start = f'{year}-{month:02d}-01'
            date_end = next_m.strftime('%Y-%m-%d')
            try:
                self.cursor.execute("""
                    SELECT DISTINCT im.CUSTCODE, COALESCE(cm.NAME, '') AS CUSTNAME
                    FROM INVOICE_MASTER im
                    LEFT JOIN CUST_MASTER cm ON cm.CUSTCODE = im.CUSTCODE
                    WHERE im.INVDATE >= %s AND im.INVDATE < %s
                    ORDER BY im.CUSTCODE
                """, (date_start, date_end))
                rows = self.cursor.fetchall()
                if not rows:
                    messagebox.showinfo("提示", f"{year}年{month:02d}月 没有发票数据")
                    return
                items = []
                for r in rows:
                    code = r['CUSTCODE'] if isinstance(r, dict) else r[0]
                    name = r['CUSTNAME'] if isinstance(r, dict) else r[1]
                    items.append((code, f"{name}  [{code}]"))
                selected = self.show_multiselect_dialog("选择要导出的客户", items)
                if selected is not None:
                    self._invoice_selected_codes = selected
                    self._invoice_selected_names = {}
                    for r in rows:
                        code = r['CUSTCODE'] if isinstance(r, dict) else r[0]
                        name = r['CUSTNAME'] if isinstance(r, dict) else r[1]
                        if code in selected:
                            self._invoice_selected_names[code] = name
                    self.invoice_cust_label.configure(text=f"已选 {len(selected)} 个客户", fg='#52c41a')
            except Exception as e:
                messagebox.showerror("错误", str(e))

        self._invoice_selected_codes = []
        self._invoice_selected_names = {}

        tk.Button(invoice_date_frame, text="选择客户", font=('Microsoft YaHei UI', 10),
                  bg='#1890ff', fg='white', relief='flat', padx=12, pady=3,
                  cursor='hand2', command=invoice_select_customer).pack(side=tk.LEFT, padx=(5, 0))

        # 输出目录行
        invoice_dir_frame = tk.Frame(invoice_control, bg='#f5f5f5')
        invoice_dir_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(invoice_dir_frame, text="输出目录：", font=('Microsoft YaHei UI', 11),
                 bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT)

        self.invoice_dir_var = tk.StringVar(value=self.config.get('invoice_output_dir', ''))
        self.invoice_dir_entry = tk.Entry(invoice_dir_frame, textvariable=self.invoice_dir_var,
                                          font=('Microsoft YaHei UI', 10), width=55, state='readonly')
        self.invoice_dir_entry.pack(side=tk.LEFT, padx=(5, 5))

        def invoice_browse_dir():
            d = filedialog.askdirectory(title="选择 Invoice 输出目录")
            if d:
                self.invoice_dir_var.set(d)
                self.config['invoice_output_dir'] = d
                self.save_config()

        tk.Button(invoice_dir_frame, text="浏览...", font=('Microsoft YaHei UI', 10),
                  bg='#f0f0f0', relief='flat', padx=10, pady=3,
                  cursor='hand2', command=invoice_browse_dir).pack(side=tk.LEFT)

        # 周分段输入区
        period_label_frame = tk.Frame(invoice_control, bg='#f5f5f5')
        period_label_frame.pack(fill=tk.X, pady=(0, 5))

        tk.Label(period_label_frame, text="周分段日期（手动输入每段起止日期）：",
                 font=('Microsoft YaHei UI', 11), bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT)

        tk.Button(period_label_frame, text="+ 添加分段", font=('Microsoft YaHei UI', 10),
                  bg='#52c41a', fg='white', relief='flat', padx=10, pady=2,
                  cursor='hand2', command=lambda: _add_invoice_period_row(periods_frame)).pack(side=tk.RIGHT)

        tk.Button(period_label_frame, text="保存分段", font=('Microsoft YaHei UI', 10),
                  bg='#faad14', fg='white', relief='flat', padx=10, pady=2,
                  cursor='hand2', command=lambda: _save_invoice_periods()).pack(side=tk.RIGHT, padx=(0, 5))

        periods_frame = tk.Frame(invoice_control, bg='#f5f5f5')
        periods_frame.pack(fill=tk.X)
        self._invoice_periods_frame = periods_frame

        # 周分段行列表（每行: 起始日 Entry - 结束日 Entry + 删除按钮）
        self._invoice_period_entries = []
        self._invoice_month_labels = []

        def _refresh_period_month_labels():
            """切换年月时更新所有分段行的月份Label"""
            m = f"{int(self.invoice_month_var.get()):02d}"
            for lbl in self._invoice_month_labels:
                lbl.configure(text=f"/{m}")

        def _save_invoice_periods():
            """保存当前分段日期到配置文件"""
            periods_data = []
            for s_entry, e_entry in self._invoice_period_entries:
                s = s_entry.get().strip()
                e = e_entry.get().strip()
                if s or e:
                    periods_data.append([s, e])
            if not periods_data:
                messagebox.showwarning("提示", "没有可保存的分段日期")
                return
            ym = f"{self.invoice_year_var.get()}-{self.invoice_month_var.get()}"
            invoice_periods = self.config.get('invoice_periods', {})
            invoice_periods[ym] = periods_data
            self.config['invoice_periods'] = invoice_periods
            self.save_config()
            messagebox.showinfo("保存成功", f"已保存 {ym} 的 {len(periods_data)} 个分段日期")

        def _add_invoice_period_row(parent_frame):
            row_frame = tk.Frame(parent_frame, bg='#f5f5f5')
            row_frame.pack(fill=tk.X, pady=2)
            idx = len(self._invoice_period_entries)
            m = f"{int(self.invoice_month_var.get()):02d}"

            tk.Label(row_frame, text=f"第{idx+1}段：", font=('Microsoft YaHei UI', 10),
                     bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT, padx=(0, 3))

            start_entry = tk.Entry(row_frame, font=('Microsoft YaHei UI', 10), width=4)
            start_entry.pack(side=tk.LEFT, padx=(0, 1))

            month_label_s = tk.Label(row_frame, text=f"/{m}", font=('Microsoft YaHei UI', 10),
                                     bg='#f5f5f5', fg='#595959')
            month_label_s.pack(side=tk.LEFT, padx=(0, 8))

            tk.Label(row_frame, text="至", font=('Microsoft YaHei UI', 10),
                     bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT, padx=(0, 8))

            end_entry = tk.Entry(row_frame, font=('Microsoft YaHei UI', 10), width=4)
            end_entry.pack(side=tk.LEFT, padx=(0, 1))

            month_label_e = tk.Label(row_frame, text=f"/{m}", font=('Microsoft YaHei UI', 10),
                                     bg='#f5f5f5', fg='#595959')
            month_label_e.pack(side=tk.LEFT, padx=(0, 10))

            self._invoice_month_labels.append(month_label_s)
            self._invoice_month_labels.append(month_label_e)

            def remove_row(frame=row_frame, entries_ref=self._invoice_period_entries,
                           lbl_s=month_label_s, lbl_e=month_label_e):
                if len(entries_ref) <= 1:
                    return
                frame.destroy()
                entries_ref.remove((start_entry, end_entry))
                self._invoice_month_labels.remove(lbl_s)
                self._invoice_month_labels.remove(lbl_e)

            tk.Button(row_frame, text="✕", font=('Microsoft YaHei UI', 9),
                      bg='#ff4d4f', fg='white', relief='flat', padx=6, pady=1,
                      cursor='hand2', command=remove_row).pack(side=tk.LEFT)

            self._invoice_period_entries.append((start_entry, end_entry))

        def _apply_periods(periods_data):
            """清空现有分段行并用 periods_data 重建"""
            for widget in periods_frame.winfo_children():
                widget.destroy()
            self._invoice_period_entries.clear()
            self._invoice_month_labels.clear()
            for s, e in periods_data:
                _add_invoice_period_row(periods_frame)
                self._invoice_period_entries[-1][0].insert(0, s)
                self._invoice_period_entries[-1][1].insert(0, e)

        def _get_default_periods():
            """获取默认的4个空分段，最后一段结束日为月末"""
            last_day = calendar.monthrange(int(self.invoice_year_var.get()),
                                            int(self.invoice_month_var.get()))[1]
            return [['', ''], ['', ''], ['', ''], ['', str(last_day)]]

        def _on_ym_changed():
            """切换年月时：只更新月份Label和最后一段月末，保留已输入的日"""
            _refresh_period_month_labels()
            if self._invoice_period_entries:
                last_day = calendar.monthrange(int(self.invoice_year_var.get()),
                                                int(self.invoice_month_var.get()))[1]
                self._invoice_period_entries[-1][1].delete(0, tk.END)
                self._invoice_period_entries[-1][1].insert(0, str(last_day))

        # 绑定年月切换事件
        self.invoice_month_combo.bind('<<ComboboxSelected>>', lambda e: _on_ym_changed())
        self.invoice_year_combo.bind('<<ComboboxSelected>>', lambda e: _on_ym_changed())

        # 初始化：加载当前年月的分段，没有则默认4行
        ym_init = f"{self.invoice_year_var.get()}-{self.invoice_month_var.get()}"
        invoice_periods_init = self.config.get('invoice_periods', {})
        if ym_init in invoice_periods_init:
            _apply_periods(invoice_periods_init[ym_init])
        else:
            _apply_periods(_get_default_periods())

        # 导出按钮行
        invoice_btn_frame = tk.Frame(invoice_control, bg='#f5f5f5')
        invoice_btn_frame.pack(fill=tk.X, pady=(10, 0))

        self.invoice_status_var = tk.StringVar(value="")
        tk.Label(invoice_btn_frame, textvariable=self.invoice_status_var,
                 font=('Microsoft YaHei UI', 10), bg='#f5f5f5', fg='#1890ff').pack(side=tk.LEFT)

        tk.Button(invoice_btn_frame, text="开始导出 Invoice", font=('Microsoft YaHei UI', 11, 'bold'),
                  bg='#1890ff', fg='white', relief='flat', padx=20, pady=6,
                  cursor='hand2', command=self.export_invoice_excel).pack(side=tk.RIGHT)

        # 结果文本框
        invoice_result_frame = tk.Frame(invoice_card, bg='#e6f7ff', padx=20, pady=15)
        invoice_result_frame.pack(fill=tk.X)

        invoice_result_title = tk.Label(
            invoice_result_frame,
            text="📋 导出结果",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#e6f7ff',
            fg='#262626'
        )
        invoice_result_title.pack(anchor='w', pady=(0, 10))

        self.invoice_result_text = tk.Text(
            invoice_result_frame,
            height=6,
            font=('Consolas', 10),
            bg='#ffffff',
            fg='#262626',
            relief='solid',
            bd=1,
            padx=10,
            pady=10
        )
        self.invoice_result_text.pack(fill=tk.X)
        self.invoice_result_text.insert(tk.END, "选择客户、输出目录和周分段日期后，点击「开始导出 Invoice」...")
        self.invoice_result_text.configure(state='disabled')

    def show_driver_mgmt(self):
        """显示司机管理页面 - 司机月结单导出"""
        canvas = tk.Canvas(self.main_content, bg='#f0f2f5', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f2f5')
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=1180)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        title_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))
        tk.Label(title_frame, text="司机管理", font=('Microsoft YaHei UI', 20, 'bold'), bg='#f0f2f5', fg='#262626').pack(side=tk.LEFT)

        export_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        export_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        card_header = tk.Frame(export_card, bg='#ffffff', padx=20, pady=15)
        card_header.pack(fill=tk.X)
        tk.Label(card_header, text="🚗 司机月结单导出", font=('Microsoft YaHei UI', 14, 'bold'), bg='#ffffff', fg='#262626').pack(side=tk.LEFT)

        desc_frame = tk.Frame(export_card, bg='#ffffff', padx=20, pady=5)
        desc_frame.pack(fill=tk.X)
        tk.Label(desc_frame, text="导出指定月份的司机月结单Excel文件，每位司机一个独立文件（如：楊洲溢202601.xlsx）。", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#8c8c8c', wraplength=1000).pack(anchor='w')

        control_frame = tk.Frame(export_card, bg='#ffffff', padx=20, pady=10)
        control_frame.pack(fill=tk.X)

        date_frame = tk.Frame(control_frame, bg='#ffffff')
        date_frame.pack(side=tk.LEFT)
        tk.Label(date_frame, text="选择年月:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        from datetime import datetime as _dt
        current_date = _dt.now()
        year_list = [str(y) for y in range(2003, 2051)]
        self.driver_year_var = tk.StringVar(value=str(current_date.year))
        self.driver_year_combo = ttk.Combobox(date_frame, textvariable=self.driver_year_var, values=year_list, width=8, state='readonly', font=('Microsoft YaHei UI', 11))
        try:
            self.driver_year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.driver_year_combo.current(0)
        self.driver_year_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(date_frame, text="年", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 15))

        month_list = [f"{m:02d}" for m in range(1, 13)]
        self.driver_month_var = tk.StringVar(value=f"{current_date.month:02d}")
        self.driver_month_combo = ttk.Combobox(date_frame, textvariable=self.driver_month_var, values=month_list, width=6, state='readonly', font=('Microsoft YaHei UI', 11))
        try:
            self.driver_month_combo.current(month_list.index(f"{current_date.month:02d}"))
        except ValueError:
            self.driver_month_combo.current(0)
        self.driver_month_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(date_frame, text="月", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 20))

        self.driver_export_btn = tk.Button(
            control_frame, text="📤 导出司机月结单",
            font=('Microsoft YaHei UI', 11, 'bold'), bg='#1890ff', fg='white',
            relief='flat', padx=20, pady=6, cursor='hand2',
            command=self.export_driver_monthly_excel
        )
        self.driver_export_btn.pack(side=tk.LEFT, padx=(0, 10))

        # 导出范围选择
        self.driver_export_mode_var = tk.StringVar(value='all')
        tk.Radiobutton(control_frame, text="全部司机", variable=self.driver_export_mode_var, value='all',
                       font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#262626',
                       activebackground='#ffffff', cursor='hand2').pack(side=tk.LEFT)
        tk.Radiobutton(control_frame, text="选择司机", variable=self.driver_export_mode_var, value='select',
                       font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#262626',
                       activebackground='#ffffff', cursor='hand2').pack(side=tk.LEFT, padx=(2, 20))

        self.driver_export_status_var = tk.StringVar(value="就绪")
        tk.Label(control_frame, textvariable=self.driver_export_status_var, font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#52c41a').pack(side=tk.LEFT)

        dir_frame = tk.Frame(export_card, bg='#ffffff', padx=20, pady=5)
        dir_frame.pack(fill=tk.X)
        tk.Label(dir_frame, text="导出目录:", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 8))
        self.driver_export_dir_var = tk.StringVar(value=self.config.get('driver_export_dir', r'X:\司机月结单'))
        tk.Entry(dir_frame, textvariable=self.driver_export_dir_var, font=('Microsoft YaHei UI', 10), width=60, relief='solid').pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(dir_frame, text="浏览...", font=('Microsoft YaHei UI', 10), bg='#f0f0f0', fg='#262626', relief='flat', padx=10, pady=3, cursor='hand2', command=self.browse_driver_export_dir).pack(side=tk.LEFT)

        sep = tk.Frame(export_card, bg='#f0f0f0', height=1)
        sep.pack(fill=tk.X, padx=20, pady=(10, 0))

        log_frame = tk.Frame(export_card, bg='#ffffff', padx=20, pady=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        tk.Label(log_frame, text="导出日志:", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#595959').pack(anchor='w', pady=(0, 5))
        log_inner = tk.Frame(log_frame, bg='#ffffff')
        log_inner.pack(fill=tk.BOTH, expand=True)
        self.driver_export_result_text = tk.Text(log_inner, font=('Consolas', 9), bg='#fafafa', fg='#262626', relief='solid', bd=1, height=18, state='disabled', wrap='word')
        log_scroll = ttk.Scrollbar(log_inner, orient='vertical', command=self.driver_export_result_text.yview)
        self.driver_export_result_text.configure(yscrollcommand=log_scroll.set)
        self.driver_export_result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # ── 司机运杂费汇总表导出 ──
        summary_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        summary_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        sc_header = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=15)
        sc_header.pack(fill=tk.X)
        tk.Label(sc_header, text="📊 司机运杂费汇总表导出", font=('Microsoft YaHei UI', 14, 'bold'), bg='#ffffff', fg='#262626').pack(side=tk.LEFT)

        sc_desc = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=5)
        sc_desc.pack(fill=tk.X)
        tk.Label(sc_desc, text="导出指定月份的司机运杂费汇总统计表Excel（一个文件，含自營車/平板車/街車分组汇总）。", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#8c8c8c', wraplength=1000).pack(anchor='w')

        sc_ctrl = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=10)
        sc_ctrl.pack(fill=tk.X)

        # 第一行：年月 + 导出按钮
        sc_date = tk.Frame(sc_ctrl, bg='#ffffff')
        sc_date.pack(fill=tk.X, pady=(0, 8))

        tk.Label(sc_date, text="选择年月:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        self.summary_year_var = tk.StringVar(value=str(current_date.year))
        self.summary_year_combo = ttk.Combobox(sc_date, textvariable=self.summary_year_var, values=year_list, width=8, state='readonly', font=('Microsoft YaHei UI', 11))
        try:
            self.summary_year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.summary_year_combo.current(0)
        self.summary_year_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(sc_date, text="年", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 15))

        self.summary_month_var = tk.StringVar(value=f"{current_date.month:02d}")
        self.summary_month_combo = ttk.Combobox(sc_date, textvariable=self.summary_month_var, values=month_list, width=6, state='readonly', font=('Microsoft YaHei UI', 11))
        try:
            self.summary_month_combo.current(month_list.index(f"{current_date.month:02d}"))
        except ValueError:
            self.summary_month_combo.current(0)
        self.summary_month_combo.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(sc_date, text="月", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 20))

        self.driver_summary_btn = tk.Button(
            sc_date, text="📊 导出司机运杂费汇总表",
            font=('Microsoft YaHei UI', 11, 'bold'), bg='#52c41a', fg='white',
            relief='flat', padx=20, pady=6, cursor='hand2',
            command=self.export_driver_summary_excel
        )
        self.driver_summary_btn.pack(side=tk.LEFT, padx=(0, 20))

        self.driver_summary_status_var = tk.StringVar(value="就绪")
        tk.Label(sc_date, textvariable=self.driver_summary_status_var, font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#52c41a').pack(side=tk.LEFT)

        # 第二行：分成比例设置
        pct_frame = tk.Frame(sc_ctrl, bg='#f6f8fa', bd=1, relief='groove')
        pct_frame.pack(fill=tk.X, pady=(5, 0))

        pct_title = tk.Frame(pct_frame, bg='#f6f8fa', padx=15, pady=8)
        pct_title.pack(fill=tk.X)
        tk.Label(pct_title, text="分成比例设置（公司% / 司机%，应用于運費合計）", font=('Microsoft YaHei UI', 10, 'bold'), bg='#f6f8fa', fg='#262626').pack(side=tk.LEFT)

        pct_inputs = tk.Frame(pct_frame, bg='#f6f8fa', padx=15, pady=5)
        pct_inputs.pack(fill=tk.X)

        # 营运拖头车 (编号 1-799)
        tk.Label(pct_inputs, text="营运拖头车(1-799):", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#595959').pack(side=tk.LEFT, padx=(0, 5))
        tk.Label(pct_inputs, text="公司", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        self.pct_tuo_company_var = tk.StringVar(value=self.config.get('pct_tuo_company', ''))
        self.pct_tuo_company_entry = tk.Entry(pct_inputs, textvariable=self.pct_tuo_company_var, width=6, font=('Microsoft YaHei UI', 10), justify='center')
        self.pct_tuo_company_entry.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(pct_inputs, text="%", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        tk.Label(pct_inputs, text=" 司机", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        self.pct_tuo_driver_var = tk.StringVar(value=self.config.get('pct_tuo_driver', ''))
        self.pct_tuo_driver_entry = tk.Entry(pct_inputs, textvariable=self.pct_tuo_driver_var, width=6, font=('Microsoft YaHei UI', 10), justify='center')
        self.pct_tuo_driver_entry.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(pct_inputs, text="%", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT, padx=(0, 25))

        # 营运吨车 (编号 800-899)
        tk.Label(pct_inputs, text="营运吨车(800-899):", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#595959').pack(side=tk.LEFT, padx=(0, 5))
        tk.Label(pct_inputs, text="公司", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        self.pct_dun_company_var = tk.StringVar(value=self.config.get('pct_dun_company', ''))
        self.pct_dun_company_entry = tk.Entry(pct_inputs, textvariable=self.pct_dun_company_var, width=6, font=('Microsoft YaHei UI', 10), justify='center')
        self.pct_dun_company_entry.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(pct_inputs, text="%", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        tk.Label(pct_inputs, text=" 司机", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        self.pct_dun_driver_var = tk.StringVar(value=self.config.get('pct_dun_driver', ''))
        self.pct_dun_driver_entry = tk.Entry(pct_inputs, textvariable=self.pct_dun_driver_var, width=6, font=('Microsoft YaHei UI', 10), justify='center')
        self.pct_dun_driver_entry.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(pct_inputs, text="%", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT, padx=(0, 25))

        # 平板车 (编号 900-999)
        tk.Label(pct_inputs, text="平板车(900-999):", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#595959').pack(side=tk.LEFT, padx=(0, 5))
        tk.Label(pct_inputs, text="公司", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        self.pct_flat_company_var = tk.StringVar(value=self.config.get('pct_flat_company', ''))
        self.pct_flat_company_entry = tk.Entry(pct_inputs, textvariable=self.pct_flat_company_var, width=6, font=('Microsoft YaHei UI', 10), justify='center')
        self.pct_flat_company_entry.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(pct_inputs, text="%", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        tk.Label(pct_inputs, text=" 司机", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)
        self.pct_flat_driver_var = tk.StringVar(value=self.config.get('pct_flat_driver', ''))
        self.pct_flat_driver_entry = tk.Entry(pct_inputs, textvariable=self.pct_flat_driver_var, width=6, font=('Microsoft YaHei UI', 10), justify='center')
        self.pct_flat_driver_entry.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(pct_inputs, text="%", font=('Microsoft YaHei UI', 10), bg='#f6f8fa', fg='#8c8c8c').pack(side=tk.LEFT)

        # 第三行：导出目录设置
        dir_frame = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=5)
        dir_frame.pack(fill=tk.X)
        tk.Label(dir_frame, text="导出目录:", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 8))
        self.driver_summary_dir_var = tk.StringVar(value=self.config.get('driver_summary_dir', r'X:\司机汇总表'))
        tk.Entry(dir_frame, textvariable=self.driver_summary_dir_var, font=('Microsoft YaHei UI', 10), width=60, relief='solid').pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(dir_frame, text="浏览...", font=('Microsoft YaHei UI', 10), bg='#f0f0f0', fg='#262626', relief='flat', padx=10, pady=3, cursor='hand2', command=self.browse_driver_summary_dir).pack(side=tk.LEFT)

        sc_sep = tk.Frame(summary_card, bg='#f0f0f0', height=1)
        sc_sep.pack(fill=tk.X, padx=20, pady=(10, 0))

        sc_log = tk.Frame(summary_card, bg='#ffffff', padx=20, pady=10)
        sc_log.pack(fill=tk.BOTH, expand=True)
        tk.Label(sc_log, text="导出日志:", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#595959').pack(anchor='w', pady=(0, 5))
        sc_log_inner = tk.Frame(sc_log, bg='#ffffff')
        sc_log_inner.pack(fill=tk.BOTH, expand=True)
        self.driver_summary_log_text = tk.Text(sc_log_inner, font=('Consolas', 9), bg='#fafafa', fg='#262626', relief='solid', bd=1, height=10, state='disabled', wrap='word')
        sc_log_scroll = ttk.Scrollbar(sc_log_inner, orient='vertical', command=self.driver_summary_log_text.yview)
        self.driver_summary_log_text.configure(yscrollcommand=sc_log_scroll.set)
        self.driver_summary_log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sc_log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # ── 司机业务量统计导出 ──
        biz_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        biz_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        biz_header = tk.Frame(biz_card, bg='#ffffff', padx=20, pady=15)
        biz_header.pack(fill=tk.X)
        tk.Label(biz_header, text="📋 司机业务量统计", font=('Microsoft YaHei UI', 14, 'bold'), bg='#ffffff', fg='#262626').pack(side=tk.LEFT)

        biz_desc = tk.Frame(biz_card, bg='#ffffff', padx=20, pady=5)
        biz_desc.pack(fill=tk.X)
        tk.Label(biz_desc, text="统计每个司机每月的业务量（票数、运费、杂费），可按月或按年导出为 Excel。", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#8c8c8c', wraplength=1000).pack(anchor='w')

        biz_ctrl = tk.Frame(biz_card, bg='#ffffff', padx=20, pady=10)
        biz_ctrl.pack(fill=tk.X)

        # 年份选择
        dbiz_date = tk.Frame(biz_ctrl, bg='#ffffff')
        dbiz_date.pack(side=tk.LEFT)

        tk.Label(dbiz_date, text="统计年份:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 6))

        self.dbiz_year_var = tk.StringVar(value=str(current_date.year))
        self.dbiz_year_combo = ttk.Combobox(dbiz_date, textvariable=self.dbiz_year_var, values=year_list, width=6, state='readonly', font=('Microsoft YaHei UI', 11))
        try:
            self.dbiz_year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.dbiz_year_combo.current(0)
        self.dbiz_year_combo.pack(side=tk.LEFT, padx=(0, 4))

        tk.Label(dbiz_date, text="年", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 16))

        # 月份选择（全年/单月）
        tk.Label(dbiz_date, text="统计月份:", font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 6))

        dbiz_month_list = ['全年'] + [f"{m}月" for m in range(1, 13)]
        self.dbiz_month_var = tk.StringVar(value='全年')
        self.dbiz_month_combo = ttk.Combobox(dbiz_date, textvariable=self.dbiz_month_var, values=dbiz_month_list, width=6, state='readonly', font=('Microsoft YaHei UI', 11))
        self.dbiz_month_combo.current(0)
        self.dbiz_month_combo.pack(side=tk.LEFT, padx=(0, 20))

        # 导出按钮
        self.dbiz_export_btn = tk.Button(
            biz_ctrl, text="📋 导出司机业务量统计",
            font=('Microsoft YaHei UI', 11, 'bold'), bg='#722ed1', fg='white',
            relief='flat', padx=20, pady=6, cursor='hand2',
            command=self.export_driver_biz_stats_gui
        )
        self.dbiz_export_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.dbiz_status_var = tk.StringVar(value="就绪")
        tk.Label(biz_ctrl, textvariable=self.dbiz_status_var, font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#52c41a').pack(side=tk.LEFT)

        # 导出目录
        dbiz_dir_frame = tk.Frame(biz_card, bg='#ffffff', padx=20, pady=5)
        dbiz_dir_frame.pack(fill=tk.X)
        tk.Label(dbiz_dir_frame, text="导出目录:", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 8))
        self.dbiz_dir_var = tk.StringVar(value=self.config.get('driver_biz_stats_dir', r'X:\司机业务量'))
        tk.Entry(dbiz_dir_frame, textvariable=self.dbiz_dir_var, font=('Microsoft YaHei UI', 10), width=60, relief='solid').pack(side=tk.LEFT, padx=(0, 8))

        def browse_dbiz_dir():
            from tkinter import filedialog
            d = filedialog.askdirectory(title="选择导出目录")
            if d:
                self.dbiz_dir_var.set(d)
                self.config['driver_biz_stats_dir'] = d
                self.save_config()

        tk.Button(dbiz_dir_frame, text="浏览...", font=('Microsoft YaHei UI', 10), bg='#f0f0f0', fg='#262626', relief='flat', padx=10, pady=3, cursor='hand2', command=browse_dbiz_dir).pack(side=tk.LEFT)

        # 日志
        dbiz_sep = tk.Frame(biz_card, bg='#f0f0f0', height=1)
        dbiz_sep.pack(fill=tk.X, padx=20, pady=(10, 0))

        dbiz_log = tk.Frame(biz_card, bg='#ffffff', padx=20, pady=10)
        dbiz_log.pack(fill=tk.BOTH, expand=True)
        tk.Label(dbiz_log, text="导出日志:", font=('Microsoft YaHei UI', 10), bg='#ffffff', fg='#595959').pack(anchor='w', pady=(0, 5))
        dbiz_log_inner = tk.Frame(dbiz_log, bg='#ffffff')
        dbiz_log_inner.pack(fill=tk.BOTH, expand=True)
        self.dbiz_log_text = tk.Text(dbiz_log_inner, font=('Consolas', 9), bg='#fafafa', fg='#262626', relief='solid', bd=1, height=10, state='disabled', wrap='word')
        dbiz_log_scroll = ttk.Scrollbar(dbiz_log_inner, orient='vertical', command=self.dbiz_log_text.yview)
        self.dbiz_log_text.configure(yscrollcommand=dbiz_log_scroll.set)
        self.dbiz_log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        dbiz_log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    def append_driver_summary_log(self, text):
        """向司机汇总表导出日志框追加文字"""
        self.driver_summary_log_text.configure(state='normal')
        self.driver_summary_log_text.insert(tk.END, text + '\n')
        self.driver_summary_log_text.see(tk.END)
        self.driver_summary_log_text.configure(state='disabled')
        self.root.update_idletasks()

    def show_multiselect_dialog(self, title, items, label_func=None):
        """弹出多选列表对话框，返回选中的条目列表（用户取消则返回None）
        items: [(key, display_text), ...]
        label_func: 可选，传入item生成显示文字的函数
        """
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.resizable(True, True)
        dialog.grab_set()
        dialog.focus_set()

        # 计算居中位置
        dialog.update_idletasks()
        w, h = 480, 520
        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        dialog.geometry(f"{w}x{h}+{x}+{y}")
        dialog.configure(bg='#f0f2f5')

        # 标题
        tk.Label(dialog, text=title, font=('Microsoft YaHei UI', 12, 'bold'),
                 bg='#f0f2f5', fg='#262626').pack(pady=(15, 5))
        tk.Label(dialog, text="按住 Ctrl 点击可多选，按住 Shift 可范围选择",
                 font=('Microsoft YaHei UI', 9), bg='#f0f2f5', fg='#8c8c8c').pack(pady=(0, 8))

        # 全选/全不选按钮行
        btn_row = tk.Frame(dialog, bg='#f0f2f5')
        btn_row.pack(fill=tk.X, padx=20, pady=(0, 5))

        list_frame = tk.Frame(dialog, bg='#f0f2f5')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20)

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical')
        listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE,
                             font=('Microsoft YaHei UI', 10),
                             bg='#ffffff', fg='#262626',
                             selectbackground='#1890ff', selectforeground='white',
                             relief='solid', bd=1, activestyle='none',
                             yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 填充数据
        keys = []
        for item in items:
            key, display = item
            keys.append(key)
            listbox.insert(tk.END, display)

        def select_all():
            listbox.select_set(0, tk.END)
        def select_none():
            listbox.selection_clear(0, tk.END)

        tk.Button(btn_row, text="全选", font=('Microsoft YaHei UI', 9),
                  bg='#f0f0f0', relief='flat', padx=10, pady=2, cursor='hand2',
                  command=select_all).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(btn_row, text="全不选", font=('Microsoft YaHei UI', 9),
                  bg='#f0f0f0', relief='flat', padx=10, pady=2, cursor='hand2',
                  command=select_none).pack(side=tk.LEFT)

        # 底部确认/取消
        result = {'selected': None}

        def on_confirm():
            sel_indices = listbox.curselection()
            result['selected'] = [keys[i] for i in sel_indices]
            dialog.destroy()

        def on_cancel():
            result['selected'] = None
            dialog.destroy()

        btn_frame = tk.Frame(dialog, bg='#f0f2f5')
        btn_frame.pack(fill=tk.X, padx=20, pady=15)
        tk.Button(btn_frame, text="确定", font=('Microsoft YaHei UI', 10, 'bold'),
                  bg='#1890ff', fg='white', relief='flat', padx=20, pady=5,
                  cursor='hand2', command=on_confirm).pack(side=tk.RIGHT, padx=(5, 0))
        tk.Button(btn_frame, text="取消", font=('Microsoft YaHei UI', 10),
                  bg='#f0f0f0', fg='#262626', relief='flat', padx=20, pady=5,
                  cursor='hand2', command=on_cancel).pack(side=tk.RIGHT)

        dialog.wait_window()
        return result['selected']

    def browse_driver_export_dir(self):
        """选择司机月结单导出目录"""
        from tkinter import filedialog
        selected_dir = filedialog.askdirectory(title="选择司机月结单导出目录", initialdir=self.driver_export_dir_var.get())
        if selected_dir:
            self.driver_export_dir_var.set(selected_dir)
            self.config['driver_export_dir'] = selected_dir
            self.save_config()

    def browse_driver_summary_dir(self):
        """选择司机汇总表导出目录"""
        from tkinter import filedialog
        selected_dir = filedialog.askdirectory(title="选择司机汇总表导出目录", initialdir=self.driver_summary_dir_var.get())
        if selected_dir:
            self.driver_summary_dir_var.set(selected_dir)
            self.config['driver_summary_dir'] = selected_dir
            self.save_config()

    def append_driver_export_result(self, text):
        """向司机月结单导出日志框追加文字"""
        self.driver_export_result_text.configure(state='normal')
        self.driver_export_result_text.insert(tk.END, text + '\n')
        self.driver_export_result_text.see(tk.END)
        self.driver_export_result_text.configure(state='disabled')
        self.root.update_idletasks()

    def export_driver_monthly_excel(self):
        """导出司机月结单Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            from datetime import datetime, timedelta
            import os, re

            year = int(self.driver_year_var.get())
            month = int(self.driver_month_combo.get())
            if not (1 <= month <= 12):
                messagebox.showerror("错误", "月份必须在 1-12 之间")
                return

            yyyymm = f"{year}{month:02d}"
            year_cn = f"{year}年{month}月"

            # 清空日志
            self.driver_export_result_text.configure(state='normal')
            self.driver_export_result_text.delete(1.0, tk.END)
            self.driver_export_result_text.configure(state='disabled')

            self.driver_export_status_var.set(f"正在导出 {year_cn}...")
            self.root.update()

            # 日期范围
            date_start = f'{year}-{month:02d}-01'
            if month == 12:
                next_month_dt = datetime(year + 1, 1, 1)
            else:
                next_month_dt = datetime(year, month + 1, 1)
            date_end = next_month_dt.strftime('%Y-%m-%d')
            date_end_display = (next_month_dt - timedelta(days=1)).strftime('%Y-%m-%d')

            self.append_driver_export_result(f"{'='*60}")
            self.append_driver_export_result(f"司机月结单导出 - {year_cn}")
            self.append_driver_export_result(f"{'='*60}\n")
            self.append_driver_export_result(f"日期范围: {date_start} 至 {date_end_display}")

            # 步骤1：生成 司机月结单TEMP（使用已有的"司机月结单"查询SQL）
            self.append_driver_export_result(f"\n步骤1/4: 生成司机月结单TEMP...")
            self.cursor.execute("DROP TABLE IF EXISTS `司机月结单TEMP`")

            # 先创建临时表（含运费司机 + 只有杂费司机）
            self.cursor.execute("DROP TEMPORARY TABLE IF EXISTS driver_stmt_temp")


            # 匹配规则：CON_DETAIL.DRIVER = 司機編號，直接关联 DRIVER_CP.CUSTCODE
            # 对于 cd.DRIVER 在 DRIVER_CP 中找不到的情况，回退到車牌匹配（cd.DRIVERCODE 匹配 dc.HKCP）
            # 三段式：
            #   第1段：CON_DETAIL运费，用 cd.DRIVER 直接关联 DRIVER_CP
            #   第2段：CON_DETAIL运费，cd.DRIVER 找不到时回退到車牌匹配
            #   第3段：只有雜费、没有运费的司机（INVOICE_DETAIL但不在第1/2段中）
            sql_driver_temp = """
            CREATE TEMPORARY TABLE driver_stmt_temp AS
            -- 第1段：CON_DETAIL运费，用 cd.DRIVER 直接关联 DRIVER_CP.CUSTCODE
            SELECT
                im.INVOICECODE,
                im.INVDATE,
                im.CUSTCODE,
                (SELECT NAME FROM CUST_MASTER WHERE CUSTCODE = im.CUSTCODE LIMIT 1) AS CUSTNAME,
                im.DEST,
                im.SHIPCODE,
                cd.NN AS con_nn,
                cd.DRIVERCODE AS hkcp,
                cd.DRIVERCOMM AS run_fee,
                cd.SIZE,
                cd.CONCODE,
                cd.TAKENO,
                dc.CUSTCODE AS driver_code,
                dc.NAME AS driver_name,
                dc.HKCP AS dc_hkcp,
                dc.SZCP,
                GROUP_CONCAT(
                    CASE WHEN id.DESCR IS NOT NULL AND id.DESCR != '' THEN TRIM(TRAILING ',' FROM TRIM(TRAILING '.' FROM TRIM(id.DESCR))) END
                    ORDER BY id.NN SEPARATOR '\xa7'
                ) AS dp,
                GROUP_CONCAT(
                    CASE WHEN id.PRICE IS NOT NULL THEN CAST(id.PRICE AS CHAR) END
                    ORDER BY id.NN SEPARATOR '\xa7'
                ) AS pp
            FROM CON_DETAIL cd
            JOIN INVOICE_MASTER im ON im.INVOICECODE = cd.INVOICECODE
            JOIN DRIVER_CP dc ON dc.CUSTCODE = cd.DRIVER
            LEFT JOIN INVOICE_DETAIL id ON id.INVOICECODE = cd.INVOICECODE
                AND id.DRIVERCODE = dc.CUSTCODE
            WHERE im.INVDATE >= %s AND im.INVDATE < %s
            GROUP BY cd.INVOICECODE, cd.NN, im.INVDATE, im.CUSTCODE, im.DEST, im.SHIPCODE,
                     cd.DRIVERCODE, cd.DRIVERCOMM, cd.SIZE, cd.CONCODE, cd.TAKENO,
                     dc.CUSTCODE, dc.NAME, dc.HKCP, dc.SZCP

            UNION ALL

            -- 第2段：cd.DRIVER 在 DRIVER_CP 中找不到的记录，回退到車牌匹配
            -- 用 LEFT JOIN 子查询获取 max_star_count（MySQL 5.6 不支持 ON 子句中的相关子查询，移到 WHERE）
            SELECT
                im.INVOICECODE,
                im.INVDATE,
                im.CUSTCODE,
                (SELECT NAME FROM CUST_MASTER WHERE CUSTCODE = im.CUSTCODE LIMIT 1) AS CUSTNAME,
                im.DEST,
                im.SHIPCODE,
                cd.NN AS con_nn,
                cd.DRIVERCODE AS hkcp,
                cd.DRIVERCOMM AS run_fee,
                cd.SIZE,
                cd.CONCODE,
                cd.TAKENO,
                dc.CUSTCODE AS driver_code,
                dc.NAME AS driver_name,
                dc.HKCP AS dc_hkcp,
                dc.SZCP,
                GROUP_CONCAT(
                    CASE WHEN id.DESCR IS NOT NULL AND id.DESCR != '' THEN TRIM(TRAILING ',' FROM TRIM(TRAILING '.' FROM TRIM(id.DESCR))) END
                    ORDER BY id.NN SEPARATOR '\xa7'
                ) AS dp,
                GROUP_CONCAT(
                    CASE WHEN id.PRICE IS NOT NULL THEN CAST(id.PRICE AS CHAR) END
                    ORDER BY id.NN SEPARATOR '\xa7'
                ) AS pp
            FROM CON_DETAIL cd
            JOIN INVOICE_MASTER im ON im.INVOICECODE = cd.INVOICECODE
            JOIN (
                SELECT
                    REPLACE(REPLACE(HKCP,'*',''),' ','') AS norm_hkcp,
                    CUSTCODE,
                    NAME,
                    HKCP,
                    SZCP,
                    CHAR_LENGTH(HKCP) - CHAR_LENGTH(REPLACE(HKCP,'*','')) AS star_count
                FROM DRIVER_CP
            ) dc ON dc.norm_hkcp = REPLACE(REPLACE(cd.DRIVERCODE,'*',''),' ','')
            LEFT JOIN (
                SELECT norm_hkcp, MAX(star_count) AS max_sc FROM (
                    SELECT
                        REPLACE(REPLACE(HKCP,'*',''),' ','') AS norm_hkcp,
                        CHAR_LENGTH(HKCP) - CHAR_LENGTH(REPLACE(HKCP,'*','')) AS star_count
                    FROM DRIVER_CP dcp_inner
                    WHERE EXISTS (
                        SELECT 1 FROM INVOICE_DETAIL sc_id
                        JOIN INVOICE_MASTER sc_im ON sc_id.INVOICECODE = sc_im.INVOICECODE
                        WHERE sc_im.INVDATE >= %s AND sc_im.INVDATE < %s
                        AND sc_id.DRIVERCODE = dcp_inner.CUSTCODE
                    )
                ) t1 GROUP BY norm_hkcp
            ) active_max ON active_max.norm_hkcp = dc.norm_hkcp
            LEFT JOIN (
                SELECT norm_hkcp, MAX(star_count) AS max_sc FROM (
                    SELECT
                        REPLACE(REPLACE(HKCP,'*',''),' ','') AS norm_hkcp,
                        CHAR_LENGTH(HKCP) - CHAR_LENGTH(REPLACE(HKCP,'*','')) AS star_count
                    FROM DRIVER_CP
                ) t2 GROUP BY norm_hkcp
            ) all_max ON all_max.norm_hkcp = dc.norm_hkcp
            LEFT JOIN INVOICE_DETAIL id ON id.INVOICECODE = cd.INVOICECODE
                AND id.DRIVERCODE = dc.CUSTCODE
            WHERE im.INVDATE >= %s AND im.INVDATE < %s
              AND NOT EXISTS (
                  SELECT 1 FROM DRIVER_CP dc_direct WHERE dc_direct.CUSTCODE = cd.DRIVER
              )
              AND dc.star_count = COALESCE(active_max.max_sc, all_max.max_sc)

            UNION ALL

            -- 第3段：只有雜费、没有运费的司机（INVOICE_DETAIL但不在第1/2段中）
            SELECT
                im.INVOICECODE,
                im.INVDATE,
                im.CUSTCODE,
                (SELECT NAME FROM CUST_MASTER WHERE CUSTCODE = im.CUSTCODE LIMIT 1) AS CUSTNAME,
                im.DEST,
                im.SHIPCODE,
                NULL AS con_nn,
                dc.HKCP AS hkcp,
                0 AS run_fee,
                NULL AS SIZE,
                NULL AS CONCODE,
                NULL AS TAKENO,
                id_grp.DRIVERCODE AS driver_code,
                dc.NAME AS driver_name,
                dc.HKCP AS dc_hkcp,
                dc.SZCP,
                GROUP_CONCAT(
                    CASE WHEN id_grp.DESCR IS NOT NULL AND id_grp.DESCR != '' THEN TRIM(TRAILING ',' FROM TRIM(TRAILING '.' FROM TRIM(id_grp.DESCR))) END
                    ORDER BY id_grp.NN SEPARATOR '\xa7'
                ) AS dp,
                GROUP_CONCAT(
                    CASE WHEN id_grp.PRICE IS NOT NULL THEN CAST(id_grp.PRICE AS CHAR) END
                    ORDER BY id_grp.NN SEPARATOR '\xa7'
                ) AS pp
            FROM INVOICE_DETAIL id_grp
            JOIN INVOICE_MASTER im ON im.INVOICECODE = id_grp.INVOICECODE
            LEFT JOIN DRIVER_CP dc ON dc.CUSTCODE = id_grp.DRIVERCODE
            WHERE im.INVDATE >= %s AND im.INVDATE < %s
              AND dc.CUSTCODE IS NOT NULL
              AND (im.INVOICECODE, id_grp.DRIVERCODE) NOT IN (
                  SELECT cd2.INVOICECODE, cd2.DRIVER
                  FROM CON_DETAIL cd2
                  WHERE cd2.DRIVER IS NOT NULL AND cd2.DRIVER != ''
              )
            GROUP BY im.INVOICECODE, im.INVDATE, im.CUSTCODE, im.DEST, im.SHIPCODE,
                     id_grp.DRIVERCODE, dc.NAME, dc.HKCP, dc.SZCP
            """
            self.cursor.execute(sql_driver_temp, (date_start, date_end, date_start, date_end, date_start, date_end, date_start, date_end))

            # 创建正式TEMP表
            sql_temp = """
            CREATE TABLE `司机月结单TEMP` (
                `id` INT UNSIGNED NOT NULL AUTO_INCREMENT,
                `日期` DATE,
                `發票號` VARCHAR(15) NOT NULL,
                `客戶編號` VARCHAR(10),
                `客戶名稱` VARCHAR(40),
                `司機編號` VARCHAR(10),
                `司機姓名` VARCHAR(40),
                `香港車牌` VARCHAR(60),
                `大陸車牌` VARCHAR(60),
                `地區` VARCHAR(40),
                `櫃號` VARCHAR(200),
                `櫃尺碼` VARCHAR(20),
                `提單號` VARCHAR(40),
                `船名` VARCHAR(30),
                `托運號` VARCHAR(40),
                `運費` DECIMAL(10,2),
                `雜費名稱1` LONGTEXT, `雜費金額1` DECIMAL(12,2),
                `雜費名稱2` LONGTEXT, `雜費金額2` DECIMAL(12,2),
                `雜費名稱3` LONGTEXT, `雜費金額3` DECIMAL(12,2),
                `雜費名稱4` LONGTEXT, `雜費金額4` DECIMAL(12,2),
                `雜費名稱5` LONGTEXT, `雜費金額5` DECIMAL(12,2),
                `雜費名稱6` LONGTEXT, `雜費金額6` DECIMAL(12,2),
                `雜費名稱7` LONGTEXT, `雜費金額7` DECIMAL(12,2),
                `雜費名稱8` LONGTEXT, `雜費金額8` DECIMAL(12,2),
                `雜費名稱9` LONGTEXT, `雜費金額9` DECIMAL(12,2),
                `雜費名稱10` LONGTEXT, `雜費金額10` DECIMAL(12,2),
                `雜費名稱11` LONGTEXT, `雜費金額11` DECIMAL(12,2),
                `雜費名稱12` LONGTEXT, `雜費金額12` DECIMAL(12,2),
                `雜費名稱13` LONGTEXT, `雜費金額13` DECIMAL(12,2),
                `雜費名稱14` LONGTEXT, `雜費金額14` DECIMAL(12,2),
                `雜費名稱15` LONGTEXT, `雜費金額15` DECIMAL(12,2),
                `雜費名稱16` LONGTEXT, `雜費金額16` DECIMAL(12,2),
                `雜費名稱17` LONGTEXT, `雜費金額17` DECIMAL(12,2),
                `雜費名稱18` LONGTEXT, `雜費金額18` DECIMAL(12,2),
                `雜費名稱19` LONGTEXT, `雜費金額19` DECIMAL(12,2),
                `雜費名稱20` LONGTEXT, `雜費金額20` DECIMAL(12,2),
                `運雜費合計` DECIMAL(33,2),
                PRIMARY KEY (`id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8
            """
            self.cursor.execute(sql_temp)

            # 从临时表插入数据（使用 § 分隔解析杂费，和客户月结单一致）
            sql_insert = """
            INSERT INTO `司机月结单TEMP`
                (`日期`, `發票號`, `客戶編號`, `客戶名稱`, `司機編號`, `司機姓名`,
                 `香港車牌`, `大陸車牌`, `地區`, `櫃號`, `櫃尺碼`, `提單號`,
                 `船名`, `托運號`, `運費`,
                 `雜費名稱1`, `雜費金額1`, `雜費名稱2`, `雜費金額2`,
                 `雜費名稱3`, `雜費金額3`, `雜費名稱4`, `雜費金額4`,
                 `雜費名稱5`, `雜費金額5`, `雜費名稱6`, `雜費金額6`,
                 `雜費名稱7`, `雜費金額7`, `雜費名稱8`, `雜費金額8`,
                 `雜費名稱9`, `雜費金額9`, `雜費名稱10`, `雜費金額10`,
                 `雜費名稱11`, `雜費金額11`, `雜費名稱12`, `雜費金額12`,
                 `雜費名稱13`, `雜費金額13`, `雜費名稱14`, `雜費金額14`,
                 `雜費名稱15`, `雜費金額15`, `雜費名稱16`, `雜費金額16`,
                 `雜費名稱17`, `雜費金額17`, `雜費名稱18`, `雜費金額18`,
                 `雜費名稱19`, `雜費金額19`, `雜費名稱20`, `雜費金額20`,
                 `運雜費合計`)
            SELECT
                dt.INVDATE AS `日期`,
                dt.INVOICECODE AS `發票號`,
                dt.CUSTCODE AS `客戶編號`,
                COALESCE(dt.CUSTNAME, '') AS `客戶名稱`,
                dt.driver_code AS `司機編號`,
                COALESCE(dt.driver_name, '') AS `司機姓名`,
                dt.dc_hkcp AS `香港車牌`,
                dt.SZCP AS `大陸車牌`,
                dt.DEST AS `地區`,
                dt.CONCODE AS `櫃號`,
                dt.SIZE AS `櫃尺碼`,
                dt.TAKENO AS `提單號`,
                NULL AS `船名`,
                dt.SHIPCODE AS `托運號`,
                dt.run_fee AS `運費`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 0,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 1), '\xa7', -1), '') AS `雜費名稱1`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 0,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 1), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額1`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 1,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 2), '\xa7', -1), '') AS `雜費名稱2`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 1,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 2), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額2`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 2,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 3), '\xa7', -1), '') AS `雜費名稱3`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 2,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 3), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額3`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 3,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 4), '\xa7', -1), '') AS `雜費名稱4`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 3,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 4), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額4`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 4,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 5), '\xa7', -1), '') AS `雜費名稱5`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 4,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 5), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額5`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 5,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 6), '\xa7', -1), '') AS `雜費名稱6`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 5,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 6), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額6`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 6,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 7), '\xa7', -1), '') AS `雜費名稱7`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 6,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 7), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額7`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 7,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 8), '\xa7', -1), '') AS `雜費名稱8`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 7,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 8), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額8`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 8,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 9), '\xa7', -1), '') AS `雜費名稱9`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 8,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 9), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額9`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 9,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 10), '\xa7', -1), '') AS `雜費名稱10`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 9,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 10), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額10`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 10,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 11), '\xa7', -1), '') AS `雜費名稱11`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 10,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 11), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額11`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 11,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 12), '\xa7', -1), '') AS `雜費名稱12`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 11,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 12), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額12`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 12,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 13), '\xa7', -1), '') AS `雜費名稱13`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 12,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 13), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額13`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 13,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 14), '\xa7', -1), '') AS `雜費名稱14`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 13,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 14), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額14`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 14,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 15), '\xa7', -1), '') AS `雜費名稱15`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 14,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 15), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額15`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 15,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 16), '\xa7', -1), '') AS `雜費名稱16`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 15,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 16), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額16`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 16,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 17), '\xa7', -1), '') AS `雜費名稱17`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 16,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 17), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額17`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 17,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 18), '\xa7', -1), '') AS `雜費名稱18`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 17,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 18), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額18`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 18,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 19), '\xa7', -1), '') AS `雜費名稱19`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 18,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 19), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額19`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 19,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(dt.dp, '\xa7', 20), '\xa7', -1), '') AS `雜費名稱20`,
                IF((CHAR_LENGTH(dt.dp) - CHAR_LENGTH(REPLACE(dt.dp, '\xa7', ''))) >= 19,
                    CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(dt.pp, '\xa7', 20), '\xa7', -1) AS DECIMAL(12,2)), NULL) AS `雜費金額20`,
                COALESCE(dt.run_fee, 0) AS `運雜費合計`
            FROM driver_stmt_temp dt
            ORDER BY dt.INVDATE, dt.driver_code, dt.con_nn
            """
            self.cursor.execute(sql_insert)
            temp_count = self.cursor.rowcount
            self.conn.commit()
            self.append_driver_export_result(f"  完成: {temp_count} 条记录")

            if temp_count == 0:
                self.append_driver_export_result(f"\n该月没有司机月结单数据，导出结束。")
                self.driver_export_status_var.set("无数据")
                return

            # 步骤2：归一化雜費关键字（代码内执行，不依赖存储过程）
            self.append_driver_export_result(f"\n步骤2/4: 归一化雜費关键字...")
            self._normalize_misc_names("司机月结单TEMP")
            self.conn.commit()
            self.append_driver_export_result(f"  完成")

            # 步骤3：pivot_driver_misc_names（转宽表）
            self.append_driver_export_result(f"\n步骤3/4: 生成司机月结单PIVOT...")
            self.cursor.execute("DROP TABLE IF EXISTS `司机月结单PIVOT`")
            self.cursor.execute("CALL pivot_driver_misc_names()")
            self.conn.commit()
            self.append_driver_export_result(f"  完成")

            # 步骤4：按司机导出Excel
            self.append_driver_export_result(f"\n步骤4/4: 导出Excel文件...")

            # 查询PIVOT表的所有列名（跳过 _id, id）
            self.cursor.execute("SELECT * FROM `司机月结单PIVOT` LIMIT 0")
            all_cols = [d[0] for d in self.cursor.description]
            skip_cols = {'_id', 'id'}
            data_cols = [c for c in all_cols if c not in skip_cols]
            # 固定列 + 动态雜费列（按指定顺序排列）
            fixed_col_names = [
                '日期', '發票號', '司機編號', '司機姓名', '香港車牌', '大陸車牌',
                '客戶名稱', '地區', '櫃號', '櫃尺碼', '托運號', '運費'
            ]
            # 按顺序提取固定列（仅保留PIVOT表中实际存在的列）
            fixed_cols = [c for c in fixed_col_names if c in data_cols]
            misc_cols = [c for c in data_cols if c not in fixed_col_names and c != '運雜費合計']
            has_total = '運雜費合計' in data_cols

            # 查询所有不重复司机（用司機編號+司機姓名组合去重，排除公司代墊和合計行）
            self.cursor.execute("SELECT DISTINCT `司機編號`, `司機姓名` FROM `司机月结单PIVOT` WHERE `司機姓名` IS NOT NULL AND `司機姓名` != '' AND `司機姓名` != '合計' AND `司機姓名` NOT LIKE '%%\r%%' AND `司機姓名` NOT LIKE '%%\n%%' ORDER BY `司機編號`")
            drivers_all = self.cursor.fetchall()

            # 若选择模式，弹出多选对话框
            if self.driver_export_mode_var.get() == 'select':
                items = []
                for d in drivers_all:
                    code = d['司機編號'] if isinstance(d, dict) else d[0]
                    name = d['司機姓名'] if isinstance(d, dict) else d[1]
                    items.append((code, f"{name}  [{code}]"))
                selected_codes = self.show_multiselect_dialog("选择要导出的司机", items)
                if selected_codes is None:
                    self.driver_export_status_var.set("已取消")
                    return
                if not selected_codes:
                    messagebox.showwarning("提示", "未选择任何司机，导出取消")
                    self.driver_export_status_var.set("已取消")
                    return
                selected_set = set(selected_codes)
                drivers = [d for d in drivers_all if (d['司機編號'] if isinstance(d, dict) else d[0]) in selected_set]
            else:
                drivers = drivers_all

            self.append_driver_export_result(f"  共 {len(drivers)} 位司机")

            # 统计同名司机的司機編號数量，用于文件名加 A/B/C 后缀
            name_code_map = {}  # {司機姓名: [司機編號1, 司機編號2, ...]}
            for drv in drivers:
                name = drv['司機姓名']
                if name not in name_code_map:
                    name_code_map[name] = []
                name_code_map[name].append(drv['司機編號'])

            # 导出目录：在用户选择的目录下创建 yyyymm 子目录，重复则加后缀 A/B/C...
            base_export_dir = self.driver_export_dir_var.get()
            if not os.path.exists(base_export_dir):
                os.makedirs(base_export_dir)
            # 生成不重复的子目录名
            subdir_name = yyyymm
            suffix_idx = 0
            while os.path.exists(os.path.join(base_export_dir, subdir_name)):
                suffix_idx += 1
                subdir_name = yyyymm + chr(ord('A') + suffix_idx - 1)
            export_dir = os.path.join(base_export_dir, subdir_name)
            os.makedirs(export_dir)

            num_fmt = '#,##0.00'
            success_count = 0

            # ── 辅助函数：提取司機編號的英文字母前缀 ──
            def get_alpha_prefix(code):
                """提取连续英文字母前缀，如 C1053->C, KT3248->KT, X5927->X"""
                prefix = ''
                for ch in (code or ''):
                    if ch.isalpha():
                        prefix += ch
                    else:
                        break
                return prefix

            # ── 辅助函数：将一个sheet的数据写入worksheet ──
            def write_driver_sheet(ws, rows, data_col_names, fixed_cols_in, misc_cols_in, has_total_in):
                """将司机数据写入一个worksheet"""
                active_misc_cols = [c for c in misc_cols_in if any(
                    r.get(c) is not None and r.get(c) != '' and r.get(c) != 0 and r.get(c) != 0.0
                    for r in rows
                )]
                header = list(fixed_cols_in) + active_misc_cols
                if has_total_in:
                    header.append('運雜費合計')
                num_cols_set = {'運費', '運雜費合計'} | set(active_misc_cols)

                # 表头行
                ws.row_dimensions[1].height = ROW_HEIGHT
                for col_idx, col_name in enumerate(header, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font      = font_header
                    cell.alignment = align_center
                    cell.border    = thin_border
                    if col_name in num_cols_set:
                        cell.number_format = num_fmt

                # 数据行
                data_row_count = len(rows)
                for row_idx, row in enumerate(rows, 2):
                    ws.row_dimensions[row_idx].height = ROW_HEIGHT
                    for col_idx, col_name in enumerate(header, 1):
                        if col_name not in data_col_names:
                            cell = ws.cell(row=row_idx, column=col_idx, value='')
                        else:
                            val = row.get(col_name, '')
                            if val is None:
                                val = ''
                            if col_name in active_misc_cols and (val == 0 or val == 0.0):
                                val = ''
                            if col_name == '日期' and hasattr(val, 'strftime'):
                                val = val.strftime('%Y-%m-%d')
                            if col_name in ('托運號', '發票號', '司機編號') and val != '':
                                val = str(val)
                            cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.font   = font_base
                        cell.border = thin_border
                        if col_name in num_cols_set:
                            cell.number_format = num_fmt
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_vcenter

                # 運雜費合計列（计算实际值）
                if has_total_in:
                    total_col = len(header)
                    freight_cols = ['運費'] + active_misc_cols
                    col_totals = {c: 0.0 for c in freight_cols}
                    grand_total = 0.0
                    for row_idx in range(2, 2 + data_row_count):
                        row = rows[row_idx - 2]
                        row_total = 0.0
                        for ac in freight_cols:
                            try:
                                v = float(row.get(ac, 0) or 0)
                                row_total += v
                                col_totals[ac] += v
                            except (ValueError, TypeError):
                                pass
                        grand_total += row_total
                        cell = ws.cell(row=row_idx, column=total_col)
                        cell.value         = row_total
                        cell.number_format = num_fmt
                        cell.font          = font_base
                        cell.border        = thin_border
                        cell.alignment     = align_right

                # 合計行
                summary_row = 2 + data_row_count
                ws.row_dimensions[summary_row].height = ROW_HEIGHT
                cell_a = ws.cell(row=summary_row, column=1, value='合計')
                cell_a.font      = font_total
                cell_a.border    = thin_border
                cell_a.alignment = align_vcenter
                freight_col_idx = fixed_cols_in.index('運費') + 1 if '運費' in fixed_cols_in else 2
                for col_idx in range(freight_col_idx, len(header) + 1):
                    col_name = header[col_idx - 1]
                    cell = ws.cell(row=summary_row, column=col_idx)
                    if col_name == '運雜費合計':
                        cell.value         = grand_total
                    else:
                        cell.value         = col_totals.get(col_name, 0.0)
                    cell.font          = font_total
                    cell.border        = thin_border
                    cell.number_format = num_fmt
                    cell.alignment     = align_right

                # 自适应列宽（含合计行）
                total_rows = data_row_count + 2  # 数据行 + 表头 + 合計行
                for col_idx, col_name in enumerate(header, 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for row_idx in range(1, total_rows + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            if isinstance(cell.value, str):
                                cell_len = len(str(cell.value))
                            else:
                                try:
                                    formatted = format(float(cell.value), ',.2f')
                                    cell_len = len(formatted)
                                except (ValueError, TypeError):
                                    cell_len = len(str(cell.value))
                            if cell_len > max_len:
                                max_len = cell_len
                    header_text = header[col_idx - 1]
                    header_display_len = sum(2 if ord(c) > 127 else 1 for c in header_text)
                    width = max(max_len, header_display_len)
                    ws.column_dimensions[col_letter].width = width * 0.75 + 2.5

                return data_row_count

            # ── 将司机分为：独立Excel（数字开头/X开头）、合并Excel（其他英文开头）──
            digit_drivers = []   # 独立Excel
            alpha_drivers = []   # 合并Excel
            for drv in drivers:
                code = drv['司機編號']
                prefix = get_alpha_prefix(code).upper() if code and code[0].isalpha() else ''
                if prefix == 'X':
                    # X开头的独立按司机导出
                    digit_drivers.append(drv)
                elif prefix:
                    alpha_drivers.append(drv)
                else:
                    digit_drivers.append(drv)

            # ── 英文开头司机按字母前缀分组 ──
            # 每个司機編號取完整英文字母前缀（如C1053→C, KT3248→KT），按前缀分组
            alpha_groups = {}  # {prefix: [driver_dicts]}
            for drv in alpha_drivers:
                code = drv['司機編號']
                prefix = get_alpha_prefix(code).upper()
                if prefix not in alpha_groups:
                    alpha_groups[prefix] = []
                alpha_groups[prefix].append(drv)

            # 按前缀排序，确保文件名 A/B/C 顺序一致
            sorted_prefixes = sorted(alpha_groups.keys())

            # ── 样式定义（参照参考Excel）──────────────────────────
            font_base   = Font(name='等线', size=11, bold=False)
            font_header = Font(name='等线', size=11, bold=True)
            font_total  = Font(name='等线', size=11, bold=True)
            thin_side   = Side(style='thin')
            thin_border = Border(left=thin_side, right=thin_side,
                                 top=thin_side, bottom=thin_side)
            align_center = Alignment(horizontal='center', vertical='center')
            align_vcenter = Alignment(vertical='center')
            align_right  = Alignment(horizontal='right', vertical='center')
            ROW_HEIGHT   = 23.25

            # ── 1) 先导出英文开头的合并Excel ──
            # 同一前缀的所有司机合并到同一个sheet，按日期排序
            for idx, prefix in enumerate(sorted_prefixes):
                group = alpha_groups[prefix]
                wb = Workbook()
                ws = wb.active
                ws.title = 'Sheet1'

                # 收集所有司机的数据行
                all_rows = []
                data_col_names = None
                driver_count = 0
                for drv in group:
                    driver_code = drv['司機編號']
                    driver_name = drv['司機姓名']
                    self.cursor.execute(
                        "SELECT * FROM `司机月结单PIVOT` WHERE `司機編號` = %s AND `司機姓名` = %s ORDER BY `日期`, `發票號`",
                        (driver_code, driver_name)
                    )
                    rows = self.cursor.fetchall()
                    if rows:
                        if data_col_names is None:
                            data_col_names = [d[0] for d in self.cursor.description]
                        all_rows.extend(rows)
                        driver_count += 1

                if not all_rows:
                    continue

                # 按日期、發票號排序（确保跨司机混合后也按日期排）
                all_rows.sort(key=lambda r: (
                    r.get('日期') or datetime.min,
                    r.get('發票號') or ''
                ))

                row_count = write_driver_sheet(ws, all_rows, data_col_names, fixed_cols, misc_cols, has_total)
                filename = f'{prefix}组司机月结单{yyyymm}.xlsx'
                filepath = os.path.join(export_dir, filename)
                wb.save(filepath)
                success_count += 1
                driver_info = ', '.join([f"{d['司機編號']}{d['司機姓名']}" for d in group])
                self.append_driver_export_result(f"  ✓ {filename} ({row_count} 行, {driver_count} 位司机: {driver_info})")

            # ── 2) 再导出数字开头的独立Excel ──
            for drv in digit_drivers:
                driver_code = drv['司機編號']
                driver_name = drv['司機姓名']
                self.cursor.execute(
                    "SELECT * FROM `司机月结单PIVOT` WHERE `司機編號` = %s AND `司機姓名` = %s ORDER BY `日期`, `發票號`",
                    (driver_code, driver_name)
                )
                rows = self.cursor.fetchall()
                if not rows:
                    continue

                data_col_names = [d[0] for d in self.cursor.description]
                wb = Workbook()
                ws = wb.active
                ws.title = 'Sheet1'
                data_row_count = write_driver_sheet(ws, rows, data_col_names, fixed_cols, misc_cols, has_total)

                # 同名司机有多个司機編號时，文件名加 A/B/C 后缀区分
                name_clean = re.sub(r'[\\/:*?"<>|]', '_', driver_name)
                codes_for_name = name_code_map.get(driver_name, [])
                if len(codes_for_name) > 1:
                    suffix = chr(ord('A') + codes_for_name.index(driver_code))
                    filename = name_clean + yyyymm + suffix + '.xlsx'
                else:
                    filename = name_clean + yyyymm + '.xlsx'
                filepath = os.path.join(export_dir, filename)
                wb.save(filepath)
                success_count += 1
                self.append_driver_export_result(f"  ✓ {filename} ({data_row_count} 行)")

            self.append_driver_export_result(f"\n{'='*60}")
            self.append_driver_export_result(f"导出完成: 共 {success_count} 个文件")
            self.append_driver_export_result(f"  其中: {len(digit_drivers)} 位数字开头司机(独立Excel), {len(sorted_prefixes)} 组英文前缀(合并Excel,同前缀司机同sheet)")
            self.append_driver_export_result(f"目录: {export_dir}")
            self.append_driver_export_result(f"{'='*60}")
            self.driver_export_status_var.set(f"导出完成: {success_count} 个文件")

        except Exception as e:
            self.append_driver_export_result(f"\n❌ 导出出错: {str(e)}")
            self.driver_export_status_var.set("导出失败")
            import traceback
            traceback.print_exc()

    def export_driver_summary_excel(self):
        """导出司机运杂费汇总表Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
            from openpyxl.utils import get_column_letter
            from datetime import datetime, timedelta
            import os, re

            year = int(self.summary_year_var.get())
            month = int(self.summary_month_combo.get())
            if not (1 <= month <= 12):
                messagebox.showerror("错误", "月份必须在 1-12 之间")
                return

            yyyymm = f"{year}{month:02d}"
            year_cn = f"{year}年{month}月"

            # 清空日志
            self.driver_summary_log_text.configure(state='normal')
            self.driver_summary_log_text.delete(1.0, tk.END)
            self.driver_summary_log_text.configure(state='disabled')

            self.driver_summary_status_var.set(f"正在导出 {year_cn}...")
            self.root.update()

            date_start = f'{year}-{month:02d}-01'
            if month == 12:
                next_month_dt = datetime(year + 1, 1, 1)
            else:
                next_month_dt = datetime(year, month + 1, 1)
            date_end = next_month_dt.strftime('%Y-%m-%d')
            date_end_display = (next_month_dt - timedelta(days=1)).strftime('%Y-%m-%d')

            # ── 读取分成比例 ──
            def parse_pct(var, label):
                """解析%比输入，返回0.0-1.0的浮点数"""
                raw = var.get().strip().replace('%', '')
                if raw == '':
                    return None
                try:
                    val = float(raw)
                    return val / 100.0
                except ValueError:
                    return None

            pct_tuo_c = parse_pct(self.pct_tuo_company_var, '营运拖头车公司%')
            pct_tuo_d = parse_pct(self.pct_tuo_driver_var, '营运拖头车司机%')
            pct_dun_c = parse_pct(self.pct_dun_company_var, '营运吨车公司%')
            pct_dun_d = parse_pct(self.pct_dun_driver_var, '营运吨车司机%')
            pct_flat_c = parse_pct(self.pct_flat_company_var, '平板车公司%')
            pct_flat_d = parse_pct(self.pct_flat_driver_var, '平板车司机%')

            def get_pct_for_code(code):
                """根据司机编号返回 (公司%比, 司机%比)，不需分成的返回 (None, None)"""
                code_s = code.strip()
                if code_s == '021':
                    return (None, None)
                if code_s.isdigit():
                    num = int(code_s)
                    if num < 800:
                        return (pct_tuo_c, pct_tuo_d)
                    elif num < 900:
                        return (pct_dun_c, pct_dun_d)
                    else:
                        return (pct_flat_c, pct_flat_d)
                return (None, None)  # 字母开头不分成
            self.append_driver_summary_log(f"{'='*60}\n")
            self.append_driver_summary_log(f"日期范围: {date_start} 至 {date_end_display}")

            self._check_connection()

            # ── 查询1：運費 + 公司運費（按司机汇总，来源和TEMP表一样） ──
            self.append_driver_summary_log("查询運费数据...")
            sql_freight = """
            SELECT
                dc.CUSTCODE AS driver_code,
                dc.NAME AS driver_name,
                dc.HKCP AS hkcp,
                COALESCE(SUM(cd.DRIVERCOMM), 0) AS total_freight,
                COALESCE(SUM(im.DRIVER_COMM), 0) AS total_company_freight
            FROM CON_DETAIL cd
            JOIN INVOICE_MASTER im ON im.INVOICECODE = cd.INVOICECODE
            JOIN DRIVER_CP dc ON dc.CUSTCODE = cd.DRIVER
            WHERE im.INVDATE >= %s AND im.INVDATE < %s
              AND cd.DRIVERCOMM IS NOT NULL AND cd.DRIVERCOMM != 0
            GROUP BY dc.CUSTCODE, dc.NAME, dc.HKCP
            """
            self.cursor.execute(sql_freight, (date_start, date_end))
            freight_rows = self.cursor.fetchall()
            freight_map = {}
            for r in freight_rows:
                freight_map[r['driver_code']] = {
                    'name': r['driver_name'] or '',
                    'hkcp': r['hkcp'] or '',
                    'freight': float(r['total_freight']),
                    'company_freight': float(r['total_company_freight']),
                }
            self.append_driver_summary_log(f"  運费数据: {len(freight_map)} 位司机")

            # ── 查询2：雜費和紮車加貨費（从INVOICE_DETAIL.DESCR后缀判断） ──
            self.append_driver_summary_log("查询雜費/紮車加貨费数据...")
            sql_misc = """
            SELECT
                id.DRIVERCODE AS driver_code,
                dc.NAME AS driver_name,
                dc.HKCP AS hkcp,
                id.DESCR AS descr,
                CAST(id.PRICE AS DECIMAL(12,2)) AS price,
                id.WHOPAY AS whopay
            FROM INVOICE_DETAIL id
            JOIN INVOICE_MASTER im ON im.INVOICECODE = id.INVOICECODE
            LEFT JOIN DRIVER_CP dc ON dc.CUSTCODE = id.DRIVERCODE
            WHERE im.INVDATE >= %s AND im.INVDATE < %s
              AND id.DRIVERCODE IS NOT NULL AND id.DRIVERCODE != ''
              AND id.PRICE IS NOT NULL AND id.PRICE != 0
            ORDER BY id.DRIVERCODE, im.INVDATE, id.NN
            """
            self.cursor.execute(sql_misc, (date_start, date_end))
            misc_rows = self.cursor.fetchall()

            # 按司机汇总
            driver_misc = {}
            for r in misc_rows:
                dcode = r['driver_code']
                descr = (r['descr'] or '').strip()
                price = float(r['price'])
                whopay = r['whopay']
                name = r['driver_name'] or ''
                hkcp = r['hkcp'] or ''

                if dcode not in driver_misc:
                    driver_misc[dcode] = {
                        'name': name, 'hkcp': hkcp,
                        'comma_total': 0.0,    # 逗号后缀汇总
                        'dot_total': 0.0,      # 句号后缀汇总
                        'noclass_w1': 0.0,     # 无后缀WHOPAY=1
                        'has_comma': False,
                    }

                dm = driver_misc[dcode]
                # 判断后缀
                if descr.endswith(','):
                    dm['comma_total'] += price
                    dm['has_comma'] = True
                elif descr.endswith('.') or descr.endswith('\u3002'):
                    dm['dot_total'] += price
                else:
                    if whopay == 1:
                        dm['noclass_w1'] += price

            self.append_driver_summary_log(f"  雜费数据: {len(driver_misc)} 位司机")

            # ── 合并数据 ──
            all_codes = sorted(set(list(freight_map.keys()) + list(driver_misc.keys())))
            driver_data = {}
            for code in all_codes:
                fi = freight_map.get(code, {'name': '', 'hkcp': '', 'freight': 0, 'company_freight': 0})
                mi = driver_misc.get(code, {'name': '', 'hkcp': '', 'comma_total': 0, 'dot_total': 0, 'noclass_w1': 0, 'has_comma': False})
                name = mi['name'] or fi['name'] or ''
                hkcp = mi['hkcp'] or fi['hkcp'] or ''

                # 雜费规则：有逗号项 = 逗号汇总；没有逗号项 = 无后缀WHOPAY=1
                if mi['has_comma']:
                    misc_fee = mi['comma_total']
                else:
                    misc_fee = mi['noclass_w1']

                freight = fi['freight']
                company_freight = fi['company_freight']
                binding = mi['dot_total']
                freight_total = freight + binding
                misc_freight_total = misc_fee + freight + binding

                driver_data[code] = {
                    'name': name,
                    'hkcp': hkcp,
                    'misc': misc_fee,
                    'freight': freight,
                    'company_freight': company_freight,
                    'binding': binding,
                    'freight_total': freight_total,
                    'misc_freight_total': misc_freight_total,
                }

            # 去除全为0的司机（无運费无雜费无紮車）
            driver_data = {k: v for k, v in driver_data.items()
                           if v['freight'] != 0 or v['misc'] != 0 or v['binding'] != 0}

            self.append_driver_summary_log(f"  合并后有效司机: {len(driver_data)} 位")

            # ── 分组 ──
            def get_group(code):
                """返回 (组名, 排序键, 是否字母组)"""
                code_s = code.strip()
                # 021公司代墊单独处理，不归入任何组
                if code_s == '021':
                    return ('company_advance', 'zzz_021', False)
                if code_s.isdigit():
                    num = int(code_s)
                    if num < 900:
                        return ('self_owned', f'{num:06d}', False)
                    else:
                        return ('flatbed', f'{num:06d}', False)
                else:
                    # 字母开头
                    prefix = re.match(r'^([A-Za-z]+)', code_s)
                    if prefix:
                        p = prefix.group(1).upper()
                        # X开头 = 散车
                        if p == 'X':
                            return ('scatter', f'{p}_{code_s}', True)
                        else:
                            return (p, f'{p}_{code_s}', True)
                    return ('other', code_s, True)

            groups_order = ['self_owned', 'flatbed', 'scatter']
            groups = {}
            for code, data in driver_data.items():
                gname, gsort, is_letter = get_group(code)
                if gname not in groups:
                    groups[gname] = []
                groups[gname].append((code, gsort, data))

            # 排序每组内的司机
            for gname in groups:
                groups[gname].sort(key=lambda x: x[1])

            # ── 创建Excel（严格按样版格式：13列A-M） ──
            wb = Workbook()
            ws = wb.active
            ws.title = f'司机产值{year_cn}A'

            # 样式
            title_font = Font(name='Microsoft YaHei UI', size=14, bold=True)
            header_font = Font(name='Microsoft YaHei UI', size=10, bold=True)
            normal_font = Font(name='Microsoft YaHei UI', size=10)
            subtotal_font = Font(name='Microsoft YaHei UI', size=10, bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            num_fmt = '#,##0.00'

            # Row 1: 标题（合并A1:M1）
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
            ws.cell(row=1, column=1, value=f'{year}年{month}月司机运杂费汇总统计表').font = title_font
            ws.cell(row=1, column=1).alignment = center_align

            # Row 2: 表头（13列，L2:M2合并为備注）
            headers = ['司機編號', '司機姓名', '香港車牌', '雜費', '運費', '紮車加貨費',
                       '運費合計', '運雜費合計', '公司分成合計', '司機分成合計', '公司運費', '備注', None]
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align
                cell.fill = header_fill
            # L2:M2 合并为"備注"
            ws.merge_cells('L2:M2')

            # Row 3+: 数据
            row_idx = 3

            # 街車合计累加
            street_misc = 0.0
            street_freight = 0.0
            street_binding = 0.0
            street_comp_freight = 0.0

            def write_data_row(ws, row, code, name, hkcp, misc, freight, binding,
                               company_freight=0.0, company_share=0.0, driver_share=0.0,
                               is_subtotal=False):
                ft = freight + binding
                mft = misc + freight + binding
                font = subtotal_font if is_subtotal else normal_font
                # K列(第11列)公司運費：有值则显示，0则留空
                k_val = company_freight if company_freight != 0 else None
                # I列(9)公司分成合計、J列(10)司機分成合計：有值显示，0留空
                i_val = company_share if company_share != 0 else None
                j_val = driver_share if driver_share != 0 else None
                vals = [code, name, hkcp, misc, freight, binding, ft, mft, i_val, j_val, k_val, None, None]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.font = font
                    cell.border = thin_border
                    if c <= 3:
                        cell.alignment = left_align
                    elif 4 <= c <= 10:
                        cell.alignment = right_align
                        cell.number_format = num_fmt
                    elif c == 11 and v is not None:
                        cell.alignment = right_align
                        cell.number_format = num_fmt
                    else:
                        cell.alignment = left_align
                # 小计行合并A:C
                if is_subtotal and code:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                return mft

            # 自營車组（纯数字 < 900，按编号排序）- 拖头车(1-799)和吨车(800-899)使用各自%比
            so_misc = so_freight = so_binding = so_comp_freight = 0.0
            so_company_share = so_driver_share = 0.0
            if 'self_owned' in groups:
                for code, gsort, data in groups['self_owned']:
                    d = data
                    pc, pd = get_pct_for_code(code)
                    ft = d['freight'] + d['binding']
                    cs = round(ft * pc, 2) if pc else 0.0
                    ds = round(ft * pd, 2) if pd else 0.0
                    write_data_row(ws, row_idx, code, d['name'], d['hkcp'],
                                   d['misc'], d['freight'], d['binding'], d['company_freight'], cs, ds)
                    so_misc += d['misc']; so_freight += d['freight']; so_binding += d['binding']
                    so_comp_freight += d['company_freight']
                    so_company_share += cs; so_driver_share += ds
                    row_idx += 1
            write_data_row(ws, row_idx, '自營車合計', '', '', so_misc, so_freight, so_binding, so_comp_freight, so_company_share, so_driver_share, True)
            row_idx += 1

            # 平板車组（纯数字 900-999）- 使用平板车%比
            fb_misc = fb_freight = fb_binding = fb_comp_freight = 0.0
            fb_company_share = fb_driver_share = 0.0
            if 'flatbed' in groups:
                for code, gsort, data in groups['flatbed']:
                    d = data
                    pc, pd = get_pct_for_code(code)
                    ft = d['freight'] + d['binding']
                    cs = round(ft * pc, 2) if pc else 0.0
                    ds = round(ft * pd, 2) if pd else 0.0
                    write_data_row(ws, row_idx, code, d['name'], d['hkcp'],
                                   d['misc'], d['freight'], d['binding'], d['company_freight'], cs, ds)
                    fb_misc += d['misc']; fb_freight += d['freight']; fb_binding += d['binding']
                    fb_comp_freight += d['company_freight']
                    fb_company_share += cs; fb_driver_share += ds
                    row_idx += 1
            write_data_row(ws, row_idx, '平板車合計', '', '', fb_misc, fb_freight, fb_binding, fb_comp_freight, fb_company_share, fb_driver_share, True)
            row_idx += 1

            # 街车 - 标题行（合并A:K，L=入線費及抵數，M=实际应付）
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=11)
            street_header_cell = ws.cell(row=row_idx, column=1, value='街车')
            street_header_cell.font = header_font
            street_header_cell.border = thin_border
            street_header_cell.alignment = left_align
            ws.cell(row=row_idx, column=12, value='入線費及抵數').font = header_font
            ws.cell(row=row_idx, column=12).border = thin_border
            ws.cell(row=row_idx, column=12).alignment = left_align
            ws.cell(row=row_idx, column=13, value='实际应付').font = header_font
            ws.cell(row=row_idx, column=13).border = thin_border
            ws.cell(row=row_idx, column=13).alignment = left_align
            row_idx += 1

            # 字母分组（除X外，按首字母分组）
            letter_groups = sorted([g for g in groups if g not in ('self_owned', 'flatbed', 'scatter', 'company_advance') and g != 'other'])
            for lg in letter_groups:
                lg_misc = lg_freight = lg_binding = lg_comp_freight = 0.0
                for code, gsort, data in groups[lg]:
                    d = data
                    write_data_row(ws, row_idx, code, d['name'], d['hkcp'],
                                   d['misc'], d['freight'], d['binding'], d['company_freight'])
                    lg_misc += d['misc']; lg_freight += d['freight']; lg_binding += d['binding']
                    lg_comp_freight += d['company_freight']
                    street_misc += d['misc']; street_freight += d['freight']; street_binding += d['binding']
                    street_comp_freight += d['company_freight']
                    row_idx += 1
                write_data_row(ws, row_idx, f'{lg}組合計', '', '', lg_misc, lg_freight, lg_binding, lg_comp_freight, 0.0, 0.0, True)
                row_idx += 1

            # 散车组（X开头）
            if 'scatter' in groups:
                sc_misc = sc_freight = sc_binding = sc_comp_freight = 0.0
                for code, gsort, data in groups['scatter']:
                    d = data
                    write_data_row(ws, row_idx, code, d['name'], d['hkcp'],
                                   d['misc'], d['freight'], d['binding'], d['company_freight'])
                    sc_misc += d['misc']; sc_freight += d['freight']; sc_binding += d['binding']
                    sc_comp_freight += d['company_freight']
                    street_misc += d['misc']; street_freight += d['freight']; street_binding += d['binding']
                    street_comp_freight += d['company_freight']
                    row_idx += 1
                write_data_row(ws, row_idx, '散車合計', '', '', sc_misc, sc_freight, sc_binding, sc_comp_freight, 0.0, 0.0, True)
                row_idx += 1

            # 街車合計
            street_ft = street_freight + street_binding
            street_mft = street_misc + street_freight + street_binding
            write_data_row(ws, row_idx, '街車合計', '', '', street_misc, street_freight, street_binding, street_comp_freight, 0.0, 0.0, True)
            row_idx += 1

            # 021公司代墊（如果存在）— 不合并A:C，A=021 B=公司代墊
            d021 = driver_data.get('021')
            if d021:
                write_data_row(ws, row_idx, '021', '公司代墊', '',
                               d021['misc'], d021['freight'], d021['binding'], d021['company_freight'])
                row_idx += 1

            # 总计 = 自營車 + 平板車 + 街車 + 021
            grand_misc = so_misc + fb_misc + street_misc
            grand_freight = so_freight + fb_freight + street_freight
            grand_binding = so_binding + fb_binding + street_binding
            grand_comp_freight = so_comp_freight + fb_comp_freight + street_comp_freight
            grand_company_share = so_company_share + fb_company_share
            grand_driver_share = so_driver_share + fb_driver_share
            if d021:
                grand_misc += d021['misc']
                grand_freight += d021['freight']
                grand_binding += d021['binding']
                grand_comp_freight += d021['company_freight']

            write_data_row(ws, row_idx, '合計', '', '', grand_misc, grand_freight, grand_binding, grand_comp_freight, grand_company_share, grand_driver_share, True)
            row_idx += 1

            # 自适应列宽（13列），第1列司機編號固定10
            ws.column_dimensions['A'].width = 10
            for col_idx in range(2, 14):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for row_num in range(1, row_idx):
                    cell = ws.cell(row=row_num, column=col_idx)
                    val = cell.value
                    if val is not None:
                        # 数字格式化为带千分位的字符串来估算显示宽度
                        if isinstance(val, (int, float)):
                            display = f'{val:,.2f}'
                        else:
                            display = str(val)
                        # 中文字符占2个字符宽度
                        cell_len = 0
                        for ch in display:
                            if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                                cell_len += 2
                            else:
                                cell_len += 1
                        if cell_len > max_len:
                            max_len = cell_len
                # 设置列宽 = 内容宽度 + padding，最小8，最大40
                final_width = max(8, min(40, max_len + 2))
                ws.column_dimensions[col_letter].width = final_width

            # 保存%比到配置
            self.config['pct_tuo_company'] = self.pct_tuo_company_var.get().strip().replace('%', '')
            self.config['pct_tuo_driver'] = self.pct_tuo_driver_var.get().strip().replace('%', '')
            self.config['pct_dun_company'] = self.pct_dun_company_var.get().strip().replace('%', '')
            self.config['pct_dun_driver'] = self.pct_dun_driver_var.get().strip().replace('%', '')
            self.config['pct_flat_company'] = self.pct_flat_company_var.get().strip().replace('%', '')
            self.config['pct_flat_driver'] = self.pct_flat_driver_var.get().strip().replace('%', '')
            self.save_config()

            # 保存到设定的导出目录（重复文件名加后缀 A/B/C...）
            base_dir = self.driver_summary_dir_var.get().strip()
            if not base_dir:
                messagebox.showerror("错误", "请先设置导出目录")
                return
            if not os.path.exists(base_dir):
                os.makedirs(base_dir)

            base_name = f'司机汇总表{yyyymm}'
            ext = '.xlsx'
            save_path = os.path.join(base_dir, base_name + ext)
            suffix_idx = 0
            while os.path.exists(save_path):
                suffix_idx += 1
                save_path = os.path.join(base_dir, base_name + chr(ord('A') + suffix_idx - 1) + ext)
            wb.save(save_path)

            self.append_driver_summary_log(f"\n{'='*60}")
            self.append_driver_summary_log(f"导出完成!")
            self.append_driver_summary_log(f"  文件: {save_path}")
            self.append_driver_summary_log(f"  司机数: {len(driver_data)}")
            self.append_driver_summary_log(f"  自營車: {len(groups.get('self_owned', []))} 位, 雜費={so_misc:.2f}, 運費={so_freight:.0f}, 公司運費={so_comp_freight:.0f}, 公司分成={so_company_share:.2f}, 司機分成={so_driver_share:.2f}")
            self.append_driver_summary_log(f"  平板車: {len(groups.get('flatbed', []))} 位, 雜費={fb_misc:.2f}, 運費={fb_freight:.0f}, 公司運費={fb_comp_freight:.0f}, 公司分成={fb_company_share:.2f}, 司機分成={fb_driver_share:.2f}")
            self.append_driver_summary_log(f"  街車: 雜費={street_misc:.2f}, 運費={street_freight:.0f}, 公司運費={street_comp_freight:.0f}")
            self.append_driver_summary_log(f"  总计: 雜費={grand_misc:.2f}, 運費={grand_freight:.0f}, 公司運費={grand_comp_freight:.0f}, 公司分成={grand_company_share:.2f}, 司機分成={grand_driver_share:.2f}")
            self.append_driver_summary_log(f"{'='*60}")

            self.driver_summary_status_var.set("导出完成")
            os.startfile(save_path)

        except ImportError as e:
            messagebox.showerror("错误", f"缺少必要的库: {e}\n\n请安装: pip install openpyxl")
            self.driver_summary_status_var.set("导出失败")
        except Exception as e:
            self.append_driver_summary_log(f"\n导出出错: {str(e)}")
            self.driver_summary_status_var.set("导出失败")
            import traceback
            traceback.print_exc()

    def show_custom_table(self):
        """显示自定义表页面 — 支持多客户专属月结单导出
        每个专属客户可使用 misc_name_rules 中与该客户绑定的特殊规则
        """
        canvas = tk.Canvas(self.main_content, bg='#f0f2f5', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f2f5')
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=1180)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 页面标题
        title_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))
        tk.Label(
            title_frame,
            text="自定义表",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg='#f0f2f5',
            fg='#262626'
        ).pack(side=tk.LEFT)
        tk.Label(
            title_frame,
            text="  专属客户月结单 — 支持客户专属规则",
            font=('Microsoft YaHei UI', 12),
            bg='#f0f2f5',
            fg='#8c8c8c'
        ).pack(side=tk.LEFT, padx=(15, 0))

        # ========== 主内容区：左右布局 ==========
        main_row = tk.Frame(scrollable_frame, bg='#f0f2f5')
        main_row.pack(fill=tk.BOTH, expand=True)

        # -------- 左侧：专属客户列表 (280px) --------
        left_panel = tk.Frame(main_row, bg='#ffffff', bd=1, relief='solid', width=280)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 15))
        left_panel.pack_propagate(False)

        left_header = tk.Frame(left_panel, bg='#ffffff', padx=12, pady=10)
        left_header.pack(fill=tk.X)
        tk.Label(left_header, text="专属客户列表", font=('Microsoft YaHei UI', 12, 'bold'),
                 bg='#ffffff', fg='#262626').pack(side=tk.LEFT)

        # 添加客户区域
        add_frame = tk.Frame(left_panel, bg='#fafafa', padx=10, pady=8)
        add_frame.pack(fill=tk.X, padx=10, pady=(0, 8))
        tk.Label(add_frame, text="客户编号:", font=('Microsoft YaHei UI', 9),
                 bg='#fafafa', fg='#595959').pack(anchor='w')
        add_row = tk.Frame(add_frame, bg='#fafafa')
        add_row.pack(fill=tk.X, pady=(4, 0))
        self.custom_add_code_var = tk.StringVar()
        tk.Entry(add_row, textvariable=self.custom_add_code_var, width=10,
                 font=('Microsoft YaHei UI', 10), relief='solid', bd=1).pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(add_row, text="+", font=('Microsoft YaHei UI', 10, 'bold'),
                  bg='#52c41a', fg='white', relief='flat', padx=8, pady=2,
                  cursor='hand2', command=self.add_custom_customer).pack(side=tk.LEFT, padx=(5, 0))

        # 客户列表（Treeview）
        cust_list_frame = tk.Frame(left_panel, bg='#ffffff')
        cust_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 5))

        cust_cols = ('cust_code', 'cust_name')
        self.cust_tree = ttk.Treeview(cust_list_frame, columns=cust_cols, show='headings',
                                       selectmode='browse', height=10)
        self.cust_tree.heading('cust_code', text='编号')
        self.cust_tree.heading('cust_name', text='客户名称')
        self.cust_tree.column('cust_code', width=60, anchor='center')
        self.cust_tree.column('cust_name', width=180)
        cust_vsb = ttk.Scrollbar(cust_list_frame, orient='vertical', command=self.cust_tree.yview)
        self.cust_tree.configure(yscrollcommand=cust_vsb.set)
        self.cust_tree.grid(row=0, column=0, sticky='nsew')
        cust_vsb.grid(row=0, column=1, sticky='ns')
        cust_list_frame.columnconfigure(0, weight=1)
        cust_list_frame.rowconfigure(0, weight=1)
        self.cust_tree.bind('<<TreeviewSelect>>', self.on_cust_select)
        self.cust_tree.bind('<ButtonRelease-1>', self._on_cust_click)

        # 删除按钮
        cust_btn_frame = tk.Frame(left_panel, bg='#ffffff', padx=10, pady=5)
        cust_btn_frame.pack(fill=tk.X)
        tk.Button(cust_btn_frame, text="移除选中客户", font=('Microsoft YaHei UI', 9),
                  bg='#ff4d4f', fg='white', relief='flat', padx=10, pady=4,
                  cursor='hand2', command=self.remove_custom_customer).pack(fill=tk.X)

        # -------- 右侧：导出设置 + 日志 --------
        right_panel = tk.Frame(main_row, bg='#f0f2f5')
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 当前选中客户信息
        cust_info_frame = tk.Frame(right_panel, bg='#e6f7ff', padx=15, pady=10)
        cust_info_frame.pack(fill=tk.X, pady=(0, 10))
        self.custom_selected_label = tk.Label(cust_info_frame, text="当前客户: 未选择",
                                               font=('Microsoft YaHei UI', 11, 'bold'),
                                               bg='#e6f7ff', fg='#0050b3')
        self.custom_selected_label.pack(side=tk.LEFT)

        # 导出设置卡片
        config_card = tk.Frame(right_panel, bg='#ffffff', bd=1, relief='solid')
        config_card.pack(fill=tk.X, pady=(0, 15), ipady=10)

        config_header = tk.Frame(config_card, bg='#ffffff', padx=20, pady=15)
        config_header.pack(fill=tk.X)
        tk.Label(
            config_header,
            text="📅 导出设置",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff',
            fg='#262626'
        ).pack(side=tk.LEFT)

        config_control = tk.Frame(config_card, bg='#f5f5f5', padx=20, pady=15)
        config_control.pack(fill=tk.X)

        # 年份选择
        current_date = datetime.now()
        row1 = tk.Frame(config_control, bg='#f5f5f5')
        row1.pack(fill=tk.X, pady=(0, 10))
        tk.Label(row1, text="年份：", font=('Microsoft YaHei UI', 11),
                 bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT)
        self.custom_year_var = tk.StringVar(value=str(current_date.year))
        year_combo = ttk.Combobox(row1, textvariable=self.custom_year_var,
                                  values=[str(y) for y in range(2003, 2051)],
                                  width=8, state='readonly', font=('Microsoft YaHei UI', 10))
        year_combo.pack(side=tk.LEFT, padx=(0, 20))

        # 月份选择
        tk.Label(row1, text="月份：", font=('Microsoft YaHei UI', 11),
                 bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT)
        self.custom_month_var = tk.StringVar(value=str(current_date.month))
        month_combo = ttk.Combobox(row1, textvariable=self.custom_month_var,
                                   values=[str(m) for m in range(1, 13)],
                                   width=5, state='readonly', font=('Microsoft YaHei UI', 10))
        month_combo.pack(side=tk.LEFT, padx=(0, 0))

        # 输出目录
        row2 = tk.Frame(config_control, bg='#f5f5f5')
        row2.pack(fill=tk.X, pady=(0, 10))
        tk.Label(row2, text="输出目录：", font=('Microsoft YaHei UI', 11),
                 bg='#f5f5f5', fg='#595959').pack(side=tk.LEFT)
        self.custom_output_dir_var = tk.StringVar(value=self.config.get('custom_table_dir', ''))
        self.custom_dir_entry = tk.Entry(row2, textvariable=self.custom_output_dir_var,
                                         font=('Microsoft YaHei UI', 10), width=50, state='readonly')
        self.custom_dir_entry.pack(side=tk.LEFT, padx=(5, 5))

        def browse_output_dir():
            d = filedialog.askdirectory(title="选择输出目录")
            if d:
                self.custom_output_dir_var.set(d)
                self.config['custom_table_dir'] = d
                self.save_config()

        tk.Button(row2, text="浏览...", font=('Microsoft YaHei UI', 10),
                  bg='#f0f0f0', relief='flat', padx=10, pady=3,
                  cursor='hand2', command=browse_output_dir).pack(side=tk.LEFT)

        # 导出按钮
        row3 = tk.Frame(config_control, bg='#f5f5f5')
        row3.pack(fill=tk.X, pady=(5, 0))
        self.custom_status_var = tk.StringVar(value="")
        tk.Label(row3, textvariable=self.custom_status_var,
                 font=('Microsoft YaHei UI', 10), bg='#f5f5f5', fg='#1890ff').pack(side=tk.LEFT)
        tk.Button(row3, text="开始导出", font=('Microsoft YaHei UI', 11, 'bold'),
                  bg='#1890ff', fg='white', relief='flat', padx=20, pady=6,
                  cursor='hand2', command=self.export_custom_table).pack(side=tk.RIGHT)

        # ========== 导出日志 ==========
        result_card = tk.Frame(right_panel, bg='#e6f7ff', padx=20, pady=15)
        result_card.pack(fill=tk.BOTH, expand=True)
        tk.Label(result_card, text="📋 导出日志",
                 font=('Microsoft YaHei UI', 12, 'bold'), bg='#e6f7ff', fg='#262626').pack(anchor='w', pady=(0, 10))
        self.custom_result_text = tk.Text(result_card, font=('Microsoft YaHei UI', 10), width=70, height=10,
                                          bg='#ffffff', fg='#333333', relief='solid', bd=1)
        self.custom_result_text.pack(fill=tk.BOTH, expand=True)
        self.custom_result_text.insert(tk.END, "从左侧列表选择客户，设置年月和输出目录后，点击「开始导出」...")
        self.custom_result_text.configure(state='disabled')

        # 初始化客户列表
        self._refresh_cust_tree()

    # ========== 客户列表管理方法 ==========

    def _refresh_cust_tree(self, select_code=None):
        """刷新专属客户列表，可选自动选中指定客户"""
        for item in self.cust_tree.get_children():
            self.cust_tree.delete(item)
        customers = self.config.get('custom_table_customers', ['0332'])
        for cc in customers:
            try:
                self._check_connection()
                self.cursor.execute("SELECT NAME FROM CUST_MASTER WHERE CUSTCODE = %s LIMIT 1", (cc,))
                row = self.cursor.fetchone()
                name = (row['NAME'] if isinstance(row, dict) else row[0]) if row else '(未找到)'
            except Exception:
                name = '(查询失败)'
            item = self.cust_tree.insert('', tk.END, values=(cc, name))
            if select_code and cc == select_code:
                self.cust_tree.selection_set(item)
                self.cust_tree.focus(item)

    def on_cust_select(self, event):
        """客户列表选择事件"""
        sel = self.cust_tree.selection()
        if sel:
            item = self.cust_tree.item(sel[0])
            values = item['values']
            self.selected_cust_code = values[0]
            if hasattr(self, 'custom_selected_label') and self.custom_selected_label.winfo_exists():
                self.custom_selected_label.config(
                    text=f"当前客户: {values[0]} — {values[1]}",
                    fg='#0050b3'
                )
        else:
            self.selected_cust_code = None
            if hasattr(self, 'custom_selected_label') and self.custom_selected_label.winfo_exists():
                self.custom_selected_label.config(text="当前客户: 未选择", fg='#8c8c8c')

    def _on_cust_click(self, event):
        """鼠标点击客户列表时的备用处理"""
        # 延迟执行，确保 Treeview 选择已更新
        self.root.after(50, self._sync_cust_selection)

    def _sync_cust_selection(self):
        """同步 Treeview 选中状态到 selected_cust_code"""
        if not hasattr(self, 'cust_tree') or not self.cust_tree.winfo_exists():
            return
        sel = self.cust_tree.selection()
        if sel:
            item = self.cust_tree.item(sel[0])
            values = item['values']
            if values:
                self.selected_cust_code = values[0]
                if hasattr(self, 'custom_selected_label') and self.custom_selected_label.winfo_exists():
                    self.custom_selected_label.config(
                        text=f"当前客户: {values[0]} — {values[1]}",
                        fg='#0050b3'
                    )
        else:
            self.selected_cust_code = None
            if hasattr(self, 'custom_selected_label') and self.custom_selected_label.winfo_exists():
                self.custom_selected_label.config(text="当前客户: 未选择", fg='#8c8c8c')

    def add_custom_customer(self):
        """添加专属客户"""
        code = self.custom_add_code_var.get().strip()
        if not code:
            messagebox.showwarning("提示", "请输入客户编号")
            return
        # 验证客户存在
        try:
            self._check_connection()
            self.cursor.execute("SELECT NAME FROM CUST_MASTER WHERE CUSTCODE = %s LIMIT 1", (code,))
            row = self.cursor.fetchone()
            if not row:
                messagebox.showwarning("提示", f"客户编号 {code} 在 CUST_MASTER 中不存在")
                return
            name = (row['NAME'] if isinstance(row, dict) else row[0])
        except Exception as e:
            messagebox.showerror("错误", f"查询客户失败: {e}")
            return

        customers = self.config.get('custom_table_customers', [])
        if code in customers:
            messagebox.showinfo("提示", f"客户 {code} 已在列表中")
            return
        customers.append(code)
        self.config['custom_table_customers'] = customers
        if not self.save_config():
            customers.pop()  # 回滚最后添加的
            return
        self.custom_add_code_var.set('')
        self._refresh_cust_tree(select_code=code)
        self.append_custom_log(f"已添加专属客户: {code} — {name}")

    def remove_custom_customer(self):
        """移除专属客户 — 直接从 Treeview 读取选中项"""
        sel = self.cust_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先在左侧列表中点击选中要移除的客户")
            return
        item = self.cust_tree.item(sel[0])
        values = item['values']
        code = str(values[0])
        cust_name = str(values[1]) if len(values) > 1 else ''

        customers = self.config.get('custom_table_customers', [])

        idx = None
        for i, c in enumerate(customers):
            if str(c) == code:
                idx = i
                break

        if idx is None:
            all_items = self.cust_tree.get_children()
            for i, tree_item in enumerate(all_items):
                if tree_item == sel[0] and i < len(customers):
                    idx = i
                    break

        if idx is None:
            messagebox.showwarning("提示", f"客户 {code} 不在专属列表中")
            return

        result = messagebox.askyesno("确认删除", f"确认将客户 {code}（{cust_name}）从专属列表中移除？\n\n"
                                      f"该客户的专属规则不会被删除，仍可在规则管理中管理。")
        if not result:
            return
        customers.pop(idx)
        if not self.save_config():
            customers.insert(idx, code)
            return
        if hasattr(self, 'custom_selected_label') and self.custom_selected_label.winfo_exists():
            self.custom_selected_label.config(text="当前客户: 未选择", fg='#8c8c8c')
        self._refresh_cust_tree()
        self.append_custom_log(f"已移除专属客户: {code}（{cust_name}）")

    def append_custom_log(self, text):
        """向自定义表日志框追加文字"""
        self.custom_result_text.configure(state='normal')
        self.custom_result_text.insert(tk.END, text + '\n')
        self.custom_result_text.see(tk.END)
        self.custom_result_text.configure(state='disabled')
        self.root.update_idletasks()

    def export_custom_table(self):
        """导出当前选中客户的专属月结单"""
        if not hasattr(self, 'selected_cust_code') or not self.selected_cust_code:
            messagebox.showwarning("提示", "请先在左侧列表中选择要导出的客户")
            return

        CUSTCODE = self.selected_cust_code
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime, timedelta

            year = int(self.custom_year_var.get())
            month = int(self.custom_month_var.get())
            if not (1 <= month <= 12):
                messagebox.showerror("错误", "月份必须在 1-12 之间")
                return
            yyyymm = f"{year}{month:02d}"
            year_cn = f"{year}年{month}月"

            # 日期范围
            date_start = f'{year}-{month:02d}-01'
            if month == 12:
                next_month = datetime(year + 1, 1, 1)
            else:
                next_month = datetime(year, month + 1, 1)
            date_end = next_month.strftime('%Y-%m-%d')
            date_end_display = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')

            output_dir = self.custom_output_dir_var.get().strip()
            if not output_dir:
                messagebox.showerror("错误", "请先选择输出目录")
                return

            # 获取客户名称
            self._check_connection()
            self.cursor.execute("SELECT NAME FROM CUST_MASTER WHERE CUSTCODE = %s LIMIT 1", (CUSTCODE,))
            cust_row = self.cursor.fetchone()
            cust_name = (cust_row['NAME'] if isinstance(cust_row, dict) else cust_row[0]) if cust_row else CUSTCODE

            # 清空日志
            self.custom_result_text.configure(state='normal')
            self.custom_result_text.delete('1.0', tk.END)
            self.custom_result_text.configure(state='disabled')

            self.custom_status_var.set(f"正在导出 {cust_name}({CUSTCODE}) {year_cn}...")
            self.root.update()

            self.append_custom_log(f"{'='*60}")
            self.append_custom_log(f"自定义表导出 - {cust_name}（{CUSTCODE}）")
            self.append_custom_log(f"{'='*60}\n")
            self.append_custom_log(f"日期范围: {date_start} 至 {date_end_display}")

            # 步骤1：生成TEMP表（限定当前客户）
            self.append_custom_log(f"\n步骤1/4: 生成客户月结单TEMP...")
            self.cursor.execute("DROP TABLE IF EXISTS `客户月结单TEMP`")

            sql_temp = """
            CREATE TABLE `客户月结单TEMP` AS
            SELECT
                t.INVDATE AS `日期`, t.INVOICECODE AS `發票號`, t.CUSTCODE AS `客戶編號`,
                COALESCE(cm.NAME, '') AS `客戶名稱`,
                t.NAME AS `司機姓名`, t.HKCP AS `香港車牌`, t.SZCP AS `大陸車牌`,
                t.DEST AS `地區`, t.CONCODE AS `櫃號`, t.SIZE AS `櫃尺碼`,
                t.TAKENO AS `提單號`, t.SHIP AS `船名`, t.SHIPCODE AS `托運號`, t.FEE AS `運費`,
                IF(t.dp IS NOT NULL, SUBSTRING_INDEX(t.dp,'§',1), NULL) AS `雜費名稱1`,
                IF(t.pp IS NOT NULL, CAST(SUBSTRING_INDEX(t.pp,'§',1) AS DECIMAL(12,2)), NULL) AS `雜費金額1`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=1, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',2),'§',-1), NULL) AS `雜費名稱2`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=1, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',2),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額2`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=2, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',3),'§',-1), NULL) AS `雜費名稱3`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=2, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',3),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額3`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=3, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',4),'§',-1), NULL) AS `雜費名稱4`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=3, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',4),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額4`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=4, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',5),'§',-1), NULL) AS `雜費名稱5`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=4, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',5),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額5`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=5, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',6),'§',-1), NULL) AS `雜費名稱6`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=5, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',6),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額6`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=6, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',7),'§',-1), NULL) AS `雜費名稱7`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=6, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',7),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額7`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=7, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',8),'§',-1), NULL) AS `雜費名稱8`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=7, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',8),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額8`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=8, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',9),'§',-1), NULL) AS `雜費名稱9`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=8, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',9),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額9`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=9, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',10),'§',-1), NULL) AS `雜費名稱10`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=9, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',10),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額10`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=10, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',11),'§',-1), NULL) AS `雜費名稱11`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=10, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',11),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額11`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=11, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',12),'§',-1), NULL) AS `雜費名稱12`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=11, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',12),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額12`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=12, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',13),'§',-1), NULL) AS `雜費名稱13`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=12, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',13),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額13`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=13, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',14),'§',-1), NULL) AS `雜費名稱14`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=13, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',14),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額14`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=14, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',15),'§',-1), NULL) AS `雜費名稱15`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=14, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',15),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額15`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=15, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',16),'§',-1), NULL) AS `雜費名稱16`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=15, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',16),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額16`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=16, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',17),'§',-1), NULL) AS `雜費名稱17`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=16, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',17),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額17`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=17, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',18),'§',-1), NULL) AS `雜費名稱18`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=17, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',18),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額18`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=18, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',19),'§',-1), NULL) AS `雜費名稱19`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=18, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',19),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額19`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=19, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',20),'§',-1), NULL) AS `雜費名稱20`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=19, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',20),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額20`
            FROM (
                SELECT
                    im.INVDATE, im.INVOICECODE, im.CUSTCODE, cd.DRIVER,
                    dc.NAME, dc.HKCP, dc.SZCP, im.DEST, cd.CONCODE, cd.SIZE, cd.TAKENO, im.SHIP, im.SHIPCODE, im.FEE,
                    GROUP_CONCAT(CASE WHEN id.WHOPAY=1 THEN id.DESCR END ORDER BY id.NN ASC SEPARATOR '§') AS dp,
                    GROUP_CONCAT(CASE WHEN id.WHOPAY=1 THEN id.PRICE END ORDER BY id.NN ASC SEPARATOR '§') AS pp
                FROM INVOICE_MASTER im
                LEFT JOIN CON_DETAIL cd ON cd.INVOICECODE = im.INVOICECODE AND cd.NN = 1
                LEFT JOIN DRIVER_CP dc ON dc.CUSTCODE = cd.DRIVER
                LEFT JOIN INVOICE_DETAIL id ON id.INVOICECODE = im.INVOICECODE
                WHERE im.INVDATE >= %s AND im.INVDATE < %s
                  AND im.CUSTCODE = %s
                GROUP BY im.INVOICECODE, im.INVDATE, im.CUSTCODE, cd.DRIVER, dc.NAME, dc.HKCP, dc.SZCP,
                    im.DEST, cd.CONCODE, cd.SIZE, cd.TAKENO, im.SHIP, im.SHIPCODE, im.FEE
            ) t
            LEFT JOIN CUST_MASTER cm ON cm.CUSTCODE = t.CUSTCODE
            ORDER BY t.INVDATE ASC, t.INVOICECODE ASC
            """
            self.cursor.execute(sql_temp, (date_start, date_end, CUSTCODE))
            self.cursor.execute("ALTER TABLE `客户月结单TEMP` ADD COLUMN `id` INT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY FIRST")
            self.conn.commit()

            self.cursor.execute("SELECT COUNT(*) as cnt FROM `客户月结单TEMP`")
            temp_count = self.cursor.fetchone()['cnt']
            self.append_custom_log(f"[OK] TEMP 生成完成 ({temp_count} 条记录)")

            if temp_count == 0:
                self.append_custom_log(f"\n⚠️ {year}年{month}月 {CUSTCODE} 无数据，导出取消")
                self.custom_status_var.set("无数据")
                return

            # 步骤2：归一化雜費名稱（使用当前客户专属规则）
            self.append_custom_log(f"\n步骤2/4: 归一化雜費名稱（{CUSTCODE}专属规则）...")
            self._normalize_misc_names("客户月结单TEMP", cust_code=CUSTCODE)
            self.conn.commit()
            self.append_custom_log("[OK] 雜費名稱归一化 完成")

            # 步骤3：PIVOT杂费（使用当前客户专属规则）
            self.append_custom_log(f"\n步骤3/4: 执行 pivot_misc_names（{CUSTCODE}专属）...")
            self._pivot_misc_names(cust_code=CUSTCODE)
            self.conn.commit()
            self.cursor.execute("SELECT COUNT(*) as cnt FROM `客户月结单PIVOT`")
            pivot_count = self.cursor.fetchone()['cnt']
            self.append_custom_log(f"[OK] PIVOT 完成 ({pivot_count} 条记录)")

            # 步骤4：导出Excel
            self.append_custom_log(f"\n步骤4/4: 导出Excel...")
            self._export_custom_excel(CUSTCODE, cust_name, year, month, yyyymm, year_cn, date_start, date_end, output_dir)

        except ImportError as e:
            messagebox.showerror("错误", f"缺少必要的库: {e}\n\n请安装: pip install pandas openpyxl")
            self.custom_status_var.set("导出失败")
        except Exception as e:
            self.append_custom_log(f"\n[错误] 导出失败: {e}")
            self.custom_status_var.set("导出失败")
            import traceback
            self.append_custom_log(traceback.format_exc())
            messagebox.showerror("错误", str(e))

    def _export_custom_excel(self, CUSTCODE, cust_name, year, month, yyyymm, year_cn, date_start, date_end, output_dir):
        """将指定客户PIVOT数据导出为Excel（单文件）"""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        FIXED_COLS = ['日期', '發票號', '司機姓名', '香港車牌', '大陸車牌', '地區', '櫃號', '櫃尺碼', '提單號', '船名', '托運號', '運費']
        TOTAL_COL = '運雜費合計'

        def font_func(size=11, bold=False):
            return Font(name='宋体', size=size, bold=bold)

        def align_func(horizontal=None, vertical=None, wrap=False):
            return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

        def thin_border_func():
            s = Side(style='thin')
            return Border(top=s, bottom=s, left=s, right=s)

        # 获取客户名称
        self.cursor.execute("SELECT NAME FROM CUST_MASTER WHERE CUSTCODE = %s LIMIT 1", (CUSTCODE,))
        cust_row = self.cursor.fetchone()
        cust_name = cust_row['NAME'] if cust_row else '深圳中電前海倉儲運營有限公司'

        # 获取所有杂费列（排除固定列和合计列）
        self.cursor.execute("""
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = 'SWT' AND TABLE_NAME = '客户月结单PIVOT'
            AND COLUMN_NAME NOT IN ('id', '_id', '客戶編號', '客戶名稱',
                '日期', '發票號', '司機姓名', '香港車牌', '大陸車牌', '地區', '櫃號', '櫃尺碼',
                '提單號', '船名', '托運號', '運費', '運雜費合計')
            ORDER BY ORDINAL_POSITION
        """)
        all_misc_cols = []
        for row in self.cursor.fetchall():
            col = row['COLUMN_NAME'] if isinstance(row, dict) else row[0]
            all_misc_cols.append(col)

        # 确定实际有数据的杂费列（列名可能含 %，需转义为 %% 防 pymysql 格式化冲突）
        misc_cols_str = ', '.join([f'`{c.replace("%", "%%")}`' for c in all_misc_cols])
        select_cols = ', '.join(FIXED_COLS + [misc_cols_str, '`運雜費合計`'])
        self.cursor.execute(f"""
            SELECT {select_cols}
            FROM 客户月结单PIVOT
            WHERE 客戶編號 = %s AND 日期 >= %s AND 日期 < %s
            ORDER BY 日期, 發票號
        """, (CUSTCODE, date_start, date_end))
        rows = self.cursor.fetchall()
        column_names = [desc[0] for desc in self.cursor.description]

        if not rows:
            self.append_custom_log("⚠️ PIVOT中无数据")
            self.custom_status_var.set("无数据")
            return

        # 去重（按發票號）
        seen = set()
        deduped = []
        for r in rows:
            inv = r.get('發票號') if isinstance(r, dict) else None
            if inv not in seen:
                seen.add(inv)
                deduped.append(r)
        rows = deduped
        if isinstance(rows[0], dict):
            df = pd.DataFrame(rows)
        else:
            df = pd.DataFrame(rows, columns=column_names)
        cols = df.columns.tolist()

        # 筛选有数据的金额列
        amount_cols = []
        for col in cols:
            if col not in FIXED_COLS and col != TOTAL_COL:
                try:
                    s = pd.to_numeric(df[col], errors='coerce')
                    if s.fillna(0).sum() > 0:
                        amount_cols.append(col)
                except:
                    pass

        all_cols = FIXED_COLS + amount_cols + [TOTAL_COL]

        self.append_custom_log(f"数据: {len(rows)} 行, 杂费列: {amount_cols}")

        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '客户明细表2016A'
        last_col_letter = get_column_letter(len(all_cols))

        # 第1行：标题
        ws.merge_cells(f'A1:{last_col_letter}1')
        c = ws['A1']
        c.value = f'深華交通實業有限公司  {year_cn} 對帳單'
        c.font = font_func(18, bold=True)
        c.alignment = align_func('center')
        ws.row_dimensions[1].height = 22.5

        # 第2-5行：公司信息
        info_rows = [
            '香港地址:香港新界上水新運路188號劍橋廣場A619室  香港電話:23689666 香港傳真:26689135',
            '深圳地址:深圳市福田区皇岗口岸皇御苑一栋205室    深圳电话:83859938 深圳传真:83858928',
            '香港中銀戶名:SUM WAH TRANSPORT ENTERPRISES CO LTD 帳号:01960410087700',
            '深圳市工商银行皇岗支行人民币户名:深圳市深华运输有限公司 帳号:4000026319200117006'
        ]
        for i, text in enumerate(info_rows, start=2):
            ws.merge_cells(f'A{i}:{last_col_letter}{i}')
            c = ws[f'A{i}']
            c.value = text
            c.font = font_func(11)
            c.alignment = align_func('center')

        # 第6行：客户名称
        ws[f'A6'].value = f'客戶名稱: {cust_name}'
        ws[f'A6'].font = font_func(11)

        # 第7行：表头
        ws.row_dimensions[7].height = 30
        for col_idx, col_name in enumerate(all_cols, start=1):
            c = ws.cell(row=7, column=col_idx)
            c.value = col_name
            c.font = font_func(11, bold=(col_idx <= 12))
            c.alignment = align_func('center', 'center')
            c.border = thin_border_func()

        # 第8行起：数据行
        DATA_START = 8
        num_cols = amount_cols + ['運費']
        col_totals = {ac: 0.0 for ac in num_cols}
        grand_total = 0.0
        for row_offset, row_data in enumerate(rows):
            r = DATA_START + row_offset
            ws.row_dimensions[r].height = 30
            row_dict = row_data if isinstance(row_data, dict) else dict(zip(cols, row_data))

            row_total = 0.0
            for ac in num_cols:
                try:
                    v = float(row_dict.get(ac, 0) or 0)
                    row_total += v
                    col_totals[ac] += v
                except (ValueError, TypeError):
                    pass
            grand_total += row_total

            for col_idx, col_name in enumerate(all_cols, start=1):
                c = ws.cell(row=r, column=col_idx)
                c.font = font_func(11)
                c.border = thin_border_func()

                if col_name == TOTAL_COL:
                    c.value = row_total
                    c.alignment = align_func('right', 'center')
                    c.number_format = '#,##0.00'
                elif col_name == '運費':
                    val = row_dict.get(col_name, 0)
                    try:
                        fv = float(val) if val is not None else 0
                        c.value = '0.00' if fv == 0 else fv
                    except (TypeError, ValueError):
                        c.value = '0.00'
                    c.alignment = align_func('right', 'center', wrap=True)
                    c.number_format = '#,##0.00'
                elif col_name in amount_cols:
                    val = row_dict.get(col_name, 0)
                    try:
                        fv = float(val) if val is not None else 0
                        c.value = '' if fv == 0 else fv
                    except (TypeError, ValueError):
                        c.value = ''
                    c.alignment = align_func('right', 'center', wrap=True)
                    c.number_format = '#,##0.00'
                else:
                    val = row_dict.get(col_name, '')
                    if col_name == '日期' and val is not None and val != '':
                        c.value = str(val)[:10]
                    elif val is not None and val != '':
                        c.value = str(val)
                    else:
                        c.value = ''
                    c.alignment = align_func(None, 'center', wrap=True)

        # 合计行
        total_row = DATA_START + len(rows)
        ws.cell(row=total_row, column=1).value = '合計'
        ws.cell(row=total_row, column=1).font = font_func(11)
        ws.cell(row=total_row, column=1).border = thin_border_func()

        for col_idx in range(2, len(all_cols) + 1):
            c = ws.cell(row=total_row, column=col_idx)
            c.font = font_func(11)
            c.border = thin_border_func()
            c.alignment = align_func('right', 'center')
            col_name = all_cols[col_idx - 1]
            if col_name == TOTAL_COL:
                c.value = grand_total
                c.number_format = '#,##0.00'
            elif col_name in amount_cols or col_name == '運費':
                c.value = col_totals.get(col_name, 0.0)
                c.number_format = '#,##0.00'

        # 自适应列宽（中文宽度×2 + 英文宽度×1，含合计行）
        for col_idx, col_name in enumerate(all_cols, start=1):
            col_letter = get_column_letter(col_idx)
            max_display_len = 0   # 用于表头（中文×2）
            max_ascii_len = 0     # 用于数字列（纯字符数）
            # 表头宽度
            header_display_len = sum(2 if ord(c) > 127 else 1 for c in str(col_name))
            max_display_len = max(max_display_len, header_display_len)
            # 数据行宽度 + 累计合计值
            col_sum = 0
            for row_data in rows:
                rd = row_data if isinstance(row_data, dict) else dict(zip(cols, row_data))
                val = rd.get(col_name)
                if val is not None and val != '':
                    if col_name in amount_cols + [TOTAL_COL, '運費']:
                        try:
                            num_val = float(val)
                            col_sum += num_val
                            s = f'{num_val:,.2f}' if num_val != 0 else '0.00'
                            max_ascii_len = max(max_ascii_len, len(s))
                        except:
                            pass
                    else:
                        cell_str = str(val)[:10] if col_name == '日期' else str(val)
                        display_len = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                        max_display_len = max(max_display_len, display_len)
            # 合计行宽度（格式化后的总和）
            if col_name in amount_cols + [TOTAL_COL, '運費']:
                total_str = f'{col_sum:,.2f}'
                max_ascii_len = max(max_ascii_len, len(total_str))
            # 列宽计算：表头用 display_len * 0.75 + 2.5，数字列用 ascii_len + 2
            width_from_header = max_display_len * 0.75 + 2.5
            if max_ascii_len > 0:
                width_from_data = max_ascii_len + 2
                final_width = max(width_from_header, width_from_data)
            else:
                final_width = width_from_header
            final_width = min(final_width, 80)
            ws.column_dimensions[col_letter].width = final_width

        # 保存
        os.makedirs(output_dir, exist_ok=True)
        base_filename = f'{cust_name}{yyyymm}.xlsx'
        filepath = os.path.join(output_dir, base_filename)
        suffix_idx = ord('A')
        while os.path.exists(filepath):
            suffix = chr(suffix_idx)
            filepath = os.path.join(output_dir, f'{cust_name}{yyyymm}{suffix}.xlsx')
            suffix_idx += 1
            if suffix_idx > ord('Z'):
                suffix_idx = ord('A')

        wb.save(filepath)

        self.append_custom_log(f"\n{'='*60}")
        self.append_custom_log(f"✅ 导出完成!")
        self.append_custom_log(f"文件: {filepath}")
        self.append_custom_log(f"数据: {len(rows)} 行, 杂费列: {amount_cols}")
        self.append_custom_log(f"{'='*60}")
        self.custom_status_var.set(f"✓ 导出完成 - {len(rows)} 行")

    def show_company_stats(self):
        """显示公司统计页面 - 客户汇总表导出"""
        canvas = tk.Canvas(self.main_content, bg='#f0f2f5', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f2f5')
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=1180)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 页面标题
        title_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))
        tk.Label(title_frame, text="公司统计", font=('Microsoft YaHei UI', 20, 'bold'),
                 bg='#f0f2f5', fg='#262626').pack(side=tk.LEFT)

    # ========== 规则管理方法 ==========

    def refresh_rules(self):
        """刷新规则列表"""
        if self.conn is None or self.cursor is None:
            self.stats_label.config(text="数据库未连接，请检查设置")
            return
        try:
            self.cursor.execute("""
                SELECT id, sort_order, pattern_type, keyword, replacement, cust_code
                FROM misc_name_rules
                ORDER BY sort_order
            """)
            rules = self.cursor.fetchall()

            # 清空列表
            for item in self.tree.get_children():
                self.tree.delete(item)

            # 填充列表 - 处理 DictCursor 返回的字典
            for rule in rules:
                if isinstance(rule, dict):
                    # DictCursor 返回字典，提取各字段值
                    values = (
                        rule.get('id'),
                        rule.get('sort_order'),
                        rule.get('pattern_type'),
                        rule.get('keyword'),
                        rule.get('replacement'),
                        rule.get('cust_code') or ''
                    )
                else:
                    # 元组情况，直接使用
                    values = rule
                self.tree.insert('', tk.END, values=values)

            # 更新统计信息
            self.stats_label.config(text=f"共 {len(rules)} 条规则")

        except Exception as e:
            messagebox.showerror("错误", f"加载规则失败: {e}")

    def search_rules(self):
        """搜索规则"""
        keyword = self.search_var.get().strip()

        if not keyword:
            self.refresh_rules()
            return

        try:
            self.cursor.execute("""
                SELECT id, sort_order, pattern_type, keyword, replacement, cust_code
                FROM misc_name_rules
                WHERE keyword LIKE %s OR replacement LIKE %s
                ORDER BY sort_order
            """, (f'%{keyword}%', f'%{keyword}%'))
            rules = self.cursor.fetchall()

            # 清空列表
            for item in self.tree.get_children():
                self.tree.delete(item)

            # 填充列表
            for rule in rules:
                if isinstance(rule, dict):
                    values = (
                        rule.get('id'), rule.get('sort_order'), rule.get('pattern_type'),
                        rule.get('keyword'), rule.get('replacement'), rule.get('cust_code') or ''
                    )
                else:
                    values = rule
                self.tree.insert('', tk.END, values=values)

            # 更新统计信息
            self.stats_label.config(text=f"找到 {len(rules)} 条匹配规则")

        except Exception as e:
            messagebox.showerror("错误", f"搜索失败: {e}")

    def on_rule_select(self, event):
        """规则选择事件"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            values = item['values']

            self.selected_rule_id = values[0]
            self.id_var.set(values[0])
            self.sort_order_var.set(values[1])
            self.pattern_type_var.set(values[2])
            self.keyword_var.set(values[3])
            self.replacement_var.set(values[4])
            self.cust_code_var.set(values[5] if len(values) > 5 and values[5] else '')

    def clear_form(self):
        """清空表单"""
        self.selected_rule_id = None
        self.id_var.set('')
        self.sort_order_var.set('')
        self.pattern_type_var.set('prefix')
        self.keyword_var.set('')
        self.replacement_var.set('')
        self.cust_code_var.set('')

    def add_rule(self):
        """添加新规则"""
        try:
            # 自动生成排序顺序
            self.cursor.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 as cnt FROM misc_name_rules")
            sort_order = self.cursor.fetchone()['cnt']

            pattern_type = 'prefix'
            keyword = self.keyword_var.get().strip()
            replacement = self.replacement_var.get().strip()
            cust_code = self.cust_code_var.get().strip()
            if not cust_code:
                cust_code = None

            if not keyword:
                messagebox.showwarning("警告", "关键字不能为空")
                return

            if not replacement:
                messagebox.showwarning("警告", "替换值不能为空")
                return

            # 检查关键字是否重复（同客户范围内）
            if cust_code:
                self.cursor.execute(
                    "SELECT COUNT(*) as cnt FROM misc_name_rules WHERE keyword = %s AND cust_code = %s",
                    (keyword, cust_code))
            else:
                self.cursor.execute(
                    "SELECT COUNT(*) as cnt FROM misc_name_rules WHERE keyword = %s AND cust_code IS NULL",
                    (keyword,))
            if self.cursor.fetchone()['cnt'] > 0:
                messagebox.showwarning("警告", f"关键字「{keyword}」已存在，请勿重复添加")
                return

            # 插入规则
            self.cursor.execute("""
                INSERT INTO misc_name_rules (sort_order, pattern_type, keyword, replacement, memo, cust_code)
                VALUES (%s, %s, %s, %s, NULL, %s)
            """, (sort_order, pattern_type, keyword, replacement, cust_code))
            self.conn.commit()

            messagebox.showinfo("成功", "规则添加成功！")
            self.clear_form()
            self.refresh_rules()

        except Exception as e:
            self.conn.rollback()
            messagebox.showerror("错误", f"添加失败: {e}")

    def edit_rule(self):
        """修改规则"""
        if self.selected_rule_id is None:
            messagebox.showwarning("警告", "请先选择要修改的规则")
            return

        try:
            sort_order = int(self.sort_order_var.get().strip())
            pattern_type = self.pattern_type_var.get().strip()
            keyword = self.keyword_var.get().strip()
            replacement = self.replacement_var.get().strip()
            cust_code = self.cust_code_var.get().strip()
            if not cust_code:
                cust_code = None

            if not keyword:
                messagebox.showwarning("警告", "关键字不能为空")
                return

            if not replacement:
                messagebox.showwarning("警告", "替换值不能为空")
                return

            # 检查关键字是否重复（排除自身）
            if cust_code:
                self.cursor.execute(
                    "SELECT COUNT(*) as cnt FROM misc_name_rules WHERE keyword = %s AND cust_code = %s AND id != %s",
                    (keyword, cust_code, self.selected_rule_id))
            else:
                self.cursor.execute(
                    "SELECT COUNT(*) as cnt FROM misc_name_rules WHERE keyword = %s AND cust_code IS NULL AND id != %s",
                    (keyword, self.selected_rule_id))
            if self.cursor.fetchone()['cnt'] > 0:
                messagebox.showwarning("警告", f"关键字「{keyword}」已存在，请使用其他关键字")
                return

            # 更新规则
            self.cursor.execute("""
                UPDATE misc_name_rules
                SET sort_order = %s, pattern_type = %s, keyword = %s, replacement = %s, cust_code = %s
                WHERE id = %s
            """, (sort_order, pattern_type, keyword, replacement, cust_code, self.selected_rule_id))
            self.conn.commit()

            messagebox.showinfo("成功", "规则修改成功！")
            self.clear_form()
            self.refresh_rules()

        except Exception as e:
            self.conn.rollback()
            messagebox.showerror("错误", f"修改失败: {e}")

    def delete_rule(self):
        """删除规则"""
        if self.selected_rule_id is None:
            messagebox.showwarning("警告", "请先选择要删除的规则")
            return

        # 获取规则信息
        self.cursor.execute("""
            SELECT keyword, replacement
            FROM misc_name_rules
            WHERE id = %s
        """, (self.selected_rule_id,))
        rule = self.cursor.fetchone()

        if not rule:
            messagebox.showerror("错误", "规则不存在")
            return

        # 确认删除
        result = messagebox.askyesno(
            "确认删除",
            f"确认删除以下规则?\n\n关键字: {rule[0]}\n替换值: {rule[1]}"
        )

        if result:
            try:
                self.cursor.execute("DELETE FROM misc_name_rules WHERE id = %s", (self.selected_rule_id,))
                self.conn.commit()

                messagebox.showinfo("成功", "规则删除成功！")
                self.clear_form()
                self.refresh_rules()

            except Exception as e:
                self.conn.rollback()
                messagebox.showerror("错误", f"删除失败: {e}")

    def update_procedures(self):
        """已废弃：归一化逻辑已内嵌代码（_normalize_misc_names），不再需要存储过程"""
        messagebox.showinfo("提示", "归一化功能已改为代码内执行，不再需要更新存储过程。\n\n"
                            "修改 misc_name_rules 表后自动生效。")

    def export_rules(self):
        """导出规则到文件"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            initialfile=f"misc_name_rules_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )

        if not filename:
            return

        try:
            self.cursor.execute("""
                SELECT id, sort_order, pattern_type, keyword, replacement, cust_code
                FROM misc_name_rules
                ORDER BY sort_order
            """)
            rules = self.cursor.fetchall()

            with open(filename, 'w', encoding='utf-8') as f:
                f.write("# 杂费名称规则导出\n")
                f.write(f"# 导出时间: {datetime.now()}\n")
                f.write(f"# 共 {len(rules)} 条规则\n")
                f.write("# 格式: sort_order|pattern_type|keyword|replacement|cust_code\n\n")

                for rule in rules:
                    if isinstance(rule, dict):
                        cc = rule.get('cust_code') or ''
                        line = f"{rule['sort_order']}|{rule['pattern_type']}|{rule['keyword']}|{rule['replacement']}|{cc}"
                    else:
                        cc = rule[5] if len(rule) > 5 and rule[5] else ''
                        line = f"{rule[1]}|{rule[2]}|{rule[3]}|{rule[4]}|{cc}"
                    f.write(line + '\n')

            messagebox.showinfo("成功", f"规则导出成功！\n\n文件: {filename}\n共 {len(rules)} 条规则")

        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    def import_rules(self):
        """从文件导入规则"""
        filename = filedialog.askopenfilename(
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if not filename:
            return

        try:
            with open(filename, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            import_count = 0
            error_count = 0
            error_lines = []

            for line_num, line in enumerate(lines, 1):
                line = line.strip()

                if not line or line.startswith('#'):
                    continue

                parts = line.split('|')
                if len(parts) >= 4:
                    try:
                        sort_order = int(parts[0])
                        pattern_type = parts[1]
                        keyword = parts[2]
                        replacement = parts[3]
                        cust_code = parts[4].strip() if len(parts) >= 5 else None
                        if not cust_code:
                            cust_code = None

                        self.cursor.execute("""
                            INSERT INTO misc_name_rules (sort_order, pattern_type, keyword, replacement, memo, cust_code)
                            VALUES (%s, %s, %s, %s, NULL, %s)
                        """, (sort_order, pattern_type, keyword, replacement, cust_code))
                        import_count += 1
                    except Exception as e:
                        error_count += 1
                        error_lines.append(f"第 {line_num} 行: {line}")
                else:
                    error_count += 1
                    error_lines.append(f"第 {line_num} 行 (格式错误): {line}")

            self.conn.commit()

            message = f"规则导入完成！\n\n成功: {import_count} 条\n失败: {error_count} 条"

            if error_count > 0 and len(error_lines) <= 10:
                message += "\n\n失败详情:\n" + "\n".join(error_lines[:10])
                if len(error_lines) > 10:
                    message += f"\n... 还有 {len(error_lines) - 10} 条错误"

            if import_count > 0:
                self.refresh_rules()

            messagebox.showinfo("导入结果", message)

        except FileNotFoundError:
            messagebox.showerror("错误", f"文件不存在: {filename}")
        except Exception as e:
            self.conn.rollback()
            messagebox.showerror("错误", f"导入失败: {e}")

    # ========== 月度检查方法 ==========

    def check_monthly(self):
        """执行月度检查"""
        try:
            year = int(self.year_combo.get())
            month = int(self.month_combo.get())

            if not (1 <= month <= 12):
                messagebox.showerror("错误", "月份必须在 1-12 之间")
                return

            yyyymm = f"{year}{month:02d}"

            # 清空结果文本框
            self.check_result_text.configure(state='normal')
            self.check_result_text.delete(1.0, tk.END)
            self.check_result_text.configure(state='disabled')

            self.check_status_var.set(f"正在检查 {year}年{month}月...")
            self.root.update()

            date_start = f'{year}-{month:02d}-01'
            if month == 12:
                date_end = f'{year+1}-01-01'
            else:
                date_end = f'{year}-{month+1:02d}-01'

            # 步骤1
            self.append_check_result(f"{'='*60}")
            self.append_check_result(f"客户月结单检查 - {year}年{month}月")
            self.append_check_result(f"{'='*60}\n")
            self.append_check_result("步骤1: 删除旧的客户月结单TEMP表...")
            self.cursor.execute("DROP TABLE IF EXISTS `客户月结单TEMP`")
            self.append_check_result("[OK] 客户月结单TEMP表已删除")

            # 步骤2
            self.append_check_result(f"\n步骤2: 执行2客户月结单查询，生成客户月结单TEMP ({year}-{month:02d})...")

            self.cursor.execute(f"""
                CREATE TEMPORARY TABLE temp_raw AS
                SELECT
                    t.INVDATE AS `日期`, t.INVOICECODE AS `發票號`, t.CUSTCODE AS `客戶編號`,
                    COALESCE(cm.NAME, '') AS `客戶名稱`,
                    t.NAME AS `司機姓名`, t.HKCP AS `香港車牌`, t.SZCP AS `大陸車牌`,
                    t.DEST AS `地區`, t.CONCODE AS `櫃號`, t.SIZE AS `櫃尺碼`,
                    t.TAKENO AS `提單號`, t.SHIP AS `船名`, t.SHIPCODE AS `托運號`, t.FEE AS `運費`,
                    t.dp, t.pp, t.`運雜費合計`
                FROM (
                    SELECT
                        im.INVDATE, im.INVOICECODE, im.CUSTCODE, cd.DRIVER,
                        dc.NAME, dc.HKCP, dc.SZCP, im.DEST, cd.CONCODE, cd.SIZE, cd.TAKENO, im.SHIP, im.SHIPCODE, im.FEE,
                        GROUP_CONCAT(CASE WHEN id.WHOPAY=1 THEN id.DESCR END ORDER BY id.NN ASC SEPARATOR '§') AS dp,
                        GROUP_CONCAT(CASE WHEN id.WHOPAY=1 THEN id.PRICE END ORDER BY id.NN ASC SEPARATOR '§') AS pp,
                        COALESCE(im.FEE,0) + COALESCE(SUM(CASE WHEN id.WHOPAY=1 THEN id.PRICE ELSE 0 END),0) AS `運雜費合計`
                    FROM INVOICE_MASTER im
                    LEFT JOIN CON_DETAIL cd ON cd.INVOICECODE = im.INVOICECODE AND cd.NN = 1
                    LEFT JOIN DRIVER_CP dc ON dc.CUSTCODE = cd.DRIVER
                    LEFT JOIN INVOICE_DETAIL id ON id.INVOICECODE = im.INVOICECODE
                    WHERE im.INVDATE >= %s AND im.INVDATE < %s
                    GROUP BY im.INVOICECODE, im.INVDATE, im.CUSTCODE, cd.DRIVER, dc.NAME, dc.HKCP, dc.SZCP,
                        im.DEST, cd.CONCODE, cd.SIZE, cd.TAKENO, im.SHIP, im.SHIPCODE, im.FEE
                ) t
                LEFT JOIN CUST_MASTER cm ON cm.CUSTCODE = t.CUSTCODE
            """, (date_start, date_end))

            self.cursor.execute(f"""
                CREATE TABLE `客户月结单TEMP` AS
                SELECT `日期`, `發票號`, `客戶編號`, `客戶名稱`, `司機姓名`, `香港車牌`, `大陸車牌`,
                    `地區`, `櫃號`, `櫃尺碼`, `提單號`, `船名`, `托運號`, `運費`,
                    IF(dp IS NOT NULL, SUBSTRING_INDEX(dp,'§',1), NULL) AS `雜費名稱1`,
                    IF(pp IS NOT NULL, CAST(SUBSTRING_INDEX(pp,'§',1) AS DECIMAL(12,2)), NULL) AS `雜費金額1`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=1, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',2),'§',-1), NULL) AS `雜費名稱2`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=1, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',2),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額2`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=2, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',3),'§',-1), NULL) AS `雜費名稱3`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=2, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',3),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額3`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=3, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',4),'§',-1), NULL) AS `雜費名稱4`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=3, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',4),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額4`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=4, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',5),'§',-1), NULL) AS `雜費名稱5`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=4, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',5),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額5`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=5, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',6),'§',-1), NULL) AS `雜費名稱6`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=5, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',6),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額6`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=6, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',7),'§',-1), NULL) AS `雜費名稱7`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=6, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',7),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額7`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=7, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',8),'§',-1), NULL) AS `雜費名稱8`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=7, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',8),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額8`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=8, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',9),'§',-1), NULL) AS `雜費名稱9`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=8, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',9),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額9`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=9, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',10),'§',-1), NULL) AS `雜費名稱10`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=9, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',10),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額10`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=10, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',11),'§',-1), NULL) AS `雜費名稱11`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=10, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',11),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額11`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=11, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',12),'§',-1), NULL) AS `雜費名稱12`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=11, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',12),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額12`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=12, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',13),'§',-1), NULL) AS `雜費名稱13`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=12, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',13),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額13`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=13, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',14),'§',-1), NULL) AS `雜費名稱14`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=13, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',14),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額14`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=14, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',15),'§',-1), NULL) AS `雜費名稱15`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=14, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',15),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額15`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=15, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',16),'§',-1), NULL) AS `雜費名稱16`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=15, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',16),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額16`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=16, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',17),'§',-1), NULL) AS `雜費名稱17`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=16, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',17),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額17`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=17, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',18),'§',-1), NULL) AS `雜費名稱18`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=17, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',18),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額18`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=18, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',19),'§',-1), NULL) AS `雜費名稱19`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=18, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',19),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額19`,
                    IF(CHAR_LENGTH(dp)-CHAR_LENGTH(REPLACE(dp,'§',''))>=19, SUBSTRING_INDEX(SUBSTRING_INDEX(dp,'§',20),'§',-1), NULL) AS `雜費名稱20`,
                    IF(CHAR_LENGTH(pp)-CHAR_LENGTH(REPLACE(pp,'§',''))>=19, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',20),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額20`,
                    `運雜費合計`
                FROM temp_raw
            """)

            self.cursor.execute("DROP TEMPORARY TABLE temp_raw")

            self.cursor.execute("SELECT COUNT(*) as cnt FROM `客户月结单TEMP`")
            temp_count = self.cursor.fetchone()['cnt']
            self.append_check_result(f"[OK] 客户月结单TEMP 生成完成 ({temp_count} 条记录)")

            # 步骤3：归一化雜費名稱（代码内执行，不依赖存储过程）
            self.append_check_result(f"\n步骤3: 标准化杂费名称...")
            self._normalize_misc_names("客户月结单TEMP")
            self.append_check_result("[OK] 雜費名稱归一化 完成")

            # 步骤4：直接在Python中查询未标准化的杂费名称和发票号
            self.append_check_result(f"\n步骤4: 检查是否有新增未标准化的杂费名称...")

            # 获取杂费名称列
            self.cursor.execute("""
                SELECT COLUMN_NAME
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME = '客户月结单TEMP'
                  AND COLUMN_NAME LIKE '雜費名稱%'
                ORDER BY ORDINAL_POSITION
            """)
            misc_cols = [r['COLUMN_NAME'] for r in self.cursor.fetchall()]

            # 获取规则
            self.cursor.execute("""
                SELECT pattern_type, keyword
                FROM misc_name_rules ORDER BY sort_order
            """)
            rules = self.cursor.fetchall()

            # 构建正则表达式模式列表
            import re
            rule_patterns = []
            for rule in rules:
                kw = rule['keyword']
                ptype = rule['pattern_type']
                if ptype == 'exact':
                    pattern = f"^{re.escape(kw)}$"
                elif ptype == 'contains':
                    pattern = f".*{re.escape(kw)}.*"
                else:  # prefix
                    pattern = f"^{re.escape(kw)}"
                rule_patterns.append(re.compile(pattern, re.IGNORECASE))

            # 查询每个杂费名称列中未匹配的记录
            unmatched_data = {}  # {(杂费名称): {'count': 0, 'cols': set(), 'invoices': set()}}

            for col in misc_cols:
                # 获取该列所有非空记录
                self.cursor.execute(f"""
                    SELECT DISTINCT `{col}`, `發票號`
                    FROM `客户月结单TEMP`
                    WHERE `{col}` IS NOT NULL AND `{col}` != ''
                """)
                for row in self.cursor.fetchall():
                    misc_name = row.get(col, '')
                    invoice = row.get('發票號', '')
                    if not misc_name or not misc_name.strip():
                        continue

                    # 检查是否匹配任意规则
                    matched = False
                    for pattern in rule_patterns:
                        if pattern.match(misc_name):
                            matched = True
                            break

                    if not matched:
                        if misc_name not in unmatched_data:
                            unmatched_data[misc_name] = {'count': 0, 'cols': set(), 'invoices': set()}
                        unmatched_data[misc_name]['count'] += 1
                        unmatched_data[misc_name]['cols'].add(col)
                        if invoice:
                            unmatched_data[misc_name]['invoices'].add(invoice)

            # 转换为结果列表
            results = []
            for name, data in sorted(unmatched_data.items(), key=lambda x: x[1]['count'], reverse=True):
                invoice_list = sorted(data['invoices'])
                results.append({
                    '雜費名稱': name,
                    '出现次数': data['count'],
                    '涉及列': ', '.join(sorted(data['cols'])),
                    '发票号列表': ', '.join(invoice_list[:10]) + ('...' if len(invoice_list) > 10 else '')
                })

            if not results:
                self.append_check_result(f"\n[OK] 检查结果：无新增未标准化的杂费名称")
                self.append_check_result(f"  所有杂费名称均已标准化")
                self.check_status_var.set(f"✓ 检查完成 - 无新增未标准化项")
            else:
                self.append_check_result(f"\n[警告] 检查结果：发现 {len(results)} 个未标准化的杂费名称\n")
                self.append_check_result("="*80)
                self.append_check_result(f"{'序号':<6}{'杂费名称':<25}{'出现次数':<10}{'涉及列':<20}{'发票号'}")
                self.append_check_result("="*80)
                for idx, row in enumerate(results, 1):
                    name = row.get('雜費名稱', '')
                    count = row.get('出现次数', 0)
                    cols = row.get('涉及列', '')
                    invoices = row.get('发票号列表', '')
                    self.append_check_result(f"{idx:<6}{name:<25}{count:<10}{cols:<20}{invoices}")
                self.append_check_result("="*80)
                self.append_check_result(f"\n建议：请将以上杂费名称添加到规则中，然后更新存储过程")
                self.check_status_var.set(f"⚠ 检查完成 - 发现 {len(results)} 个未标准化项")

            self.conn.commit()

        except Exception as e:
            self.append_check_result(f"\n[错误] 执行失败: {e}")
            self.check_status_var.set("✗ 检查失败")
            messagebox.showerror("错误", f"检查失败: {e}")
            import traceback
            traceback.print_exc()

    def append_check_result(self, text):
        """追加检查结果到文本框"""
        self.check_result_text.configure(state='normal')
        self.check_result_text.insert(tk.END, text + "\n")
        self.check_result_text.see(tk.END)
        self.check_result_text.configure(state='disabled')
        self.root.update()

    # ========== 客户月结单导出方法 ==========

    def browse_export_dir(self):
        """选择输出目录"""
        from tkinter import filedialog
        selected_dir = filedialog.askdirectory(
            title="选择输出目录",
            initialdir=self.export_dir_var.get()
        )
        if selected_dir:
            self.export_dir_var.set(selected_dir)
            self.config['export_dir'] = selected_dir
            self.save_config()

    def export_invoice_excel(self):
        """导出 Invoice 对账单 Excel（按周汇总运杂费）"""
        try:
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            from datetime import datetime, timedelta
            import calendar

            year = int(self.invoice_year_var.get())
            month = int(self.invoice_month_var.get())
            yyyymm = f"{year}{month:02d}"

            if not (1 <= month <= 12):
                messagebox.showerror("错误", "月份必须在 1-12 之间")
                return

            # 检查是否选择了客户
            if not self._invoice_selected_codes:
                messagebox.showwarning("提示", "请先选择要导出的客户")
                return

            # 检查输出目录
            out_dir = self.invoice_dir_var.get()
            if not out_dir:
                messagebox.showwarning("提示", "请先选择输出目录")
                return

            # 解析周分段日期
            periods = []
            for start_entry, end_entry in self._invoice_period_entries:
                s = start_entry.get().strip()
                e = end_entry.get().strip()
                if s and e:
                    periods.append((s, e))

            if not periods:
                messagebox.showwarning("提示", "请至少填写一个周分段的起止日期")
                return

            # 计算日期范围
            date_start = f'{year}-{month:02d}-01'
            if month == 12:
                next_month = datetime(year + 1, 1, 1)
            else:
                next_month = datetime(year, month + 1, 1)
            date_end = next_month.strftime('%Y-%m-%d')
            date_end_display = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')

            # 解析分段日期为实际日期对象（用于SQL过滤）
            # 用户输入只有日（如 01, 10），月份从下拉框获取
            parsed_periods = []
            month_abbr_en = ['','JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
            display_periods = []
            for s, e in periods:
                try:
                    sd = int(s.strip())
                    ed = int(e.strip())
                    if not (1 <= sd <= 31 and 1 <= ed <= 31):
                        continue
                    p_start = datetime(year, month, sd)
                    # 结束日可能跨月/跨年
                    if ed < sd:
                        # 跨月
                        if month == 12:
                            p_end = datetime(year + 1, 1, ed)
                        else:
                            p_end = datetime(year, month + 1, ed)
                    else:
                        p_end = datetime(year, month, ed)
                    parsed_periods.append((p_start, p_end))
                    # 显示格式：日/月（跨月时结束日期用下个月份）
                    display_periods.append((f"{sd:02d}/{month:02d}", f"{ed:02d}/{p_end.month:02d}"))
                except (ValueError, IndexError):
                    continue

            if not parsed_periods:
                messagebox.showwarning("提示", "周分段日期格式不正确，请输入1-31的数字")
                return

            # 清空结果文本框
            self.invoice_result_text.configure(state='normal')
            self.invoice_result_text.delete(1.0, tk.END)
            self.invoice_result_text.configure(state='disabled')

            self.invoice_status_var.set(f"正在导出 Invoice {year}年{month:02d}月...")
            self.root.update()

            def append_log(text):
                self.invoice_result_text.configure(state='normal')
                self.invoice_result_text.insert(tk.END, text + '\n')
                self.invoice_result_text.see(tk.END)
                self.invoice_result_text.configure(state='disabled')
                self.root.update()

            append_log(f"{'='*60}")
            append_log(f"Invoice 对账单导出 - {year}年{month:02d}月")
            append_log(f"{'='*60}")
            append_log(f"日期范围: {date_start} 至 {date_end_display}")
            append_log(f"分段数: {len(parsed_periods)}")
            append_log(f"客户数: {len(self._invoice_selected_codes)}")
            append_log(f"输出目录: {out_dir}\n")

            # 查询所有选中客户的发票运杂费合计
            for cust_code in self._invoice_selected_codes:
                cust_name = self._invoice_selected_names.get(cust_code, cust_code)
                append_log(f"处理客户: {cust_name} [{cust_code}]...")

                try:
                    # 查询该客户当月所有发票的运杂费合计
                    self.cursor.execute("""
                        SELECT im.INVDATE, im.INVOICECODE,
                            COALESCE(im.FEE,0) + COALESCE(
                                SUM(CASE WHEN id.WHOPAY=1 THEN id.PRICE ELSE 0 END), 0
                            ) AS total_fee
                        FROM INVOICE_MASTER im
                        LEFT JOIN INVOICE_DETAIL id ON id.INVOICECODE = im.INVOICECODE
                        WHERE im.INVDATE >= %s AND im.INVDATE < %s
                          AND im.CUSTCODE = %s
                        GROUP BY im.INVOICECODE, im.INVDATE, im.FEE
                        ORDER BY im.INVDATE ASC
                    """, (date_start, date_end, cust_code))
                    invoices = self.cursor.fetchall()
                    invoice_count = len(invoices)

                    if invoice_count == 0:
                        append_log(f"  ⚠ 该客户无发票数据，跳过")
                        continue

                    # 按分段汇总金额
                    period_amounts = []
                    grand_total = 0.0
                    for p_start, p_end in parsed_periods:
                        p_total = 0.0
                        for inv in invoices:
                            inv_date = inv['INVDATE'] if isinstance(inv, dict) else inv[0]
                            if isinstance(inv_date, datetime):
                                inv_date = inv_date.date()
                            elif isinstance(inv_date, str):
                                inv_date = datetime.strptime(inv_date, '%Y-%m-%d').date()
                            p_start_d = p_start.date() if isinstance(p_start, datetime) else p_start
                            p_end_d = p_end.date() if isinstance(p_end, datetime) else p_end
                            if p_start_d <= inv_date <= p_end_d:
                                fee = inv['total_fee'] if isinstance(inv, dict) else inv[2]
                                p_total += float(fee) if fee else 0.0
                        period_amounts.append(p_total)
                        grand_total += p_total

                    # 创建输出目录
                    cust_dir = os.path.join(out_dir, f"{cust_code}{cust_name}")
                    os.makedirs(cust_dir, exist_ok=True)

                    # 生成 Excel
                    wb = Workbook()
                    ws = wb.active
                    ws.title = yyyymm

                    # 列宽（与模板一致）
                    ws.column_dimensions['A'].width = 19
                    ws.column_dimensions['B'].width = 18
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 13
                    ws.column_dimensions['E'].width = 18

                    # 字体样式
                    font_title = Font(name='宋体', size=18, bold=True)
                    font_addr = Font(name='宋体', size=11)
                    font_invoice = Font(name='宋体', size=16, bold=True)
                    font_invno = Font(name='宋体', size=14)
                    font_cust = Font(name='宋体', size=14, bold=True)
                    font_header = Font(name='宋体', size=12, bold=True)
                    font_normal = Font(name='宋体', size=11)
                    font_total = Font(name='宋体', size=11, bold=True)
                    font_small = Font(name='宋体', size=11)

                    align_center = Alignment(horizontal='center', vertical='center')
                    align_right = Alignment(horizontal='right', vertical='center')
                    align_left = Alignment(horizontal='left', vertical='center')
                    num_fmt = '#,##0.00_ '

                    # 边框（按用户修改版标准）
                    thin_side = Side(style='thin')
                    thin_border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                    # 表头行: A左上右, E左上右; 总单数行: A左上, E左上右
                    # 数据行: A左, E左右; 空行最后一行: A左下 + B/C/D下 + E左右下
                    # Sub-Total金额行(E): 左右下

                    # 行高（与模板一致）
                    ws.row_dimensions[1].height = 23.25
                    ws.row_dimensions[2].height = 23.25
                    for r in range(3, 6):
                        ws.row_dimensions[r].height = 12.75
                    ws.row_dimensions[6].height = 30.0
                    for r in range(7, 10):
                        ws.row_dimensions[r].height = 16.5
                    ws.row_dimensions[10].height = 30.0
                    ws.row_dimensions[11].height = 30.0
                    ws.row_dimensions[12].height = 25.5

                    # Row 1: 公司名（繁体）
                    ws.merge_cells('A1:E1')
                    c = ws['A1']
                    c.value = '深華交通實業有限公司'
                    c.font = font_title
                    c.alignment = align_center

                    # Row 2: 英文名
                    ws.merge_cells('A2:E2')
                    c = ws['A2']
                    c.value = 'SUM WAH TRANSPORT ENTERPRISES CO LTD'
                    c.font = font_title
                    c.alignment = align_center

                    # Row 3: 地址1
                    ws.merge_cells('A3:E3')
                    c = ws['A3']
                    c.value = 'RM A19 BLK A 6/F CAMBRIDGE PLAZA,'
                    c.font = font_addr
                    c.alignment = Alignment(horizontal='center')

                    # Row 4: 地址2
                    ws.merge_cells('A4:E4')
                    c = ws['A4']
                    c.value = '188 SAN WAN ROAD SHEUNG SHUI N.T.'
                    c.font = font_addr
                    c.alignment = Alignment(horizontal='center')

                    # Row 5: 联系方式
                    ws.merge_cells('A5:E5')
                    c = ws['A5']
                    c.value = 'E-mail:cw@hksumwah.com  TEL:23689666   FAX:26689135'
                    c.font = font_addr
                    c.alignment = Alignment(horizontal='center')

                    # Row 6: INVOICE
                    ws.merge_cells('A6:E6')
                    c = ws['A6']
                    c.value = 'INVOICE'
                    c.font = font_invoice
                    c.alignment = align_center

                    # Row 7: Invoice No.
                    ws.merge_cells('D7:E7')
                    c = ws['D7']
                    c.value = f'Invoice No.{yyyymm}'
                    c.font = font_invno
                    c.alignment = Alignment(horizontal='right')

                    # Row 8: 致客户 + Date
                    last_day = date_end_display
                    try:
                        ld = datetime.strptime(last_day, '%Y-%m-%d')
                        date_en = f"{ld.day:02d} {month_abbr_en[ld.month]},{ld.year}"
                    except:
                        date_en = last_day

                    c = ws['A8']
                    c.value = f'致：{cust_name}'
                    c.font = font_cust
                    c.alignment = Alignment(horizontal='left')

                    ws.merge_cells('D8:E8')
                    c = ws['D8']
                    c.value = f'Date: {date_en}'
                    c.font = font_invno
                    c.alignment = Alignment(horizontal='right')

                    # Row 9: Attn
                    c = ws['A9']
                    c.value = 'Attn：A/C Dept.'
                    c.font = font_cust
                    c.alignment = Alignment(horizontal='left')

                    # Row 10: LOCAL HAULAGE
                    ws.merge_cells('A10:E10')
                    c = ws['A10']
                    c.value = '*LOCAL HAULAGE*'
                    c.font = font_invno
                    c.alignment = align_center

                    # Row 11: 表头（Description / (HKD)）— 四边外框
                    ws.merge_cells('A11:D11')
                    c = ws['A11']
                    c.value = 'Description'
                    c.font = font_header
                    c.alignment = align_center
                    c.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                    for col in ['B', 'C', 'D']:
                        ws[f'{col}11'].border = Border(top=thin_side, bottom=thin_side)

                    c = ws['E11']
                    c.value = '(HKD)'
                    c.font = font_header
                    c.alignment = align_right
                    c.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                    # Row 12: 总单数
                    ws.merge_cells('A12:D12')
                    c = ws['A12']
                    c.value = f'{month}月份共{invoice_count}張單'
                    c.font = font_header
                    c.alignment = align_left
                    c.border = Border(left=thin_side, top=thin_side)

                    c = ws['E12']
                    c.border = Border(left=thin_side, right=thin_side, top=thin_side)

                    # Row 13+: 每周一行
                    data_start_row = 13
                    num_periods = len(display_periods)
                    for i, ((ds, de), amount) in enumerate(zip(display_periods, period_amounts)):
                        row = data_start_row + i
                        ws.row_dimensions[row].height = 25.5
                        ws.merge_cells(f'A{row}:D{row}')
                        c = ws[f'A{row}']
                        c.value = f'{ds} - {de}'
                        c.font = font_normal
                        c.alignment = align_left
                        c.border = Border(left=thin_side)

                        c = ws[f'E{row}']
                        c.value = amount if amount else 0
                        c.font = font_normal
                        c.alignment = align_right
                        c.number_format = num_fmt
                        c.border = Border(left=thin_side, right=thin_side)

                    # 6行空行（与模板一致）
                    empty_start = data_start_row + num_periods
                    empty_last_row = empty_start + 5  # 空行最后一行，关底边框
                    for r in range(empty_start, empty_start + 6):
                        ws.row_dimensions[r].height = 25.5
                        ws.merge_cells(f'A{r}:D{r}')
                        # 每个空行：A列左边框，E列左右边框
                        ws[f'A{r}'].border = Border(left=thin_side)
                        ws[f'E{r}'].border = Border(left=thin_side, right=thin_side)
                    # 空行最后一行：额外加底部关框
                    ws[f'A{empty_last_row}'].border = Border(left=thin_side, bottom=thin_side)
                    for col in ['B', 'C', 'D']:
                        ws[f'{col}{empty_last_row}'].border = Border(bottom=thin_side)
                    ws[f'E{empty_last_row}'].border = Border(left=thin_side, right=thin_side, bottom=thin_side)

                    # Sub-Total 行（合并A:D，写入左上角A列，右对齐模拟D列居中效果）
                    subtotal_row = empty_start + 6
                    ws.row_dimensions[subtotal_row].height = 30.0
                    ws.merge_cells(f'A{subtotal_row}:D{subtotal_row}')
                    c = ws[f'A{subtotal_row}']
                    c.value = 'Sub-Total(HKD)'
                    c.font = font_normal
                    c.alignment = align_right

                    sub_total = sum(a if a else 0 for a in period_amounts)
                    c = ws[f'E{subtotal_row}']
                    c.value = sub_total
                    c.font = font_total
                    c.alignment = align_right
                    c.number_format = num_fmt
                    c.border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

                    # 签名区（Sub-Total后隔2行开始，与模板一致）
                    sign_start = subtotal_row + 3
                    ws.row_dimensions[sign_start].height = 30.0
                    ws.row_dimensions[sign_start + 1].height = 30.0
                    ws.row_dimensions[sign_start + 5].height = 30.0

                    c = ws[f'C{sign_start}']
                    c.value = '深華交通實業有限公司'
                    c.font = font_small

                    c = ws[f'C{sign_start + 1}']
                    c.value = 'SUM WAH TRANSPORT ENTERPRISES CO LTD'
                    c.font = font_small

                    c = ws[f'C{sign_start + 5}']
                    c.value = 'Company Chop and signature'
                    c.font = font_small

                    # 打印设置
                    last_row = sign_start + 5
                    ws.print_area = f'A1:E{last_row}'
                    ws.page_setup.orientation = 'portrait'
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4

                    # 保存文件
                    filename = f'invoice_table_{cust_code}_{yyyymm}.xlsx'
                    filepath = os.path.join(cust_dir, filename)
                    wb.save(filepath)
                    append_log(f"  ✅ {filename}（{invoice_count}张单，合计 {grand_total:,.2f}）")

                except Exception as ex:
                    append_log(f"  ❌ 导出失败: {str(ex)}")

            append_log(f"\n{'='*60}")
            append_log(f"导出完成！共处理 {len(self._invoice_selected_codes)} 个客户")
            append_log(f"输出目录: {out_dir}")
            append_log(f"{'='*60}")
            self.invoice_status_var.set("导出完成")

        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
            self.invoice_status_var.set("导出失败")

    def export_monthly_excel(self):
        """导出客户月结单Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import os

            year = int(self.export_year_combo.get())
            month = int(self.export_month_combo.get())

            if not (1 <= month <= 12):
                messagebox.showerror("错误", "月份必须在 1-12 之间")
                return

            yyyymm = f"{year}{month:02d}"
            year_cn = f"{year}年{month}月"

            # 清空结果文本框
            self.export_result_text.configure(state='normal')
            self.export_result_text.delete(1.0, tk.END)
            self.export_result_text.configure(state='disabled')

            self.export_status_var.set(f"正在导出 {year_cn}...")
            self.root.update()

            # 计算日期范围
            from datetime import datetime, timedelta
            date_start = f'{year}-{month:02d}-01'
            if month == 12:
                next_month = datetime(year + 1, 1, 1)
            else:
                next_month = datetime(year, month + 1, 1)
            # 修复：SQL 用 INVDATE < date_end，date_end 须为下个月第一天
            # 才能正确包含当月最后一天（如3月31日）
            date_end = next_month.strftime('%Y-%m-%d')
            # 仅用于日志显示的月末最后一天
            date_end_display = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')

            self.append_export_result(f"{'='*60}")
            self.append_export_result(f"客户月结单导出 - {year_cn}")
            self.append_export_result(f"{'='*60}\n")
            self.append_export_result(f"日期范围: {date_start} 至 {date_end_display}")

            # 步骤1：生成客户月结单TEMP
            self.append_export_result(f"\n步骤1/4: 生成客户月结单TEMP...")
            self.cursor.execute("DROP TABLE IF EXISTS `客户月结单TEMP`")

            sql_temp = """
            CREATE TABLE `客户月结单TEMP` AS
            SELECT
                t.INVDATE AS `日期`, t.INVOICECODE AS `發票號`, t.CUSTCODE AS `客戶編號`,
                COALESCE(cm.NAME, '') AS `客戶名稱`,
                t.NAME AS `司機姓名`, t.HKCP AS `香港車牌`, t.SZCP AS `大陸車牌`,
                t.DEST AS `地區`, t.CONCODE AS `櫃號`, t.SIZE AS `櫃尺碼`,
                t.TAKENO AS `提單號`, t.SHIP AS `船名`, t.SHIPCODE AS `托運號`, t.FEE AS `運費`,
                IF(t.dp IS NOT NULL, SUBSTRING_INDEX(t.dp,'§',1), NULL) AS `雜費名稱1`,
                IF(t.pp IS NOT NULL, CAST(SUBSTRING_INDEX(t.pp,'§',1) AS DECIMAL(12,2)), NULL) AS `雜費金額1`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=1, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',2),'§',-1), NULL) AS `雜費名稱2`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=1, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',2),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額2`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=2, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',3),'§',-1), NULL) AS `雜費名稱3`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=2, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',3),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額3`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=3, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',4),'§',-1), NULL) AS `雜費名稱4`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=3, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',4),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額4`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=4, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',5),'§',-1), NULL) AS `雜費名稱5`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=4, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',5),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額5`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=5, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',6),'§',-1), NULL) AS `雜費名稱6`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=5, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',6),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額6`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=6, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',7),'§',-1), NULL) AS `雜費名稱7`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=6, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',7),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額7`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=7, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',8),'§',-1), NULL) AS `雜費名稱8`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=7, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',8),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額8`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=8, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',9),'§',-1), NULL) AS `雜費名稱9`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=8, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',9),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額9`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=9, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',10),'§',-1), NULL) AS `雜費名稱10`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=9, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',10),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額10`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=10, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',11),'§',-1), NULL) AS `雜費名稱11`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=10, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',11),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額11`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=11, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',12),'§',-1), NULL) AS `雜費名稱12`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=11, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',12),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額12`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=12, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',13),'§',-1), NULL) AS `雜費名稱13`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=12, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',13),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額13`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=13, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',14),'§',-1), NULL) AS `雜費名稱14`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=13, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',14),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額14`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=14, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',15),'§',-1), NULL) AS `雜費名稱15`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=14, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',15),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額15`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=15, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',16),'§',-1), NULL) AS `雜費名稱16`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=15, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',16),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額16`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=16, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',17),'§',-1), NULL) AS `雜費名稱17`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=16, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',17),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額17`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=17, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',18),'§',-1), NULL) AS `雜費名稱18`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=17, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',18),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額18`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=18, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',19),'§',-1), NULL) AS `雜費名稱19`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=18, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',19),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額19`,
                IF(CHAR_LENGTH(t.dp)-CHAR_LENGTH(REPLACE(t.dp,'§',''))>=19, SUBSTRING_INDEX(SUBSTRING_INDEX(t.dp,'§',20),'§',-1), NULL) AS `雜費名稱20`,
                IF(CHAR_LENGTH(t.pp)-CHAR_LENGTH(REPLACE(t.pp,'§',''))>=19, CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(pp,'§',20),'§',-1) AS DECIMAL(12,2)), NULL) AS `雜費金額20`,
                t.`運雜費合計`
            FROM (
                SELECT
                    im.INVDATE, im.INVOICECODE, im.CUSTCODE, cd.DRIVER,
                    dc.NAME, dc.HKCP, dc.SZCP, im.DEST, cd.CONCODE, cd.SIZE, cd.TAKENO, im.SHIP, im.SHIPCODE, im.FEE,
                    GROUP_CONCAT(CASE WHEN id.WHOPAY=1 THEN id.DESCR END ORDER BY id.NN ASC SEPARATOR '§') AS dp,
                    GROUP_CONCAT(CASE WHEN id.WHOPAY=1 THEN id.PRICE END ORDER BY id.NN ASC SEPARATOR '§') AS pp,
                    COALESCE(im.FEE,0) + COALESCE(SUM(CASE WHEN id.WHOPAY=1 THEN id.PRICE ELSE 0 END),0) AS `運雜費合計`
                FROM INVOICE_MASTER im
                LEFT JOIN CON_DETAIL cd ON cd.INVOICECODE = im.INVOICECODE AND cd.NN = 1
                LEFT JOIN DRIVER_CP dc ON dc.CUSTCODE = cd.DRIVER
                LEFT JOIN INVOICE_DETAIL id ON id.INVOICECODE = im.INVOICECODE
                WHERE im.INVDATE >= %s AND im.INVDATE < %s
                GROUP BY im.INVOICECODE, im.INVDATE, im.CUSTCODE, cd.DRIVER, dc.NAME, dc.HKCP, dc.SZCP,
                    im.DEST, cd.CONCODE, cd.SIZE, cd.TAKENO, im.SHIP, im.SHIPCODE, im.FEE
            ) t
            LEFT JOIN CUST_MASTER cm ON cm.CUSTCODE = t.CUSTCODE
            ORDER BY t.INVDATE ASC, t.INVOICECODE ASC
            """

            self.cursor.execute(sql_temp, (date_start, date_end))
            self.cursor.execute("ALTER TABLE `客户月结单TEMP` ADD COLUMN `id` INT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY FIRST")
            self.conn.commit()

            self.cursor.execute("SELECT COUNT(*) as cnt FROM `客户月结单TEMP`")
            temp_count = self.cursor.fetchone()['cnt']
            self.append_export_result(f"[OK] 客户月结单TEMP 生成完成 ({temp_count} 条记录)")

            # 步骤2：归一化雜費名稱（代码内执行，不依赖存储过程）
            self.append_export_result(f"\n步骤2/4: 归一化雜費名稱...")
            self._normalize_misc_names("客户月结单TEMP")
            self.conn.commit()
            self.append_export_result("[OK] 雜費名稱归一化 完成")

            # 步骤3：执行 pivot（代码内执行，不依赖存储过程）
            self.append_export_result(f"\n步骤3/4: 执行 pivot_misc_names...")
            self._pivot_misc_names()
            self.conn.commit()

            self.cursor.execute("SELECT COUNT(*) as cnt FROM `客户月结单PIVOT`")
            pivot_count = self.cursor.fetchone()['cnt']
            self.append_export_result(f"[OK] pivot_misc_names 完成 ({pivot_count} 条记录)")

            # 步骤4：导出Excel
            self.append_export_result(f"\n步骤4/4: 导出Excel...")

            # 查询客户列表
            self.cursor.execute("""
                SELECT DISTINCT 客戶編號, 客戶名稱
                FROM 客户月结单PIVOT
                WHERE 客戶編號 IS NOT NULL AND 客戶編號 != ''
                ORDER BY 客戶編號
            """)
            customers_all = self.cursor.fetchall()

            # 若选择模式，弹出多选对话框
            if self.cust_export_mode_var.get() == 'select':
                items = []
                for c in customers_all:
                    code = c['客戶編號'] if isinstance(c, dict) else c[0]
                    name = c['客戶名稱'] if isinstance(c, dict) else c[1]
                    items.append((code, f"{name}  [{code}]"))
                selected_codes = self.show_multiselect_dialog("选择要导出的客户", items)
                if selected_codes is None:
                    self.export_status_var.set("已取消")
                    return
                if not selected_codes:
                    messagebox.showwarning("提示", "未选择任何客户，导出取消")
                    self.export_status_var.set("已取消")
                    return
                selected_set = set(selected_codes)
                customers = [c for c in customers_all if (c['客戶編號'] if isinstance(c, dict) else c[0]) in selected_set]
            else:
                customers = customers_all

            # 固定列
            FIXED_COLS = ['日期', '發票號', '司機姓名', '香港車牌', '大陸車牌', '地區', '櫃號', '櫃尺碼', '提單號', '船名', '托運號', '運費']
            TOTAL_COL = '運雜費合計'

            # 样式函数
            def font_func(size=11, bold=False):
                return Font(name='宋体', size=size, bold=bold)

            def align_func(horizontal=None, vertical=None, wrap=False):
                return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

            def thin_border_func():
                s = Side(style='thin')
                return Border(top=s, bottom=s, left=s, right=s)

            # 输出目录（重复年月时自动加后缀 A/B/C...）
            try:
                base_dir = self.export_dir_var.get()
                out_dir_base = os.path.join(base_dir, f'客户月结单{yyyymm}')
                out_dir = out_dir_base
                suffix_idx = 0
                while os.path.exists(out_dir):
                    suffix_idx += 1
                    out_dir = out_dir_base + chr(ord('A') + suffix_idx - 1)
                os.makedirs(out_dir, exist_ok=True)
            except:
                messagebox.showerror("错误", f"无法创建输出目录: {base_dir}")
                return

            # 主循环
            exported = 0
            for customer in customers:
                # 由于使用 DictCursor，customers 是字典列表
                if isinstance(customer, dict):
                    cust_code = customer.get('客戶編號')
                    cust_name = customer.get('客戶名稱')
                else:
                    cust_code, cust_name = customer

                # 查询数据（指定列名，避免使用 *）
                self.cursor.execute("""
                    SELECT 日期, 發票號, 司機姓名, 香港車牌, 大陸車牌, 地區, 櫃號, 櫃尺碼, 提單號, 船名, 托運號, 運費, 運雜費合計
                    FROM 客户月结单PIVOT
                    WHERE 客戶編號 = %s AND 日期 >= %s AND 日期 < %s
                    ORDER BY 日期, 發票號
                """, (cust_code, date_start, date_end))

                rows = self.cursor.fetchall()

                # 调试：输出客户编码和数据行数
                self.append_export_result(f"DEBUG: 客户 {cust_code} ({cust_name}), 数据行数: {len(rows)}, 类型: {type(rows)}, 首行类型: {type(rows[0]) if rows else 'N/A'}")

                if not rows:
                    self.append_export_result(f"DEBUG: 跳过客户 {cust_code}, 无数据")
                    continue

                # 获取数据库返回的列名
                column_names = [desc[0] for desc in self.cursor.description]

                # 获取所有杂费列（从 FIXED_COLS 和 運雜費合計 之外的列）
                self.cursor.execute("""
                    SELECT COLUMN_NAME
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_SCHEMA = 'SWT' AND TABLE_NAME = '客户月结单PIVOT'
                    AND COLUMN_NAME NOT IN ('_id', 'id', '客戶編號', '客戶名稱', '日期', '發票號', '司機姓名', '香港車牌', '大陸車牌', '地區', '櫃號', '櫃尺碼', '提單號', '船名', '托運號', '運費', '運雜費合計')
                    ORDER BY ORDINAL_POSITION
                """)
                misc_cols_result = self.cursor.fetchall()
                # 处理 DictCursor 返回的字典或元组
                all_misc_cols = []
                for row in misc_cols_result:
                    if isinstance(row, dict):
                        all_misc_cols.append(row.get('COLUMN_NAME'))
                    else:
                        all_misc_cols.append(row[0])

                # 查询完整的杂费数据（列名可能含 %，需转义为 %% 防 pymysql 格式化冲突）
                misc_cols_str = ', '.join([f'`{col.replace("%", "%%")}`' for col in all_misc_cols])
                self.cursor.execute(f"""
                    SELECT 日期, 發票號, 司機姓名, 香港車牌, 大陸車牌, 地區, 櫃號, 櫃尺碼, 提單號, 船名, 托運號, 運費, {misc_cols_str}, 運雜費合計
                    FROM 客户月结单PIVOT
                    WHERE 客戶編號 = %s AND 日期 >= %s AND 日期 < %s
                    ORDER BY 日期, 發票號
                """, (cust_code, date_start, date_end))

                rows = self.cursor.fetchall()
                column_names = [desc[0] for desc in self.cursor.description]

                self.append_export_result(f"DEBUG: 列名: {column_names[:5]}... (共{len(column_names)}列)")

                if not rows:
                    continue

                # 按發票號去重：發票號唯一，只保留每张发票的第一行
                seen_invoices = set()
                deduped_rows = []
                for _r in rows:
                    inv = _r.get('發票號') if isinstance(_r, dict) else None
                    if inv is None:
                        # 元组情况：找 發票號 在 column_names 中的位置
                        try:
                            inv = _r[column_names.index('發票號')]
                        except (ValueError, IndexError):
                            inv = None
                    if inv not in seen_invoices:
                        seen_invoices.add(inv)
                        deduped_rows.append(_r)
                if len(deduped_rows) < len(rows):
                    self.append_export_result(f"  [去重] 客户 {cust_code}: 原 {len(rows)} 行 → 去重后 {len(deduped_rows)} 行")
                rows = deduped_rows

                # 由于使用 DictCursor，rows 是字典列表，直接创建 DataFrame
                if isinstance(rows[0], dict):
                    df = pd.DataFrame(rows)
                else:
                    df = pd.DataFrame(rows, columns=column_names)
                cols = df.columns.tolist()

                # 确定有数据的金额列（排除固定列，只对数值列求和）
                amount_cols = []
                for col in cols:
                    if col not in FIXED_COLS and col != TOTAL_COL:
                        # 检查是否为数值类型列
                        if df[col].dtype in ['float64', 'int64', 'object']:
                            # 对于object类型，尝试转换为数值
                            try:
                                numeric_series = pd.to_numeric(df[col], errors='coerce')
                                if numeric_series.fillna(0).sum() > 0:
                                    amount_cols.append(col)
                            except:
                                # 如果无法转换为数值，跳过
                                pass

                all_cols = FIXED_COLS + amount_cols + [TOTAL_COL]

                # 生成带序号的文件名，避免覆盖
                base_filename = f'{cust_name}{yyyymm}.xlsx'
                filepath = os.path.join(out_dir, base_filename)
                suffix = ''
                suffix_idx = ord('A')  # 从 'A' 开始

                # 如果文件已存在，添加后缀
                while os.path.exists(filepath):
                    suffix = chr(suffix_idx)
                    filepath = os.path.join(out_dir, f'{cust_name}{yyyymm}{suffix}.xlsx')
                    suffix_idx += 1
                    # 超过 'Z' 后重置为 'A'
                    if suffix_idx > ord('Z'):
                        suffix_idx = ord('A')
                        suffix = ''  # 清空，下次循环会添加新的前缀（如 AA, AB）

                # 创建Workbook
                wb = Workbook()
                ws = wb.active
                last_col_letter = get_column_letter(len(all_cols))

                # 第1行：公司名称 + 年月
                ws.merge_cells(f'A1:{last_col_letter}1')
                c = ws['A1']
                c.value = f'深華交通實業有限公司  {year_cn} 對帳單'
                c.font = font_func(18, bold=True)
                c.alignment = align_func('center')
                ws.row_dimensions[1].height = 22.5

                # 第2-5行：公司信息
                info_rows = [
                    '香港地址:香港新界上水新運路188號劍橋廣場A619室  香港電話:23689666 香港傳真:26689135',
                    '深圳地址:深圳市福田区皇岗口岸皇御苑一栋205室    深圳电话:83859938 深圳传真:83858928',
                    '香港中銀戶名:SUM WAH TRANSPORT ENTERPRISES CO LTD 帳号:01960410087700',
                    '深圳市工商银行皇岗支行人民币户名:深圳市深华运输有限公司 帳号:4000026319200117006'
                ]

                for i, text in enumerate(info_rows, start=2):
                    ws.merge_cells(f'A{i}:{last_col_letter}{i}')
                    c = ws[f'A{i}']
                    c.value = text
                    c.font = font_func(11)
                    c.alignment = align_func('center')

                # 第6行：客户名称
                ws[f'A6'].value = f'客戶名稱: {cust_name}'
                ws[f'A6'].font = font_func(11)

                # 第7行：表头
                ws.row_dimensions[7].height = 30
                for col_idx, col_name in enumerate(all_cols, start=1):
                    c = ws.cell(row=7, column=col_idx)
                    c.value = col_name
                    c.font = font_func(11, bold=(col_idx <= 12))
                    c.alignment = align_func('center', 'center')
                    c.border = thin_border_func()

                # 第8行起：数据行
                DATA_START = 8

                # 调试信息
                if cust_name == '深圳神彩物流有限公司':
                    self.append_export_result(f"  调试: cols={len(cols)}, all_cols={len(all_cols)}, rows={len(rows)}")
                    if len(rows) > 0:
                        self.append_export_result(f"  调试: 第一行数据 keys={row_dict.keys() if 'row_dict' in locals() else 'N/A'}")

                num_cols = amount_cols + ['運費']
                col_totals = {ac: 0.0 for ac in num_cols}
                grand_total = 0.0
                for row_offset, row_data in enumerate(rows):
                    r = DATA_START + row_offset
                    ws.row_dimensions[r].height = 30

                    # 创建列名到索引的映射
                    # 如果 row_data 是字典，直接使用；如果是元组，则用 zip 转换
                    if isinstance(row_data, dict):
                        row_dict = row_data
                    else:
                        row_dict = dict(zip(cols, row_data))

                    row_total = 0.0
                    for ac in num_cols:
                        try:
                            v = float(row_dict.get(ac, 0) or 0)
                            row_total += v
                            col_totals[ac] += v
                        except (ValueError, TypeError):
                            pass
                    grand_total += row_total

                    for col_idx, col_name in enumerate(all_cols, start=1):
                        c = ws.cell(row=r, column=col_idx)
                        c.font = font_func(11)
                        c.border = thin_border_func()

                        if col_name == TOTAL_COL:
                            c.value = row_total
                            c.alignment = align_func('right', 'center')
                            c.number_format = '#,##0.00'
                        elif col_name in amount_cols:
                            # 金额列
                            val = row_dict.get(col_name, 0)
                            if val is not None and val != '':
                                try:
                                    fv = float(val)
                                    c.value = '' if fv == 0 else fv
                                except (TypeError, ValueError):
                                    c.value = ''
                            else:
                                c.value = ''
                            c.alignment = align_func('right', 'center', wrap=True)
                            c.number_format = '#,##0.00'
                        elif col_name == '運費':
                            # 運費列：为0时显示0（不能为空）
                            val = row_dict.get(col_name, 0)
                            if val is not None and val != '':
                                try:
                                    fv = float(val)
                                    c.value = '0.00' if fv == 0 else fv
                                except (TypeError, ValueError):
                                    c.value = '0.00'
                            else:
                                c.value = 0
                            c.alignment = align_func('right', 'center', wrap=True)
                            c.number_format = '#,##0.00'
                        else:
                            # 其他列
                            val = row_dict.get(col_name, '')
                            if col_name == '日期' and val is not None and val != '':
                                c.value = str(val)[:10] if hasattr(val, '__str__') else str(val)
                            elif val is not None and val != '':
                                c.value = str(val)
                            else:
                                c.value = ''
                            c.alignment = align_func(None, 'center', wrap=True)

                # 合计行
                total_row = DATA_START + len(rows)
                ws.cell(row=total_row, column=1).value = '合計'
                ws.cell(row=total_row, column=1).font = font_func(11)
                ws.cell(row=total_row, column=1).border = thin_border_func()

                for col_idx in range(2, len(all_cols) + 1):
                    c = ws.cell(row=total_row, column=col_idx)
                    c.font = font_func(11)
                    c.border = thin_border_func()
                    c.alignment = align_func('right', 'center')
                    col_name = all_cols[col_idx - 1]

                    if col_name == TOTAL_COL:
                        c.value = grand_total
                        c.number_format = '#,##0.00'
                    elif col_name == '運費' or col_name in amount_cols:
                        c.value = col_totals.get(col_name, 0.0)
                        c.number_format = '#,##0.00'
                        c.number_format = '#,##0.00'

                # 动态调整列宽（真正自适应）
                for col_idx, col_name in enumerate(all_cols, start=1):
                    col_letter = get_column_letter(col_idx)

                    # 收集该列所有单元格的内容
                    all_cell_values = []

                    # 表头
                    all_cell_values.append(str(col_name))

                    # 数据行
                    for row_data in rows:
                        # 如果 row_data 是字典，直接使用；如果是元组，转换为字典
                        if isinstance(row_data, dict):
                            val = row_data.get(col_name)
                        else:
                            row_dict = dict(zip(cols, row_data))
                            val = row_dict.get(col_name)

                        if col_name in amount_cols + [TOTAL_COL, '運費']:
                            # 金额列：格式化为带千分位的字符串
                            if val is not None and val != '':
                                try:
                                    num_val = float(val)
                                    if num_val != 0:
                                        formatted = f'{num_val:,.2f}'  # 格式化为 1,234.56
                                        all_cell_values.append(formatted)
                                    else:
                                        all_cell_values.append('0.00')
                                except:
                                    all_cell_values.append(str(val))
                        else:
                            # 文本列
                            if val is not None and val != '':
                                cell_str = str(val)
                                if col_name == '日期':
                                    cell_str = cell_str[:10]  # 日期只取前10位
                                # 处理换行，取最长的行
                                lines = cell_str.split('\n')
                                all_cell_values.extend(lines)

                    # 合计行
                    total_cell_value = ws.cell(row=total_row, column=col_idx).value
                    if total_cell_value is not None and total_cell_value != '':
                        if col_name in amount_cols + [TOTAL_COL, '運費']:
                            # 金额合计：估算一个更大的值
                            all_cell_values.append('999,999.99')
                        else:
                            total_str = str(total_cell_value)
                            lines = total_str.split('\n')
                            all_cell_values.extend(lines)

                    # 運雜費合計列特殊处理：考虑所有金额列的最大值
                    if col_name == TOTAL_COL:
                        # 收集从運費开始到運雜費合計之前的所有金额列的最大值
                        for row_data in rows:
                            # 如果 row_data 是字典，直接使用；如果是元组，转换为字典
                            if isinstance(row_data, dict):
                                row_dict = row_data
                            else:
                                row_dict = dict(zip(cols, row_data))
                            # 检查从運費开始的所有金额列
                            for check_col in ['運費'] + amount_cols:
                                val = row_dict.get(check_col)
                                if val is not None and val != '':
                                    try:
                                        num_val = float(val)
                                        if num_val != 0:
                                            formatted = f'{num_val:,.2f}'
                                            all_cell_values.append(formatted)
                                    except:
                                        pass

                # 计算最大宽度（所有列都采用自适应列宽）
                if all_cell_values:
                    max_width = max(len(cell) for cell in all_cell_values)

                    # 添加适当padding
                    # 文本列padding小一些，金额列需要更多空间
                    if col_name in amount_cols + [TOTAL_COL, '運費']:
                        padding = 3  # 金额列需要更多padding
                    else:
                        padding = 2  # 文本列padding小一些

                    final_width = max_width + padding

                    # 限制最小和最大宽度
                    final_width = max(final_width, 8)  # 最小8
                    final_width = min(final_width, 80)  # 最大80
                else:
                    final_width = 10  # 默认宽度

                ws.column_dimensions[col_letter].width = final_width

                # 尝试保存工作簿
                try:
                    wb.save(filepath)
                    self.append_export_result(f"OK {cust_name}{yyyymm}.xlsx  ({len(rows)}行, {len(amount_cols)+1}个金额列)")
                    exported += 1
                except Exception as save_error:
                    self.append_export_result(f"[错误] 保存文件失败 {cust_name}{yyyymm}.xlsx: {save_error}")
                    import traceback
                    self.append_export_result(f"详细错误: {traceback.format_exc()}")
                    continue

            self.append_export_result(f"\n{'='*60}")
            self.append_export_result(f"全部完成！")
            self.append_export_result(f"{'='*60}")
            self.append_export_result(f"步骤1: 客户月结单TEMP 生成完成 ({temp_count} 条记录)")
            self.append_export_result(f"步骤2: normalize_misc_names 完成")
            self.append_export_result(f"步骤3: pivot_misc_names 完成 ({pivot_count} 条记录)")
            self.append_export_result(f"步骤4: Excel 导出完成 ({exported} 个文件)")
            self.append_export_result(f"年月: {year_cn}")
            base_dir = self.export_dir_var.get()
            self.append_export_result(f"输出目录: {out_dir}")

            self.export_status_var.set(f"✓ 导出完成 - {exported} 个文件")

        except ImportError as e:
            messagebox.showerror("错误", f"缺少必要的库: {e}\n\n请安装: pip install pandas openpyxl")
            self.export_status_var.set("✗ 导出失败")
        except Exception as e:
            self.append_export_result(f"\n[错误] 导出失败: {e}")
            self.export_status_var.set("✗ 导出失败")
            import traceback
            error_detail = traceback.format_exc()
            self.append_export_result(f"\n详细错误信息:\n{error_detail}")
            messagebox.showerror("错误", f"导出失败: {e}\n\n详细信息:\n{error_detail}")

    def browse_summary_dir(self):
        """选择客户汇总表导出目录"""
        from tkinter import filedialog
        selected_dir = filedialog.askdirectory(
            title="选择导出目录",
            initialdir=self.summary_dir_var.get()
        )
        if selected_dir:
            self.summary_dir_var.set(selected_dir)
            self.config['summary_dir'] = selected_dir
            self.save_config()

    def export_customer_summary_gui(self):
        """从GUI调用客户汇总表导出"""
        try:
            year = int(self.summary_year_combo.get())
            month = int(self.summary_month_combo.get())

            if not (1 <= month <= 12):
                messagebox.showerror("错误", "月份必须在 1-12 之间")
                return

            yyyymm = f"{year}{month:02d}"
            date_start = f"{year}-{month:02d}-01"

            from datetime import datetime, timedelta
            if month == 12:
                next_month = datetime(year + 1, 1, 1)
            else:
                next_month = datetime(year, month + 1, 1)
            date_end = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')

            self.summary_status_var.set("正在导出...")
            self.root.update()

            # 清空结果文本框
            self.summary_result_text.configure(state='normal')
            self.summary_result_text.delete(1.0, tk.END)
            self.summary_result_text.configure(state='disabled')

            # 显示导出信息
            self.append_summary_result(f"{'='*60}")
            self.append_summary_result(f"客户汇总表导出 - {year}年{month}月")
            self.append_summary_result(f"{'='*60}")
            self.append_summary_result(f"日期范围: {date_start} 至 {date_end}")

            # 执行导出
            result = self.export_customer_summary_core(year, month, date_start, date_end)

            self.append_summary_result(f"\n✓ 导出完成!")
            self.append_summary_result(f"文件路径: {result}")
            self.append_summary_result(f"\n{'='*60}")

            self.summary_status_var.set("准备就绪")

        except ValueError as e:
            self.append_summary_result(f"\n[错误] 输入格式错误: {e}")
            self.summary_status_var.set("准备就绪")
        except Exception as e:
            self.append_summary_result(f"\n[错误] 导出失败: {e}")
            import traceback
            error_detail = traceback.format_exc()
            self.append_summary_result(f"\n详细错误信息:\n{error_detail}")
            self.summary_status_var.set("准备就绪")

    def export_customer_summary_core(self, year, month, date_start, date_end):
        """核心导出逻辑 - 复制 export_customer_summary.py 的逻辑"""
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import timedelta

        # 删除旧临时表
        self.cursor.execute("DROP TABLE IF EXISTS `客户汇总TEMP`")

        # 创建临时表（使用参数化查询防止SQL注入）
        sql = """
CREATE TABLE `客户汇总TEMP` AS
SELECT
    a.CUSTCODE                                              AS `客户编号`,
    (SELECT MIN(cm.NAME) FROM CUST_MASTER cm WHERE cm.CUSTCODE = a.CUSTCODE) AS `客户名称`,
    COALESCE(a.fee_sum, 0)                                  AS `运费`,
    COALESCE(b.overtime_sum, 0)                             AS `超时费等`,
    COALESCE(a.fee_sum, 0) + COALESCE(b.overtime_sum, 0)   AS `运费合计`,
    COALESCE(b.misc_income_sum, 0)                          AS `杂费收入`,
    COALESCE(b.misc_sum, 0)                                 AS `杂费`,
    COALESCE(a.fee_sum, 0) + COALESCE(b.overtime_sum, 0)
    + COALESCE(b.misc_income_sum, 0) + COALESCE(b.misc_sum, 0) AS `运杂费合计`,
    COALESCE(a.agent_sum, 0)                                AS `代理费`,
    COALESCE(a.comp_fee_sum, 0)                             AS `公司杂费`,
    COALESCE(a.driver_comm_sum, 0)                          AS `公司运费`,
    COALESCE(a.driver_rec_amt_sum, 0)                       AS `司机杂费`,
    COALESCE(a.fee_sum, 0) + COALESCE(b.overtime_sum, 0)
    - COALESCE(a.agent_sum, 0) - COALESCE(a.comp_fee_sum, 0)
    - COALESCE(a.driver_comm_sum, 0) - COALESCE(a.driver_rec_amt_sum, 0) AS `净运费`
FROM (
    SELECT
        im.CUSTCODE,
        SUM(im.FEE)            AS fee_sum,
        SUM(im.AGENT_FEE)      AS agent_sum,
        SUM(im.COMP_FEE)       AS comp_fee_sum,
        SUM(im.DRIVER_COMM)    AS driver_comm_sum,
        SUM(im.DRIVER_REC_AMT) AS driver_rec_amt_sum
    FROM INVOICE_MASTER im
    WHERE im.INVDATE BETWEEN %s AND %s
    GROUP BY im.CUSTCODE
) a
LEFT JOIN (
    SELECT
        im2.CUSTCODE,
        SUM(CASE WHEN TRIM(id.DESCR) LIKE '%%.'
                  AND id.DRIVERCODE <> '021'
             THEN id.PRICE ELSE 0 END)                      AS overtime_sum,
        SUM(CASE WHEN id.DRIVERCODE = '021'
                  AND TRIM(id.DESCR) LIKE '%%.'
             THEN id.PRICE ELSE 0 END)                      AS misc_income_sum,
        SUM(CASE WHEN (id.WHOPAY = 1 AND TRIM(id.DESCR) NOT LIKE '%%.')
                       OR (id.WHOPAY = 0 AND TRIM(id.DESCR) NOT LIKE '%%.')
             THEN id.PRICE ELSE 0 END)                      AS misc_sum
    FROM INVOICE_MASTER im2
    JOIN INVOICE_DETAIL id ON id.INVOICECODE = im2.INVOICECODE
    WHERE im2.INVDATE BETWEEN %s AND %s
    GROUP BY im2.CUSTCODE
) b ON b.CUSTCODE = a.CUSTCODE

UNION ALL
SELECT
    '合计'                                                  AS `客户编号`,
    '合    计'                                              AS `客户名称`,
    SUM(COALESCE(a2.fee_sum, 0))                            AS `运费`,
    SUM(COALESCE(b2.overtime_sum, 0))                       AS `超时费等`,
    SUM(COALESCE(a2.fee_sum, 0) + COALESCE(b2.overtime_sum, 0)) AS `运费合计`,
    SUM(COALESCE(b2.misc_income_sum, 0))                    AS `杂费收入`,
    SUM(COALESCE(b2.misc_sum, 0))                           AS `杂费`,
    SUM(COALESCE(a2.fee_sum, 0) + COALESCE(b2.overtime_sum, 0)
      + COALESCE(b2.misc_income_sum, 0) + COALESCE(b2.misc_sum, 0)) AS `运杂费合计`,
    SUM(COALESCE(a2.agent_sum, 0))                          AS `代理费`,
    SUM(COALESCE(a2.comp_fee_sum, 0))                       AS `公司杂费`,
    SUM(COALESCE(a2.driver_comm_sum, 0))                    AS `公司运费`,
    SUM(COALESCE(a2.driver_rec_amt_sum, 0))                 AS `司机杂费`,
    SUM(COALESCE(a2.fee_sum, 0) + COALESCE(b2.overtime_sum, 0)
      - COALESCE(a2.agent_sum, 0) - COALESCE(a2.comp_fee_sum, 0)
      - COALESCE(a2.driver_comm_sum, 0) - COALESCE(a2.driver_rec_amt_sum, 0)) AS `净运费`
FROM (
    SELECT im.CUSTCODE,
           SUM(im.FEE)            AS fee_sum,
           SUM(im.AGENT_FEE)      AS agent_sum,
           SUM(im.COMP_FEE)       AS comp_fee_sum,
           SUM(im.DRIVER_COMM)    AS driver_comm_sum,
           SUM(im.DRIVER_REC_AMT) AS driver_rec_amt_sum
    FROM INVOICE_MASTER im
    WHERE im.INVDATE BETWEEN %s AND %s
    GROUP BY im.CUSTCODE
) a2
LEFT JOIN (
    SELECT im2.CUSTCODE,
           SUM(CASE WHEN TRIM(id.DESCR) LIKE '%%.'
                     AND id.DRIVERCODE <> '021'
                THEN id.PRICE ELSE 0 END)                   AS overtime_sum,
           SUM(CASE WHEN id.DRIVERCODE = '021'
                     AND TRIM(id.DESCR) LIKE '%%.'
                THEN id.PRICE ELSE 0 END)                   AS misc_income_sum,
           SUM(CASE WHEN (id.WHOPAY = 1 AND TRIM(id.DESCR) NOT LIKE '%%.')
                          OR (id.WHOPAY = 0 AND TRIM(id.DESCR) NOT LIKE '%%.')
                THEN id.PRICE ELSE 0 END)                   AS misc_sum
    FROM INVOICE_MASTER im2
    JOIN INVOICE_DETAIL id ON id.INVOICECODE = im2.INVOICECODE
    WHERE im2.INVDATE BETWEEN %s AND %s
    GROUP BY im2.CUSTCODE
) b2 ON b2.CUSTCODE = a2.CUSTCODE
ORDER BY (`客户编号` = '') DESC, `客户编号` ASC
"""
        # 参数化查询（4个 BETWEEN 子句，需要8个参数）
        params = (date_start, date_end, date_start, date_end, date_start, date_end, date_start, date_end)
        self.cursor.execute(sql, params)

        # 查询数据
        self.cursor.execute("""
            SELECT
                `客户编号`, `客户名称`,
                `运费`, `超时费等`, `运费合计`,
                `杂费收入`, `杂费`, `运杂费合计`,
                `代理费`, `公司杂费`, `公司运费`, `司机杂费`,
                `净运费`
            FROM `客户汇总TEMP`
            ORDER BY `客户编号`
        """)
        rows = self.cursor.fetchall()

        if not rows:
            messagebox.showwarning("警告", f"{date_start} 至 {date_end} 没有数据")
            self.cursor.execute("DROP TABLE IF EXISTS `客户汇总TEMP`")
            return "无数据"

        # 定义样式
        FONT_BASE = Font(name='宋体', size=12)
        THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

        HEADERS = ['客户编号', '客户名称', '运费', '超时费等', '运费合计',
                   '杂费收入', '杂费', '运杂费合计',
                   '代理费', '公司杂费', '公司运费', '司机杂费', '净运费']

        AMOUNT_COLS = ['运费', '超时费等', '运费合计', '杂费收入', '杂费', '运杂费合计',
                       '代理费', '公司杂费', '公司运费', '司机杂费', '净运费']

        COL_WIDTHS = [9.38, 46.4, 15.9, 13.6, 16.1, 13.9, 13.9, 16.1, 12.9, 11.6, 13.9, 11.4, 16.5]

        N_COLS = len(HEADERS)
        LAST_COL = get_column_letter(N_COLS)

        # 创建 Excel
        yyyymm = f"{year}{month:02d}"
        out_dir = os.path.join(self.summary_dir_var.get(), yyyymm)
        os.makedirs(out_dir, exist_ok=True)

        # 生成带序号的文件名，避免覆盖
        base_filename = f'客户汇总表{yyyymm}.xlsx'
        filepath = os.path.join(out_dir, base_filename)
        suffix_idx = ord('A')  # 从 'A' 开始

        # 如果文件已存在，添加后缀
        while os.path.exists(filepath):
            suffix = chr(suffix_idx)
            filepath = os.path.join(out_dir, f'客户汇总表{yyyymm}{suffix}.xlsx')
            suffix_idx += 1
            # 超过 'Z' 后重置为 'A'
            if suffix_idx > ord('Z'):
                suffix_idx = ord('A')
                suffix = ''  # 清空，下次循环会添加新的前缀（如 AA, AB）

        wb = Workbook()
        ws = wb.active

        # 标题行（合并）
        ws.merge_cells(f'A1:{LAST_COL}1')
        title_cell = ws['A1']
        title_cell.value = f'客户汇总表{year}年{month}月'
        title_cell.font = Font(name='宋体', size=12)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 14.25

        # 表头行
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = FONT_BASE
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 15

        # 数据行
        DATA_ROW_HEIGHT = 18
        for row_offset, record in enumerate(rows):
            row_num = 3 + row_offset
            ws.row_dimensions[row_num].height = DATA_ROW_HEIGHT

            c = ws.cell(row=row_num, column=1, value=record['客户编号'])
            c.font = FONT_BASE
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER

            c = ws.cell(row=row_num, column=2, value=record['客户名称'])
            c.font = FONT_BASE
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER

            for col_idx, col_name in enumerate(AMOUNT_COLS, start=3):
                c = ws.cell(row=row_num, column=col_idx)
                value = record.get(col_name, 0)
                if value is not None and value != '':
                    c.value = float(value)
                else:
                    c.value = 0
                c.font = FONT_BASE
                c.alignment = ALIGN_RIGHT
                c.border = THIN_BORDER
                c.number_format = '#,##0.00'

        # 列宽
        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(filepath)

        # 清理
        self.cursor.execute("DROP TABLE IF EXISTS `客户汇总TEMP`")

        return filepath

    def append_summary_result(self, text):
        """追加客户汇总表导出结果到文本框"""
        self.summary_result_text.configure(state='normal')
        self.summary_result_text.insert(tk.END, text + "\n")
        self.summary_result_text.see(tk.END)
        self.summary_result_text.configure(state='disabled')
        self.root.update()

    def append_export_result(self, text):
        """追加导出结果到文本框"""
        self.export_result_text.configure(state='normal')
        self.export_result_text.insert(tk.END, text + "\n")
        self.export_result_text.see(tk.END)
        self.export_result_text.configure(state='disabled')
        self.root.update()

    def show_biz_stats(self):
        """显示业务统计页面"""
        # ── 滚动容器 ──────────────────────────────────
        canvas = tk.Canvas(self.main_content, bg='#f0f2f5', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f2f5')
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=1180)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # ── 页面标题 ──────────────────────────────────
        title_frame = tk.Frame(scrollable_frame, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            title_frame,
            text="📈  业务统计",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg='#f0f2f5',
            fg='#262626'
        ).pack(side=tk.LEFT)

        # ── 导出卡片 ──────────────────────────────────
        biz_card = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        biz_card.pack(fill=tk.X, pady=(0, 20), ipady=10)

        card_header = tk.Frame(biz_card, bg='#ffffff', padx=20, pady=15)
        card_header.pack(fill=tk.X)

        tk.Label(
            card_header,
            text="📊 按月统计各客户业务量",
            font=('Microsoft YaHei UI', 13, 'bold'),
            bg='#ffffff',
            fg='#262626'
        ).pack(side=tk.LEFT)

        tk.Label(
            card_header,
            text="统计指定年份内每月每个客户的发票数量，导出为 Excel（每月一个 Sheet）",
            font=('Microsoft YaHei UI', 10),
            bg='#ffffff',
            fg='#8c8c8c'
        ).pack(side=tk.LEFT, padx=(15, 0))

        # 控制区
        biz_control_frame = tk.Frame(biz_card, bg='#ffffff', padx=20, pady=10)
        biz_control_frame.pack(fill=tk.X)

        # 年份 / 月份选择
        biz_date_frame = tk.Frame(biz_control_frame, bg='#ffffff')
        biz_date_frame.pack(side=tk.LEFT)

        tk.Label(biz_date_frame, text="统计年份:", font=('Microsoft YaHei UI', 11),
                 bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 6))

        from datetime import datetime
        current_date = datetime.now()
        year_list = [str(y) for y in range(2003, 2051)]
        self.biz_year_var = tk.StringVar(value=str(current_date.year))
        self.biz_year_combo = ttk.Combobox(biz_date_frame, textvariable=self.biz_year_var,
                                           values=year_list, width=6, state='readonly',
                                           font=('Microsoft YaHei UI', 11))
        # 显式选中当前年份
        try:
            self.biz_year_combo.current(year_list.index(str(current_date.year)))
        except ValueError:
            self.biz_year_combo.current(0)
        self.biz_year_combo.pack(side=tk.LEFT, padx=(0, 4))

        tk.Label(biz_date_frame, text="年", font=('Microsoft YaHei UI', 11),
                 bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 16))

        tk.Label(biz_date_frame, text="统计月份:", font=('Microsoft YaHei UI', 11),
                 bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 6))

        month_list = ['全年'] + [f"{m}月" for m in range(1, 13)]
        self.biz_month_var = tk.StringVar(value='全年')
        self.biz_month_combo = ttk.Combobox(biz_date_frame, textvariable=self.biz_month_var,
                                            values=month_list, width=6, state='readonly',
                                            font=('Microsoft YaHei UI', 11))
        self.biz_month_combo.current(0)   # 显式选中"全年"（索引0）
        self.biz_month_combo.pack(side=tk.LEFT, padx=(0, 20))

        # 导出目录
        biz_dir_frame = tk.Frame(biz_control_frame, bg='#ffffff')
        biz_dir_frame.pack(side=tk.LEFT, padx=(20, 0))

        tk.Label(biz_dir_frame, text="导出目录:", font=('Microsoft YaHei UI', 11),
                 bg='#ffffff', fg='#595959').pack(side=tk.LEFT, padx=(0, 10))

        self.biz_dir_var = tk.StringVar(value=self.config.get('biz_dir', r'X:\月结单'))
        biz_dir_entry = tk.Entry(biz_dir_frame, textvariable=self.biz_dir_var, width=40,
                                 font=('Microsoft YaHei UI', 11))
        biz_dir_entry.pack(side=tk.LEFT, padx=(0, 10))

        def browse_biz_dir():
            from tkinter import filedialog
            d = filedialog.askdirectory(title="选择导出目录")
            if d:
                self.biz_dir_var.set(d)
                self.config['biz_dir'] = d
                self.save_config()

        tk.Button(
            biz_dir_frame, text="浏览",
            font=('Microsoft YaHei UI', 10), bg='#1890ff', fg='#ffffff',
            activebackground='#40a9ff', activeforeground='#ffffff',
            bd=0, padx=15, pady=6, cursor='hand2',
            command=browse_biz_dir
        ).pack(side=tk.LEFT)

        # 导出按钮
        tk.Button(
            biz_control_frame, text="导出业务统计",
            font=('Microsoft YaHei UI', 11, 'bold'), bg='#722ed1', fg='#ffffff',
            activebackground='#9254de', activeforeground='#ffffff',
            bd=0, padx=30, pady=8, cursor='hand2',
            command=self.export_biz_stats_gui
        ).pack(side=tk.LEFT, padx=(20, 10))

        # 状态标签
        self.biz_status_var = tk.StringVar(value="准备就绪")
        tk.Label(
            biz_control_frame, textvariable=self.biz_status_var,
            font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#8c8c8c'
        ).pack(side=tk.LEFT, padx=(10, 0))

        # ── 结果文本框 ──────────────────────────────────
        biz_result_frame = tk.Frame(scrollable_frame, bg='#ffffff', bd=1, relief='solid')
        biz_result_frame.pack(fill=tk.X, pady=(0, 20), ipady=10)

        result_header = tk.Frame(biz_result_frame, bg='#ffffff', padx=15, pady=12)
        result_header.pack(fill=tk.X)

        tk.Label(
            result_header, text="📄 导出结果",
            font=('Microsoft YaHei UI', 12, 'bold'), bg='#ffffff', fg='#262626'
        ).pack(side=tk.LEFT)

        self.biz_result_text = tk.Text(
            biz_result_frame, height=10,
            font=('Consolas', 10), bg='#ffffff', fg='#262626',
            relief='solid', bd=1, padx=10, pady=10
        )
        self.biz_result_text.pack(fill=tk.X)
        self.biz_result_text.insert(tk.END, "点击「导出业务统计」按钮开始导出...")
        self.biz_result_text.configure(state='disabled')

    def export_biz_stats_gui(self):
        """从 GUI 调用业务统计导出"""
        try:
            year = int(self.biz_year_combo.get())
            month_str = self.biz_month_combo.get()   # '全年' 或 '1月'~'12月'
            if month_str == '全年':
                month = None
                range_label = f"{year}年（全年）"
            else:
                month = int(month_str.rstrip('月'))
                range_label = f"{year}年{month}月"

            self.biz_status_var.set("正在导出...")
            self.root.update()

            # 清空结果文本框
            self.biz_result_text.configure(state='normal')
            self.biz_result_text.delete(1.0, tk.END)
            self.biz_result_text.configure(state='disabled')

            self.append_biz_stats_result(f"{'='*60}")
            self.append_biz_stats_result(f"业务统计导出 - {range_label}")
            self.append_biz_stats_result(f"{'='*60}\n")

            result = self.export_biz_stats_core(year, month)

            self.append_biz_stats_result(f"\n✓ 导出完成!")
            self.append_biz_stats_result(f"文件路径: {result}")
            self.append_biz_stats_result(f"\n{'='*60}")

            self.biz_status_var.set("准备就绪")

        except ValueError as e:
            self.append_biz_stats_result(f"\n[错误] 输入格式错误: {e}")
            self.biz_status_var.set("准备就绪")
        except Exception as e:
            self.append_biz_stats_result(f"\n[错误] 导出失败: {e}")
            import traceback
            self.append_biz_stats_result(f"\n详细错误:\n{traceback.format_exc()}")
            self.biz_status_var.set("准备就绪")

    def export_biz_stats_core(self, year, month=None):
        """
        业务统计核心逻辑：
        month=None  → 全年汇总（全年汇总Sheet + 1~12月各Sheet）
        month=1~12  → 只导出该月的单张 Sheet
        """
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # ── 查询数据范围 ─────────────────────────────
        if month is None:
            date_start = f"{year}-01-01"
            date_end   = f"{year}-12-31"
        else:
            import calendar
            last_day   = calendar.monthrange(year, month)[1]
            date_start = f"{year}-{month:02d}-01"
            date_end   = f"{year}-{month:02d}-{last_day:02d}"

        sql = """
            SELECT
                im.CUSTCODE                                         AS custcode,
                COALESCE(MIN(cm.NAME), im.CUSTCODE)                 AS custname,
                MONTH(im.INVDATE)                                   AS inv_month,
                COUNT(DISTINCT im.INVOICECODE)                      AS inv_count
            FROM INVOICE_MASTER im
            LEFT JOIN CUST_MASTER cm ON cm.CUSTCODE = im.CUSTCODE
            WHERE im.INVDATE BETWEEN %s AND %s
            GROUP BY im.CUSTCODE, MONTH(im.INVDATE)
            ORDER BY im.CUSTCODE, MONTH(im.INVDATE)
        """
        self.cursor.execute(sql, (date_start, date_end))
        rows = self.cursor.fetchall()

        if not rows:
            from tkinter import messagebox
            label = f"{year}年{month}月" if month else f"{year}年"
            messagebox.showwarning("警告", f"{label} 没有数据")
            return "无数据"

        # ── 整理数据结构 ─────────────────────────────
        # data[custcode] = {'name': ..., months: {1:0, 2:0, ... 12:0}}
        data = {}
        cust_order = []  # 保留客户出现顺序
        for r in rows:
            code      = r['custcode']
            name      = r['custname'] or code
            inv_month = int(r['inv_month'])
            count     = int(r['inv_count'])
            if code not in data:
                data[code] = {'name': name, 'months': {m: 0 for m in range(1, 13)}}
                cust_order.append(code)
            data[code]['months'][inv_month] = count

        self.append_biz_stats_result(f"  共找到 {len(cust_order)} 个客户，正在生成 Excel...")

        # ── 样式 ─────────────────────────────────────
        FONT_HEADER = Font(name='宋体', size=11, bold=True)
        FONT_BASE   = Font(name='宋体', size=11)
        FONT_TITLE  = Font(name='宋体', size=13, bold=True)
        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
        ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center')
        ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')
        FILL_HEADER  = PatternFill(fill_type='solid', fgColor='DDEEFF')   # 淡蓝
        FILL_TOTAL   = PatternFill(fill_type='solid', fgColor='FFF2CC')   # 淡黄
        FILL_TITLE   = PatternFill(fill_type='solid', fgColor='EBF5FB')   # 极淡蓝

        # ── 创建 Excel ────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)  # 删掉默认空 Sheet

        MONTH_NAMES = ['一月','二月','三月','四月','五月','六月',
                       '七月','八月','九月','十月','十一月','十二月']

        def _write_month_sheet(ws, yr, mo, data, cust_order):
            """将指定月份数据写入一个 Sheet"""
            month_data = [(code, data[code]) for code in cust_order
                          if data[code]['months'][mo] > 0]

            # 标题行
            ws.merge_cells('A1:D1')
            c = ws['A1']
            c.value     = f"{yr}年{mo}月  客户业务量统计"
            c.font      = FONT_TITLE
            c.alignment = ALIGN_CENTER
            c.fill      = FILL_TITLE
            ws.row_dimensions[1].height = 22

            # 表头
            for col_idx, h in enumerate(['客户编号', '客户名称', '业务数量（票）', '备注'], start=1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = FONT_HEADER; cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER; cell.fill = FILL_HEADER
            ws.row_dimensions[2].height = 18

            if month_data:
                month_total = 0
                for row_offset, (code, info) in enumerate(month_data):
                    row_num = 3 + row_offset
                    count   = info['months'][mo]
                    month_total += count

                    c = ws.cell(row=row_num, column=1, value=code)
                    c.font = FONT_BASE; c.alignment = ALIGN_CENTER; c.border = THIN_BORDER

                    c = ws.cell(row=row_num, column=2, value=info['name'])
                    c.font = FONT_BASE; c.alignment = ALIGN_LEFT; c.border = THIN_BORDER

                    c = ws.cell(row=row_num, column=3, value=count)
                    c.font = FONT_BASE; c.alignment = ALIGN_CENTER; c.border = THIN_BORDER

                    c = ws.cell(row=row_num, column=4, value='')
                    c.border = THIN_BORDER
                    ws.row_dimensions[row_num].height = 16

                # 合计行
                total_row_m = 3 + len(month_data)
                ws.merge_cells(f'A{total_row_m}:B{total_row_m}')
                c = ws.cell(row=total_row_m, column=1, value='合  计')
                c.font = FONT_HEADER; c.alignment = ALIGN_CENTER
                c.border = THIN_BORDER; c.fill = FILL_TOTAL
                ws.cell(row=total_row_m, column=2).border = THIN_BORDER
                ws.cell(row=total_row_m, column=2).fill   = FILL_TOTAL

                c = ws.cell(row=total_row_m, column=3, value=month_total)
                c.font = FONT_HEADER; c.alignment = ALIGN_CENTER
                c.border = THIN_BORDER; c.fill = FILL_TOTAL

                c = ws.cell(row=total_row_m, column=4, value='')
                c.border = THIN_BORDER; c.fill = FILL_TOTAL
                ws.row_dimensions[total_row_m].height = 18

                self.append_biz_stats_result(f"  {mo:2d}月: {len(month_data)} 个客户，共 {month_total} 票")
            else:
                ws.merge_cells('A3:D3')
                c = ws['A3']
                c.value = '（本月无数据）'
                c.font = Font(name='宋体', size=11, color='888888')
                c.alignment = ALIGN_CENTER
                self.append_biz_stats_result(f"  {mo:2d}月: 无数据")

            # 列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 32
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 20

        if month is None:
            # ── 全年模式：全年汇总 Sheet + 1~12 月 Sheet ──
            # 汇总 Sheet
            ws_all = wb.create_sheet(title=f"{year}年全年汇总")

            n_data_cols = 12
            total_cols  = 2 + n_data_cols + 1
            last_col    = get_column_letter(total_cols)

            ws_all.merge_cells(f'A1:{last_col}1')
            c = ws_all['A1']
            c.value     = f"{year}年  客户业务量统计表（全年汇总）"
            c.font      = FONT_TITLE
            c.alignment = ALIGN_CENTER
            c.fill      = FILL_TITLE
            ws_all.row_dimensions[1].height = 22

            headers = ['客户编号', '客户名称'] + MONTH_NAMES + ['全年合计']
            for col_idx, h in enumerate(headers, start=1):
                cell = ws_all.cell(row=2, column=col_idx, value=h)
                cell.font      = FONT_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border    = THIN_BORDER
                cell.fill      = FILL_HEADER
            ws_all.row_dimensions[2].height = 18

            grand_month_totals = {m: 0 for m in range(1, 13)}
            for row_offset, code in enumerate(cust_order):
                row_num  = 3 + row_offset
                info     = data[code]
                row_total = sum(info['months'].values())

                ws_all.cell(row=row_num, column=1, value=code).border    = THIN_BORDER
                ws_all.cell(row=row_num, column=1).font      = FONT_BASE
                ws_all.cell(row=row_num, column=1).alignment = ALIGN_CENTER

                ws_all.cell(row=row_num, column=2, value=info['name']).border    = THIN_BORDER
                ws_all.cell(row=row_num, column=2).font      = FONT_BASE
                ws_all.cell(row=row_num, column=2).alignment = ALIGN_LEFT

                for m in range(1, 13):
                    val = info['months'][m]
                    grand_month_totals[m] += val
                    c = ws_all.cell(row=row_num, column=2 + m, value=val if val else '')
                    c.font      = FONT_BASE
                    c.alignment = ALIGN_CENTER
                    c.border    = THIN_BORDER

                c = ws_all.cell(row=row_num, column=total_cols, value=row_total)
                c.font      = Font(name='宋体', size=11, bold=True)
                c.alignment = ALIGN_CENTER
                c.border    = THIN_BORDER
                ws_all.row_dimensions[row_num].height = 16

            total_row = 3 + len(cust_order)
            ws_all.row_dimensions[total_row].height = 18

            c = ws_all.cell(row=total_row, column=1, value='合计')
            c.font = FONT_HEADER; c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER; c.fill = FILL_TOTAL

            ws_all.merge_cells(f'A{total_row}:B{total_row}')
            ws_all.cell(row=total_row, column=2).border = THIN_BORDER
            ws_all.cell(row=total_row, column=2).fill   = FILL_TOTAL

            grand_total = 0
            for m in range(1, 13):
                val = grand_month_totals[m]
                grand_total += val
                c = ws_all.cell(row=total_row, column=2 + m, value=val)
                c.font = FONT_HEADER; c.alignment = ALIGN_CENTER
                c.border = THIN_BORDER; c.fill = FILL_TOTAL

            c = ws_all.cell(row=total_row, column=total_cols, value=grand_total)
            c.font = FONT_HEADER; c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER; c.fill = FILL_TOTAL

            ws_all.column_dimensions['A'].width = 12
            ws_all.column_dimensions['B'].width = 32
            for m in range(1, 13):
                ws_all.column_dimensions[get_column_letter(2 + m)].width = 9
            ws_all.column_dimensions[get_column_letter(total_cols)].width = 10

            # 每月独立 Sheet
            for m in range(1, 13):
                ws = wb.create_sheet(title=f"{m}月")
                _write_month_sheet(ws, year, m, data, cust_order)

            base_filename = f'业务统计{year}年.xlsx'

        else:
            # ── 单月模式：只导出该月 Sheet ──
            ws = wb.create_sheet(title=f"{month}月")
            _write_month_sheet(ws, year, month, data, cust_order)
            base_filename = f'业务统计{year}年{month}月.xlsx'



        # ── 保存文件 ──────────────────────────────────
        out_dir  = self.biz_dir_var.get()
        os.makedirs(out_dir, exist_ok=True)
        # base_filename 已在上方 if/else 分支中设置
        stem, ext = os.path.splitext(base_filename)  # e.g. ('业务统计2026年', '.xlsx')
        filepath = os.path.join(out_dir, base_filename)
        suffix_idx = ord('A')
        while os.path.exists(filepath):
            suffix = chr(suffix_idx)
            filepath = os.path.join(out_dir, f'{stem}{suffix}{ext}')
            suffix_idx += 1
            if suffix_idx > ord('Z'):
                suffix_idx = ord('A')

        wb.save(filepath)
        return filepath


    def append_dbiz_log(self, text):
        """向司机业务量统计日志框追加文字"""
        self.dbiz_log_text.configure(state='normal')
        self.dbiz_log_text.insert(tk.END, text + '\n')
        self.dbiz_log_text.see(tk.END)
        self.dbiz_log_text.configure(state='disabled')
        self.root.update_idletasks()

    def export_driver_biz_stats_gui(self):
        """GUI 回调：司机业务量统计导出"""
        try:
            year = int(self.dbiz_year_combo.get())
            month_str = self.dbiz_month_combo.get()  # '全年' 或 '1月'~'12月'
            if month_str == '全年':
                month = None
                range_label = f"{year}年（全年）"
            else:
                month = int(month_str.rstrip('月'))
                range_label = f"{year}年{month}月"

            self.dbiz_status_var.set("正在导出...")
            self.root.update()

            self.dbiz_log_text.configure(state='normal')
            self.dbiz_log_text.delete(1.0, tk.END)
            self.dbiz_log_text.configure(state='disabled')

            self.append_dbiz_log(f"{'='*60}")
            self.append_dbiz_log(f"司机业务量统计 - {range_label}")
            self.append_dbiz_log(f"{'='*60}\n")

            filepath = self.export_driver_biz_stats_core(year, month)

            self.append_dbiz_log(f"\n✓ 导出完成!")
            self.append_dbiz_log(f"文件: {filepath}")
            self.append_dbiz_log(f"\n{'='*60}")
            self.dbiz_status_var.set("准备就绪")

        except ValueError as e:
            self.append_dbiz_log(f"\n[错误] 输入格式错误: {e}")
            self.dbiz_status_var.set("准备就绪")
        except Exception as e:
            self.append_dbiz_log(f"\n[错误] 导出失败: {e}")
            import traceback
            self.append_dbiz_log(f"\n详细错误:\n{traceback.format_exc()}")
            self.dbiz_status_var.set("准备就绪")

    def export_driver_biz_stats_core(self, year, month=None):
        """
        司机业务量统计核心逻辑：
        month=None  → 全年（汇总Sheet + 1~12月各Sheet）
        month=1~12  → 只导该月

        统计内容：每个司机每月的业务量（票数）、运费合计、杂费合计
        数据来源：
          - 运费：CON_DETAIL.DRIVER = 司機編號, DRIVERCOMM
          - 杂费：INVOICE_DETAIL.DRIVERCODE = 司機編號, WHOPAY IN (0,1)
        """
        import os
        import calendar
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # ── 查询日期范围 ─────────────────────────────
        if month is None:
            date_start = f"{year}-01-01"
            date_end   = f"{year}-12-31"
        else:
            last_day   = calendar.monthrange(year, month)[1]
            date_start = f"{year}-{month:02d}-01"
            date_end   = f"{year}-{month:02d}-{last_day:02d}"

        # ── SQL1: 从 CON_DETAIL 统计运费（每行=一个司机的一笔运费）──
        sql_fare = """
            SELECT
                cd.DRIVER                        AS driver_code,
                COALESCE(dc.NAME, cd.DRIVER)     AS driver_name,
                MONTH(im.INVDATE)                AS inv_month,
                COUNT(DISTINCT im.INVOICECODE)   AS inv_count,
                SUM(cd.DRIVERCOMM)               AS total_fare
            FROM CON_DETAIL cd
            JOIN INVOICE_MASTER im ON im.INVOICECODE = cd.INVOICECODE
            LEFT JOIN DRIVER_CP dc ON dc.CUSTCODE = cd.DRIVER
            WHERE im.INVDATE BETWEEN %s AND %s
              AND cd.DRIVER IS NOT NULL
              AND cd.DRIVER <> ''
            GROUP BY cd.DRIVER, MONTH(im.INVDATE)
        """

        # ── SQL2: 从 INVOICE_DETAIL 统计杂费（以逗号结尾的项）──
        sql_misc = """
            SELECT
                id.DRIVERCODE                    AS driver_code,
                COALESCE(dc.NAME, id.DRIVERCODE) AS driver_name,
                MONTH(im.INVDATE)                AS inv_month,
                SUM(CAST(id.PRICE AS DECIMAL(12,2))) AS total_misc
            FROM INVOICE_DETAIL id
            JOIN INVOICE_MASTER im ON im.INVOICECODE = id.INVOICECODE
            LEFT JOIN DRIVER_CP dc ON dc.CUSTCODE = id.DRIVERCODE
            WHERE im.INVDATE BETWEEN %s AND %s
              AND id.DRIVERCODE IS NOT NULL
              AND id.DRIVERCODE <> ''
              AND id.WHOPAY IN (0, 1)
              AND id.DESCR LIKE '%%,'
            GROUP BY id.DRIVERCODE, MONTH(im.INVDATE)
        """

        self.cursor.execute(sql_fare, (date_start, date_end))
        fare_rows = self.cursor.fetchall()
        self.cursor.execute(sql_misc, (date_start, date_end))
        misc_rows = self.cursor.fetchall()

        if not fare_rows and not misc_rows:
            from tkinter import messagebox
            label = f"{year}年{month}月" if month else f"{year}年"
            messagebox.showwarning("警告", f"{label} 没有司机数据")
            return "无数据"

        # ── 合并数据 ────────────────────────────────
        # data[driver_code] = {'name': ..., 'months': {1: {'count':0, 'fare':0, 'misc':0}, ...}}
        data = {}
        driver_order = []

        for r in fare_rows:
            code = r['driver_code']
            mo   = int(r['inv_month'])
            if code not in data:
                data[code] = {'name': r['driver_name'] or code, 'months': {m: {'count': 0, 'fare': 0.0, 'misc': 0.0} for m in range(1, 13)}}
                driver_order.append(code)
            data[code]['months'][mo]['count'] += int(r['inv_count'])
            data[code]['months'][mo]['fare']  += float(r['total_fare'] or 0)

        for r in misc_rows:
            code = r['driver_code']
            mo   = int(r['inv_month'])
            if code not in data:
                data[code] = {'name': r['driver_name'] or code, 'months': {m: {'count': 0, 'fare': 0.0, 'misc': 0.0} for m in range(1, 13)}}
                driver_order.append(code)
            data[code]['months'][mo]['misc'] += float(r['total_misc'] or 0)

        self.append_dbiz_log(f"  共找到 {len(driver_order)} 个司机")

        # ── 样式 ─────────────────────────────────────
        FONT_HEADER = Font(name='宋体', size=11, bold=True)
        FONT_BASE   = Font(name='宋体', size=11)
        FONT_TITLE  = Font(name='宋体', size=13, bold=True)
        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
        ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center')
        ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')
        FILL_HEADER  = PatternFill(fill_type='solid', fgColor='DDEEFF')
        FILL_TOTAL   = PatternFill(fill_type='solid', fgColor='FFF2CC')
        FILL_TITLE   = PatternFill(fill_type='solid', fgColor='EBF5FB')

        MONTH_NAMES = ['一月','二月','三月','四月','五月','六月',
                       '七月','八月','九月','十月','十一月','十二月']

        # ── 创建 Excel ────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        def _write_month_sheet(ws, yr, mo, data, driver_order):
            """将指定月份的司机业务量写入一个 Sheet"""
            month_data = [(code, data[code]) for code in driver_order
                          if data[code]['months'][mo]['count'] > 0 or data[code]['months'][mo]['misc'] != 0]

            ws.merge_cells('A1:F1')
            c = ws['A1']
            c.value     = f"{yr}年{mo}月  司机业务量统计"
            c.font      = FONT_TITLE
            c.alignment = ALIGN_CENTER
            c.fill      = FILL_TITLE
            ws.row_dimensions[1].height = 22

            headers = ['司机编号', '司机姓名', '业务数量（票）', '运费合计', '杂费合计', '备注']
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = FONT_HEADER; cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER; cell.fill = FILL_HEADER
            ws.row_dimensions[2].height = 18

            if month_data:
                m_count = 0; m_fare = 0.0; m_misc = 0.0
                for row_offset, (code, info) in enumerate(month_data):
                    row_num = 3 + row_offset
                    d = info['months'][mo]
                    m_count += d['count']; m_fare += d['fare']; m_misc += d['misc']

                    ws.cell(row=row_num, column=1, value=code).font = FONT_BASE
                    ws.cell(row=row_num, column=1).alignment = ALIGN_CENTER
                    ws.cell(row=row_num, column=1).border = THIN_BORDER

                    ws.cell(row=row_num, column=2, value=info['name']).font = FONT_BASE
                    ws.cell(row=row_num, column=2).alignment = ALIGN_LEFT
                    ws.cell(row=row_num, column=2).border = THIN_BORDER

                    ws.cell(row=row_num, column=3, value=d['count']).font = FONT_BASE
                    ws.cell(row=row_num, column=3).alignment = ALIGN_CENTER
                    ws.cell(row=row_num, column=3).border = THIN_BORDER

                    c = ws.cell(row=row_num, column=4, value=round(d['fare'], 2) if d['fare'] else '')
                    c.font = FONT_BASE; c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER
                    c.number_format = '#,##0.00'

                    c = ws.cell(row=row_num, column=5, value=round(d['misc'], 2) if d['misc'] else '')
                    c.font = FONT_BASE; c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER
                    c.number_format = '#,##0.00'

                    ws.cell(row=row_num, column=6, value='').border = THIN_BORDER
                    ws.row_dimensions[row_num].height = 16

                # 合计行
                tr = 3 + len(month_data)
                ws.merge_cells(f'A{tr}:B{tr}')
                c = ws.cell(row=tr, column=1, value='合  计')
                c.font = FONT_HEADER; c.alignment = ALIGN_CENTER
                c.border = THIN_BORDER; c.fill = FILL_TOTAL
                ws.cell(row=tr, column=2).border = THIN_BORDER
                ws.cell(row=tr, column=2).fill   = FILL_TOTAL

                ws.cell(row=tr, column=3, value=m_count).font = FONT_HEADER
                ws.cell(row=tr, column=3).alignment = ALIGN_CENTER
                ws.cell(row=tr, column=3).border = THIN_BORDER
                ws.cell(row=tr, column=3).fill   = FILL_TOTAL

                c = ws.cell(row=tr, column=4, value=round(m_fare, 2))
                c.font = FONT_HEADER; c.alignment = ALIGN_RIGHT
                c.border = THIN_BORDER; c.fill = FILL_TOTAL
                c.number_format = '#,##0.00'

                c = ws.cell(row=tr, column=5, value=round(m_misc, 2))
                c.font = FONT_HEADER; c.alignment = ALIGN_RIGHT
                c.border = THIN_BORDER; c.fill = FILL_TOTAL
                c.number_format = '#,##0.00'

                ws.cell(row=tr, column=6, value='').border = THIN_BORDER
                ws.cell(row=tr, column=6).fill   = FILL_TOTAL
                ws.row_dimensions[tr].height = 18

                self.append_dbiz_log(f"  {mo:2d}月: {len(month_data)} 个司机，{m_count} 票，运费 {m_fare:,.2f}，杂费 {m_misc:,.2f}")
            else:
                ws.merge_cells('A3:F3')
                c = ws['A3']
                c.value = '（本月无数据）'
                c.font = Font(name='宋体', size=11, color='888888')
                c.alignment = ALIGN_CENTER
                self.append_dbiz_log(f"  {mo:2d}月: 无数据")

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 16
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 20

        if month is None:
            # ── 全年模式 ──
            # 汇总 Sheet：每个司机一行，列=1~12月(票数) + 运费合计 + 杂费合计
            ws_all = wb.create_sheet(title=f"{year}年全年汇总")

            # 表头：司机编号 | 司机姓名 | 1月票数 | 1月运费 | ... | 12月票数 | 12月运费 | 全年票数 | 全年运费 | 全年杂费
            headers = ['司机编号', '司机姓名']
            for m in range(1, 13):
                headers.append(f'{m}月(票)')
                headers.append(f'{m}月(运费)')
            headers.extend(['全年票数', '全年运费', '全年杂费'])
            total_cols = len(headers)
            last_col = get_column_letter(total_cols)

            ws_all.merge_cells(f'A1:{last_col}1')
            c = ws_all['A1']
            c.value     = f"{year}年  司机业务量统计表（全年汇总）"
            c.font      = FONT_TITLE
            c.alignment = ALIGN_CENTER
            c.fill      = FILL_TITLE
            ws_all.row_dimensions[1].height = 22

            for col_idx, h in enumerate(headers, start=1):
                cell = ws_all.cell(row=2, column=col_idx, value=h)
                cell.font = FONT_HEADER; cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER; cell.fill = FILL_HEADER
            ws_all.row_dimensions[2].height = 18

            grand_month_counts = {m: 0 for m in range(1, 13)}
            grand_month_fares  = {m: 0.0 for m in range(1, 13)}
            grand_count = 0; grand_fare = 0.0; grand_misc = 0.0

            for row_offset, code in enumerate(driver_order):
                row_num = 3 + row_offset
                info = data[code]
                y_count = 0; y_fare = 0.0; y_misc = 0.0

                ws_all.cell(row=row_num, column=1, value=code).font = FONT_BASE
                ws_all.cell(row=row_num, column=1).alignment = ALIGN_CENTER
                ws_all.cell(row=row_num, column=1).border = THIN_BORDER

                ws_all.cell(row=row_num, column=2, value=info['name']).font = FONT_BASE
                ws_all.cell(row=row_num, column=2).alignment = ALIGN_LEFT
                ws_all.cell(row=row_num, column=2).border = THIN_BORDER

                for m in range(1, 13):
                    d = info['months'][m]
                    y_count += d['count']; y_fare += d['fare']; y_misc += d['misc']
                    grand_month_counts[m] += d['count']
                    grand_month_fares[m]  += d['fare']

                    col_count = 3 + (m - 1) * 2      # 票数列
                    col_fare  = 3 + (m - 1) * 2 + 1  # 运费列

                    c = ws_all.cell(row=row_num, column=col_count, value=d['count'] if d['count'] else '')
                    c.font = FONT_BASE; c.alignment = ALIGN_CENTER; c.border = THIN_BORDER

                    c = ws_all.cell(row=row_num, column=col_fare, value=round(d['fare'], 2) if d['fare'] else '')
                    c.font = FONT_BASE; c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER
                    c.number_format = '#,##0.00'

                grand_count += y_count; grand_fare += y_fare; grand_misc += y_misc

                # 全年汇总列
                col_y_count = 3 + 24
                col_y_fare  = 3 + 25
                col_y_misc  = 3 + 26

                c = ws_all.cell(row=row_num, column=col_y_count, value=y_count)
                c.font = Font(name='宋体', size=11, bold=True)
                c.alignment = ALIGN_CENTER; c.border = THIN_BORDER

                c = ws_all.cell(row=row_num, column=col_y_fare, value=round(y_fare, 2))
                c.font = Font(name='宋体', size=11, bold=True)
                c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER
                c.number_format = '#,##0.00'

                c = ws_all.cell(row=row_num, column=col_y_misc, value=round(y_misc, 2))
                c.font = Font(name='宋体', size=11, bold=True)
                c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER
                c.number_format = '#,##0.00'

                ws_all.row_dimensions[row_num].height = 16

            # 合计行
            tr = 3 + len(driver_order)
            ws_all.merge_cells(f'A{tr}:B{tr}')
            c = ws_all.cell(row=tr, column=1, value='合  计')
            c.font = FONT_HEADER; c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER; c.fill = FILL_TOTAL
            ws_all.cell(row=tr, column=2).border = THIN_BORDER
            ws_all.cell(row=tr, column=2).fill   = FILL_TOTAL

            for m in range(1, 13):
                col_count = 3 + (m - 1) * 2
                col_fare  = 3 + (m - 1) * 2 + 1
                c = ws_all.cell(row=tr, column=col_count, value=grand_month_counts[m])
                c.font = FONT_HEADER; c.alignment = ALIGN_CENTER; c.border = THIN_BORDER; c.fill = FILL_TOTAL
                c = ws_all.cell(row=tr, column=col_fare, value=round(grand_month_fares[m], 2))
                c.font = FONT_HEADER; c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER; c.fill = FILL_TOTAL
                c.number_format = '#,##0.00'

            col_y_count = 3 + 24
            col_y_fare  = 3 + 25
            col_y_misc  = 3 + 26
            c = ws_all.cell(row=tr, column=col_y_count, value=grand_count)
            c.font = FONT_HEADER; c.alignment = ALIGN_CENTER; c.border = THIN_BORDER; c.fill = FILL_TOTAL
            c = ws_all.cell(row=tr, column=col_y_fare, value=round(grand_fare, 2))
            c.font = FONT_HEADER; c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER; c.fill = FILL_TOTAL
            c.number_format = '#,##0.00'
            c = ws_all.cell(row=tr, column=col_y_misc, value=round(grand_misc, 2))
            c.font = FONT_HEADER; c.alignment = ALIGN_RIGHT; c.border = THIN_BORDER; c.fill = FILL_TOTAL
            c.number_format = '#,##0.00'
            ws_all.row_dimensions[tr].height = 18

            # 列宽
            ws_all.column_dimensions['A'].width = 12
            ws_all.column_dimensions['B'].width = 16
            for m in range(1, 13):
                ws_all.column_dimensions[get_column_letter(3 + (m - 1) * 2)].width = 9
                ws_all.column_dimensions[get_column_letter(3 + (m - 1) * 2 + 1)].width = 12
            ws_all.column_dimensions[get_column_letter(col_y_count)].width = 10
            ws_all.column_dimensions[get_column_letter(col_y_fare)].width = 12
            ws_all.column_dimensions[get_column_letter(col_y_misc)].width = 12

            # 每月独立 Sheet
            for m in range(1, 13):
                ws = wb.create_sheet(title=f"{m}月")
                _write_month_sheet(ws, year, m, data, driver_order)

            base_filename = f'司机业务量统计{year}年.xlsx'
        else:
            # ── 单月模式 ──
            ws = wb.create_sheet(title=f"{month}月")
            _write_month_sheet(ws, year, month, data, driver_order)
            base_filename = f'司机业务量统计{year}年{month}月.xlsx'

        # ── 保存文件 ──────────────────────────────────
        out_dir = self.dbiz_dir_var.get()
        os.makedirs(out_dir, exist_ok=True)
        stem, ext = os.path.splitext(base_filename)
        filepath = os.path.join(out_dir, base_filename)
        suffix_idx = ord('A')
        while os.path.exists(filepath):
            suffix = chr(suffix_idx)
            filepath = os.path.join(out_dir, f'{stem}{suffix}{ext}')
            suffix_idx += 1
            if suffix_idx > ord('Z'):
                suffix_idx = ord('A')

        wb.save(filepath)
        return filepath

    def append_biz_stats_result(self, text):
        """追加业务统计导出结果到文本框"""
        self.biz_result_text.configure(state='normal')
        self.biz_result_text.insert(tk.END, text + "\n")
        self.biz_result_text.see(tk.END)
        self.biz_result_text.configure(state='disabled')
        self.root.update()

    # ── 删除数据 ──────────────────────────────────────

    def show_delete_data(self):
        """显示删除数据页面 —— 搜索并删除 CON_DETAIL 记录"""
        page = tk.Frame(self.main_content, bg='#f0f2f5')
        page.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # ── 页面标题 ──────────────────────────────────
        title_frame = tk.Frame(page, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            title_frame,
            text="删除数据",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg='#f0f2f5',
            fg='#262626'
        ).pack(side=tk.LEFT)

        # ── 搜索卡片 ──────────────────────────────────
        search_card = tk.Frame(page, bg='#ffffff', bd=1, relief='solid')
        search_card.pack(fill=tk.X, pady=(0, 15), ipady=10)

        card_header = tk.Frame(search_card, bg='#ffffff', padx=20, pady=15)
        card_header.pack(fill=tk.X)

        tk.Label(
            card_header,
            text="🔎 搜索 CON_DETAIL 记录",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff',
            fg='#262626'
        ).pack(side=tk.LEFT)

        search_control = tk.Frame(search_card, bg='#ffffff', padx=20)
        search_control.pack(fill=tk.X)

        tk.Label(search_control, text="發票號(INVOICECODE):",
                 font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#595959'
                 ).pack(side=tk.LEFT, padx=(0, 8))

        self._delete_search_var = tk.StringVar()
        self._delete_search_entry = ttk.Entry(
            search_control, textvariable=self._delete_search_var,
            width=25, font=('Microsoft YaHei UI', 11)
        )
        self._delete_search_entry.pack(side=tk.LEFT, padx=(0, 10))
        self._delete_search_entry.bind("<Return>", lambda e: self._delete_do_search())

        tk.Button(search_control, text="搜索", command=self._delete_do_search,
                  font=('Microsoft YaHei UI', 10), bg='#1890ff', fg='#ffffff',
                  bd=0, padx=12, pady=4, cursor='hand2'
                  ).pack(side=tk.LEFT, padx=3)
        tk.Button(search_control, text="清空", command=self._delete_clear,
                  font=('Microsoft YaHei UI', 10), bg='#f5f5f5', fg='#595959',
                  bd=0, padx=12, pady=4, cursor='hand2'
                  ).pack(side=tk.LEFT, padx=3)
        tk.Button(search_control, text="加载全部", command=self._delete_load_all,
                  font=('Microsoft YaHei UI', 10), bg='#f5f5f5', fg='#595959',
                  bd=0, padx=12, pady=4, cursor='hand2'
                  ).pack(side=tk.LEFT, padx=3)

        hint_text = "字段: INVOICECODE(發票號) | NN | CONCODE(櫃號) | DRIVER(司機編號) | SIZE(櫃尺碼) | DRIVERCODE(香港車牌) | TAKENO(提貨號碼) | DRIVERCOMM(司機運費)"
        tk.Label(search_card, text=hint_text, font=('Microsoft YaHei UI', 9),
                 bg='#ffffff', fg='#8c8c8c').pack(anchor="w", padx=20, pady=(0, 10))

        # ── 结果表格卡片 ──────────────────────────────────
        table_card = tk.Frame(page, bg='#ffffff', bd=1, relief='solid')
        table_card.pack(fill=tk.BOTH, expand=True, ipady=10)

        table_header = tk.Frame(table_card, bg='#ffffff', padx=20, pady=15)
        table_header.pack(fill=tk.X)

        tk.Label(
            table_header,
            text="📋 查询结果",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff',
            fg='#262626'
        ).pack(side=tk.LEFT)

        self._delete_count_label = tk.Label(
            table_header,
            text="共 0 条记录",
            font=('Microsoft YaHei UI', 11),
            bg='#ffffff',
            fg='#8c8c8c'
        )
        self._delete_count_label.pack(side=tk.RIGHT)

        # Treeview
        tree_container = tk.Frame(table_card, bg='#ffffff')
        tree_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))

        columns = ("INVOICECODE", "NN", "CONCODE", "DRIVER", "SIZE", "DRIVERCODE", "TAKENO", "DRIVERCOMM")
        col_labels = {
            "INVOICECODE": "發票號",
            "NN": "NN",
            "CONCODE": "櫃號",
            "DRIVER": "司機編號",
            "SIZE": "櫃尺碼",
            "DRIVERCODE": "香港車牌",
            "TAKENO": "提貨號碼",
            "DRIVERCOMM": "司機運費",
        }
        col_widths = {
            "INVOICECODE": 130, "NN": 80, "CONCODE": 110, "DRIVER": 90, "SIZE": 80,
            "DRIVERCODE": 100, "TAKENO": 110, "DRIVERCOMM": 100,
        }

        self._delete_tree = ttk.Treeview(
            tree_container, columns=columns, show="headings",
            selectmode="browse", height=15
        )
        for col in columns:
            label = col_labels.get(col, col)
            self._delete_tree.heading(col, text=label + "\n(" + col + ")")
            self._delete_tree.column(col, width=col_widths.get(col, 100),
                                     anchor="center", minwidth=60)

        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self._delete_tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self._delete_tree.xview)
        self._delete_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._delete_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        self._delete_tree.bind("<Double-1>", self._delete_on_double_click)

        # 操作按钮
        action_frame = tk.Frame(table_card, bg='#ffffff')
        action_frame.pack(fill="x", padx=20, pady=(0, 15))

        tk.Button(
            action_frame, text="删除选中记录",
            font=('Microsoft YaHei UI', 10), bg='#ff4d4f', fg='#ffffff',
            bd=0, padx=16, pady=5, cursor='hand2',
            command=self._delete_record_from_tree
        ).pack(side="left", padx=3)
        tk.Button(
            action_frame, text="刷新",
            font=('Microsoft YaHei UI', 10), bg='#f5f5f5', fg='#595959',
            bd=0, padx=16, pady=5, cursor='hand2',
            command=self._delete_do_search
        ).pack(side="left", padx=3)

    def _delete_do_search(self):
        keyword = self._delete_search_var.get().strip()
        if not keyword:
            self._delete_load_all()
            return
        if not self.db_connected or not self.conn:
            messagebox.showwarning("未连接", "数据库未连接")
            return
        conn = self._get_connection()
        try:
            cols = "INVOICECODE, NN, CONCODE, DRIVER, SIZE, DRIVERCODE, TAKENO, DRIVERCOMM"
            sql = f"SELECT {cols} FROM CON_DETAIL WHERE INVOICECODE LIKE %s ORDER BY INVOICECODE"
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                cur.execute(sql, ("%" + keyword + "%",))
                rows = cur.fetchall()
            self._delete_populate_tree(rows)
            self._delete_count_label.config(text=f"共 {len(rows)} 条记录")
        except Exception as e:
            messagebox.showerror("搜索失败", str(e))

    def _delete_load_all(self):
        if not self.db_connected or not self.conn:
            messagebox.showwarning("未连接", "数据库未连接")
            return
        conn = self._get_connection()
        try:
            cols = "INVOICECODE, NN, CONCODE, DRIVER, SIZE, DRIVERCODE, TAKENO, DRIVERCOMM"
            sql = f"SELECT {cols} FROM CON_DETAIL ORDER BY INVOICECODE LIMIT 500"
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                cur.execute(sql)
                rows = cur.fetchall()
            self._delete_populate_tree(rows)
            self._delete_count_label.config(text=f"共 {len(rows)} 条记录")
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _delete_clear(self):
        self._delete_search_var.set("")
        for item in self._delete_tree.get_children():
            self._delete_tree.delete(item)
        self._delete_count_label.config(text="共 0 条记录")

    def _delete_populate_tree(self, rows):
        tree = self._delete_tree
        for item in tree.get_children():
            tree.delete(item)
        for row in rows:
            tree.insert("", "end", values=(
                row.get("INVOICECODE", ""),
                row.get("NN", ""),
                row.get("CONCODE", ""),
                row.get("DRIVER", ""),
                row.get("SIZE", ""),
                row.get("DRIVERCODE", ""),
                row.get("TAKENO", ""),
                row.get("DRIVERCOMM", ""),
            ))

    def _delete_record_from_tree(self):
        tree = self._delete_tree
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选中一条记录")
            return
        values = tree.item(selected[0], "values")
        nn_val = values[1]
        try:
            nn = int(nn_val)
        except (ValueError, TypeError):
            nn = 0
        if nn < 2:
            messagebox.showwarning("提示", f"只能删除 NN >= 2 的记录，当前 NN={nn}")
            return

        invoice_code = values[0]
        con_code = values[2]
        dialog = DeleteConfirmDialog(self.root, invoice_code, con_code)
        self.root.wait_window(dialog)
        if not dialog.result:
            return

        if not self.db_connected or not self.conn:
            messagebox.showwarning("未连接", "数据库未连接")
            return
        conn = self._get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM CON_DETAIL WHERE INVOICECODE = %s AND CONCODE = %s AND NN >= 2",
                    (invoice_code, con_code)
                )
                affected = cur.rowcount
                conn.commit()
            messagebox.showinfo("删除成功", f"已删除 {affected} 条记录\n發票號: {invoice_code}\n櫃號: {con_code}")
            self._delete_do_search()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("删除失败", str(e))

    def _delete_on_double_click(self, event):
        tree = self._delete_tree
        selected = tree.selection()
        if not selected:
            return
        values = tree.item(selected[0], "values")
        cols = tree["columns"]
        info_lines = [f"{col}: {values[i]}" for i, col in enumerate(cols)]
        messagebox.showinfo("记录详情", "\n".join(info_lines))

    # ── 增加司机 ──────────────────────────────────────

    def show_add_driver(self):
        """显示增加司机页面 —— 写入 DRIVER_MASTER / VEHICLE_MASTER / DRIVER_CP"""
        page = tk.Frame(self.main_content, bg='#f0f2f5')
        page.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        self._add_driver_edit_code = None

        # ── 页面标题 ──────────────────────────────────
        title_frame = tk.Frame(page, bg='#f0f2f5')
        title_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            title_frame,
            text="增加司机",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg='#f0f2f5',
            fg='#262626'
        ).pack(side=tk.LEFT)

        # ── 表单卡片 ──────────────────────────────────
        form_card = tk.Frame(page, bg='#ffffff', bd=1, relief='solid')
        form_card.pack(fill=tk.X, ipady=10)

        card_header = tk.Frame(form_card, bg='#ffffff', padx=20, pady=15)
        card_header.pack(fill=tk.X)

        self._add_driver_mode = tk.Label(
            card_header, text="📝 新增司机",
            font=('Microsoft YaHei UI', 14, 'bold'),
            bg='#ffffff', fg='#1890ff'
        )
        self._add_driver_mode.pack(side=tk.LEFT)

        form_body = tk.Frame(form_card, bg='#ffffff', padx=20)
        form_body.pack(fill=tk.X)

        fields = [
            ("司機編號", "driver_code", 20),
            ("司機姓名", "driver_name", 20),
            ("香港車牌", "hk_plate", 20),
            ("大陸車牌", "sz_plate", 20),
        ]
        self._add_driver_entries = {}
        for i, (label, key, width) in enumerate(fields):
            row = tk.Frame(form_body, bg='#ffffff')
            row.pack(fill=tk.X, pady=6)
            tk.Label(row, text=label + "：", font=('Microsoft YaHei UI', 11),
                     bg='#ffffff', fg='#595959', width=10, anchor='e'
                     ).pack(side=tk.LEFT, padx=(0, 10))
            entry = tk.Entry(row, font=('Microsoft YaHei UI', 11), width=width)
            entry.pack(side=tk.LEFT)
            self._add_driver_entries[key] = entry

        tk.Label(form_body, text="提交后同时写入 DRIVER_MASTER / VEHICLE_MASTER / DRIVER_CP 三张表",
                 font=('Microsoft YaHei UI', 9), bg='#ffffff', fg='#8c8c8c'
                 ).pack(anchor="w", pady=(10, 0))

        btn_frame = tk.Frame(form_card, bg='#ffffff', padx=20)
        btn_frame.pack(fill=tk.X)

        tk.Button(
            btn_frame, text="提交", font=('Microsoft YaHei UI', 11),
            bg='#1890ff', fg='#ffffff', bd=0, padx=30, pady=6, cursor='hand2',
            command=self._add_driver_submit
        ).pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(
            btn_frame, text="清空", font=('Microsoft YaHei UI', 11),
            bg='#f5f5f5', fg='#595959', bd=0, padx=20, pady=6, cursor='hand2',
            command=self._add_driver_clear
        ).pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(
            btn_frame, text="取消编辑", font=('Microsoft YaHei UI', 11),
            bg='#f5f5f5', fg='#595959', bd=0, padx=20, pady=6, cursor='hand2',
            command=self._add_driver_cancel_edit
        ).pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(
            btn_frame, text="删除司机", font=('Microsoft YaHei UI', 11),
            bg='#ff4d4f', fg='#ffffff', bd=0, padx=20, pady=6, cursor='hand2',
            command=self._add_driver_delete
        ).pack(side=tk.LEFT)

        # ── 结果提示区域 ──────────────────────────────────
        self._add_driver_result = tk.Label(
            page, text="", font=('Microsoft YaHei UI', 11),
            bg='#f0f2f5', fg='#262626'
        )
        self._add_driver_result.pack(pady=(15, 0))

        # ── 司机列表卡片 ──────────────────────────────────
        list_card = tk.Frame(page, bg='#ffffff', bd=1, relief='solid')
        list_card.pack(fill=tk.BOTH, expand=True, pady=(15, 0), ipady=10)

        list_header = tk.Frame(list_card, bg='#ffffff', padx=20, pady=15)
        list_header.pack(fill=tk.X)

        tk.Label(list_header, text="📋 司机信息列表（点击行可编辑）",
                 font=('Microsoft YaHei UI', 14, 'bold'),
                 bg='#ffffff', fg='#262626').pack(side=tk.LEFT)

        self._add_driver_list_count = tk.Label(
            list_header, text="共 0 条",
            font=('Microsoft YaHei UI', 11), bg='#ffffff', fg='#8c8c8c'
        )
        self._add_driver_list_count.pack(side=tk.RIGHT)

        list_container = tk.Frame(list_card, bg='#ffffff')
        list_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))

        list_cols = ("CUSTCODE", "NAME", "HKCP", "SZCP")
        list_labels = {"CUSTCODE": "司機編號", "NAME": "司機姓名", "HKCP": "香港車牌", "SZCP": "大陸車牌"}
        list_widths = {"CUSTCODE": 120, "NAME": 120, "HKCP": 120, "SZCP": 120}

        self._add_driver_tree = ttk.Treeview(
            list_container, columns=list_cols, show="headings",
            selectmode="browse", height=8
        )
        for col in list_cols:
            self._add_driver_tree.heading(col, text=list_labels[col])
            self._add_driver_tree.column(col, width=list_widths[col], anchor="center", minwidth=80)

        vsb = ttk.Scrollbar(list_container, orient="vertical", command=self._add_driver_tree.yview)
        self._add_driver_tree.configure(yscrollcommand=vsb.set)

        self._add_driver_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        list_container.grid_rowconfigure(0, weight=1)
        list_container.grid_columnconfigure(0, weight=1)

        self._add_driver_tree.bind("<ButtonRelease-1>", self._add_driver_on_select)

        # 加载列表
        self._add_driver_load_list()

    def _add_driver_clear(self):
        for entry in self._add_driver_entries.values():
            entry.delete(0, tk.END)
        self._add_driver_result.config(text="", fg='#262626')
        self._add_driver_mode.config(text="📝 新增司机", fg='#1890ff')
        self._add_driver_edit_code = None

    def _add_driver_submit(self):
        driver_code = self._add_driver_entries["driver_code"].get().strip()
        driver_name = self._add_driver_entries["driver_name"].get().strip()
        hk_plate = self._add_driver_entries["hk_plate"].get().strip()
        sz_plate = self._add_driver_entries["sz_plate"].get().strip()

        if not driver_code:
            messagebox.showwarning("提示", "请输入司機編號")
            return
        if not driver_name:
            messagebox.showwarning("提示", "请输入司機姓名")
            return
        if not hk_plate:
            messagebox.showwarning("提示", "请输入香港車牌")
            return

        if not self.db_connected or not self.conn:
            messagebox.showwarning("未连接", "数据库未连接")
            return

        conn = self._get_connection()
        try:
            # 新增模式：检查司機編號是否重复
            if not getattr(self, '_add_driver_edit_code', None):
                with conn.cursor(pymysql.cursors.DictCursor) as cur:
                    cur.execute("SELECT CUSTCODE, NAME, HKCP, SZCP FROM DRIVER_CP WHERE CUSTCODE = %s", (driver_code,))
                    existing = cur.fetchone()
                if existing:
                    self._add_driver_entries["driver_code"].delete(0, tk.END)
                    self._add_driver_entries["driver_code"].insert(0, existing.get("CUSTCODE", ""))
                    self._add_driver_entries["driver_name"].delete(0, tk.END)
                    self._add_driver_entries["driver_name"].insert(0, existing.get("NAME", ""))
                    self._add_driver_entries["hk_plate"].delete(0, tk.END)
                    self._add_driver_entries["hk_plate"].insert(0, existing.get("HKCP", ""))
                    self._add_driver_entries["sz_plate"].delete(0, tk.END)
                    self._add_driver_entries["sz_plate"].insert(0, existing.get("SZCP", "") or "")
                    self._add_driver_mode.config(text="✏️ 编辑司机", fg='#fa8c16')
                    self._add_driver_edit_code = driver_code
                    self._add_driver_result.config(
                        text="⚠ 司機編號 " + driver_code + " 已存在，已加载其信息，请修改后提交",
                        fg='#fa8c16'
                    )
                    return

            # 香港車牌去重：检查是否与其他司机重复，重复则追加 *
            edit_code = getattr(self, '_add_driver_edit_code', None)
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                while True:
                    if edit_code:
                        cur.execute(
                            "SELECT HKCP FROM DRIVER_CP WHERE HKCP = %s AND CUSTCODE != %s",
                            (hk_plate, edit_code)
                        )
                    else:
                        cur.execute("SELECT HKCP FROM DRIVER_CP WHERE HKCP = %s", (hk_plate,))
                    if not cur.fetchone():
                        break
                    hk_plate = hk_plate + "*"
            self._add_driver_entries["hk_plate"].delete(0, tk.END)
            self._add_driver_entries["hk_plate"].insert(0, hk_plate)

            # 事务写入三表
            sql1 = "INSERT INTO DRIVER_MASTER (DRIVERCODE, NAME) VALUES (%s, %s) ON DUPLICATE KEY UPDATE NAME = VALUES(NAME)"
            sql2 = "INSERT INTO VEHICLE_MASTER (VEHICLECODE, DRIVER) VALUES (%s, %s) ON DUPLICATE KEY UPDATE DRIVER = VALUES(DRIVER)"
            sql3 = "INSERT INTO DRIVER_CP (CUSTCODE, NAME, HKCP, SZCP) VALUES (%s, %s, %s, %s) ON DUPLICATE KEY UPDATE NAME = VALUES(NAME), HKCP = VALUES(HKCP), SZCP = VALUES(SZCP)"

            with conn.cursor() as cur:
                cur.execute(sql1, (driver_code, driver_name))
                cur.execute(sql2, (hk_plate, driver_code))
                cur.execute(sql3, (driver_code, driver_name, hk_plate, sz_plate))
                conn.commit()

            self._add_driver_result.config(
                text="✅ 司机 " + driver_name + "（" + driver_code + "）" + ("更新成功" if getattr(self, '_add_driver_edit_code', None) else "添加成功"),
                fg='#52c41a'
            )
            self._add_driver_clear()
            self._add_driver_load_list()
        except Exception as e:
            conn.rollback()
            self._add_driver_result.config(
                text="❌ " + ("更新" if getattr(self, '_add_driver_edit_code', None) else "添加") + "失败: " + str(e),
                fg='#ff4d4f'
            )

    def _add_driver_load_list(self):
        """加载司机列表"""
        if not self.db_connected or not self.conn:
            return
        conn = self._get_connection()
        try:
            sql = "SELECT CUSTCODE, NAME, HKCP, SZCP FROM DRIVER_CP ORDER BY CUSTCODE"
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                cur.execute(sql)
                rows = cur.fetchall()
            tree = self._add_driver_tree
            for item in tree.get_children():
                tree.delete(item)
            for row in rows:
                tree.insert("", "end", values=(
                    row.get("CUSTCODE", ""),
                    row.get("NAME", ""),
                    row.get("HKCP", ""),
                    row.get("SZCP", ""),
                ))
            self._add_driver_list_count.config(text="共 " + str(len(rows)) + " 条")
        except Exception as e:
            self._add_driver_list_count.config(text="加载失败")

    def _add_driver_on_select(self, event):
        """点击列表行 → 填入表单进入编辑模式"""
        tree = self._add_driver_tree
        selected = tree.selection()
        if not selected:
            return
        values = tree.item(selected[0], "values")
        self._add_driver_entries["driver_code"].delete(0, tk.END)
        self._add_driver_entries["driver_code"].insert(0, values[0])
        self._add_driver_entries["driver_name"].delete(0, tk.END)
        self._add_driver_entries["driver_name"].insert(0, values[1])
        self._add_driver_entries["hk_plate"].delete(0, tk.END)
        self._add_driver_entries["hk_plate"].insert(0, values[2])
        self._add_driver_entries["sz_plate"].delete(0, tk.END)
        self._add_driver_entries["sz_plate"].insert(0, values[3] if values[3] else "")
        self._add_driver_mode.config(text="✏️ 编辑司机", fg='#fa8c16')
        self._add_driver_edit_code = values[0]
        self._add_driver_result.config(text="", fg='#262626')

    def _add_driver_cancel_edit(self):
        """取消编辑，回到新增模式"""
        self._add_driver_clear()
        self._add_driver_tree.selection_remove(self._add_driver_tree.selection())

    def _add_driver_delete(self):
        """删除司机 —— 需要输入司機編號确认，删除三表记录"""
        driver_code = self._add_driver_entries["driver_code"].get().strip()
        if not driver_code:
            messagebox.showwarning("提示", "请先在表单中输入要删除的司機編號，或从列表中点击选择")
            return

        dialog = DriverDeleteConfirmDialog(self.root, driver_code)
        self.root.wait_window(dialog)
        if not dialog.result:
            return

        if not self.db_connected or not self.conn:
            messagebox.showwarning("未连接", "数据库未连接")
            return

        conn = self._get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM DRIVER_MASTER WHERE DRIVERCODE = %s", (driver_code,))
                cur.execute("DELETE FROM VEHICLE_MASTER WHERE DRIVER = %s", (driver_code,))
                cur.execute("DELETE FROM DRIVER_CP WHERE CUSTCODE = %s", (driver_code,))
                conn.commit()

            self._add_driver_result.config(
                text="✅ 司机 " + driver_code + " 已从三表中删除",
                fg='#52c41a'
            )
            self._add_driver_clear()
            self._add_driver_load_list()
        except Exception as e:
            conn.rollback()
            self._add_driver_result.config(
                text="❌ 删除失败: " + str(e),
                fg='#ff4d4f'
            )

    def close(self):
        """关闭程序"""
        self.conn.close()
        self.root.destroy()


def main():
    root = AcrylicRoot()
    app = SWTManagementSystem(root)

    # 关闭按钮调用 app.close（断开数据库）
    app._close_window = app.close

    root.mainloop()


if __name__ == '__main__':
    main()
